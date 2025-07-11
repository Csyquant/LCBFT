import streamlit as st
import requests
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter
import datetime as dt
import xml.etree.ElementTree as ET
import pandas as pd
from rapidfuzz import fuzz, process
import json
import os
import glob
import tempfile
from io import BytesIO
import matplotlib.pyplot as plt
import base64
import platform
if platform.system() == "Windows":
    import pythoncom
    import win32com.client as win32
import numpy as np
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import unicodedata
import re
from collections import Counter
import plotly.graph_objects as go
import plotly.express as px
from openpyxl.styles import Font, PatternFill
import pythoncom
import time
import random
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.io as pio
from openpyxl.styles import Border, Side, Font, Alignment
from PIL import Image
import io
from fuzzywuzzy import process
from openpyxl.utils import get_column_letter # 

# IMPORTANT: set_page_config DOIT être la première commande Streamlit
st.set_page_config(
    page_title="Anti-Money Laundering System",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="collapsed"
)


# Animation d'introduction améliorée avec le style exact du logo
splash_html = """
<style>
    /* Polices */
    @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600&display=swap');
    
    /* Animation d'entrée */
    @keyframes fadeIn {
        0% { opacity: 0; }
        100% { opacity: 1; }
    }
    
    @keyframes slideIn {
        0% { transform: translateX(-30px); opacity: 0; }
        100% { transform: translateX(0); opacity: 1; }
    }
    
    @keyframes logoReveal {
        0% { clip-path: circle(0% at center); opacity: 0; }
        100% { clip-path: circle(100% at center); opacity: 1; }
    }
    
    @keyframes swirl {
        0% { transform: rotate(0deg) scale(0.8); opacity: 0; }
        100% { transform: rotate(360deg) scale(1); opacity: 1; }
    }
    
    @keyframes pulse {
        0% { transform: scale(0.95); }
        50% { transform: scale(1.05); }
        100% { transform: scale(1); }
    }
    
    /* Conteneur principal */
    .intro-container {
        position: fixed;
        top: 0;
        left: 0;
        width: 100vw;
        height: 100vh;
        display: flex;
        justify-content: center;
        align-items: center;
        background-color: #f8f9fa;
        z-index: 9999;
        animation: fadeOut 0.8s ease-in-out 3.5s forwards;
        font-family: 'Montserrat', sans-serif;
    }
    
    @keyframes fadeOut {
        0% { opacity: 1; visibility: visible; }
        100% { opacity: 0; visibility: hidden; }
    }
    
    /* Conteneur du logo */
    .logo-container {
        display: flex;
        align-items: center;
        justify-content: center;
    }
    
    /* Logo avec les spirales */
    .logo-wrapper {
        position: relative;
        width: 220px;
        height: 220px;
        animation: pulse 1.5s ease-in-out;
    }
    
    .logo-spiral {
        position: absolute;
        top: 0;
        left: 0;
        width: 100%;
        height: 100%;
        border: 3px solid #002060;
        border-radius: 50%;
        border-top-color: transparent;
        border-bottom-color: transparent;
        animation: swirl 1.8s cubic-bezier(0.215, 0.61, 0.355, 1) forwards;
    }
    
    .logo-spiral:nth-child(1) {
        transform: rotate(45deg);
        animation-delay: 0.2s;
    }
    
    .logo-spiral:nth-child(2) {
        transform: rotate(-45deg);
        animation-delay: 0.4s;
    }
    
    /* Texte AML */
    .logo-text {
        position: absolute;
        top: 50%;
        left: 50%;
        transform: translate(-50%, -50%);
        color: #002060;
        font-weight: 600;
        font-size: 28px;
        opacity: 0;
        animation: fadeIn 0.8s ease-out 1s forwards;
    }
    
    /* Nom de l'entreprise */
    .company-name-container {
        margin-left: 25px;
        display: flex;
        flex-direction: column;
    }
    
    .company-name {
        font-size: 32px;
        font-weight: 600;
        color: #002060;
        opacity: 0;
        animation: slideIn 0.8s ease-out 1.2s forwards;
        letter-spacing: 0.5px;
    }
    
    .capital {
        font-size: 28px;
        font-weight: 500;
        color: #002060;
        opacity: 0;
        animation: slideIn 0.8s ease-out 1.4s forwards;
    }
    
    .helium {
        font-size: 20px;
        font-weight: 400;
        color: #002060;
        opacity: 0;
        animation: slideIn 0.8s ease-out 1.6s forwards;
        margin-top: 5px;
    }
    
    /* Tagline */
    .tagline {
        position: absolute;
        bottom: 40px;
        font-size: 16px;
        color: #002060;
        letter-spacing: 1.5px;
        opacity: 0;
        animation: fadeIn 0.8s ease-out 2s forwards;
        font-weight: 500;
    }
</style>

<div class="intro-container">
    <div class="logo-container">
        <div class="logo-wrapper">
            <div class="logo-spiral"></div>
            <div class="logo-spiral"></div>
            <div class="logo-text">AML</div>
        </div>
        <div class="company-name-container">
            <div class="company-name">SYQUANT</div>
            <div class="capital">Capital</div>
            <div class="helium">Helium Funds</div>
        </div>
    </div>
    <div class="tagline">ANTI-MONEY LAUNDERING SYSTEM</div>
</div>

<script>
    // Script pour afficher les éléments Streamlit après l'animation
    setTimeout(function() {
        // Réafficher les éléments Streamlit
        document.querySelectorAll('header, footer, .sidebar-content, .block-container').forEach(function(el) {
            el.style.opacity = '1';
        });
    }, 6500);
</script>
"""
# Injecter l'animation APRÈS set_page_config
st.markdown(splash_html, unsafe_allow_html=True)

# Configuration des URLs par défaut
DEFAULT_URLS = {
    'SDN List': 'https://www.treasury.gov/ofac/downloads/sdn.xml',
    'Non-SDN List': 'https://www.treasury.gov/ofac/downloads/consolidated/consolidated.xml',
    'French List': 'https://gels-avoirs.dgtresor.gouv.fr/ApiPublic/api/v1/publication/derniere-publication-fichier-json',
    'IOSCO List': 'https://www.iosco.org/i-scan/?export-to-csv&VALIDATIONDATEEND=&ID=&SUBSECTION=main&page=1&CATEGORYID=&NCA_ID=&VALIDATIONDATESTART=&PRODUCTID=&KEYWORDS='
}

def extract_date_from_filename_new_format(filename):
    """Extracts date (YYYYMMDD) from filenames like 'YYYYMMDD - screening...'"""
    match = re.match(r'^(\d{8})\s-', filename)
    if match:
        return match.group(1)
    return None

def normalize_text_for_key(text):
    """
    Normalise un texte pour l'utilisation comme clé, en le rendant insensible à la casse,
    aux accents, aux espaces multiples et aux caractères non-alphanumériques (sauf espaces).
    """
    if not text:
        return ""
    text = str(text).lower().strip()
    # Remove accents
    text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
    # Replace all non-alphanumeric characters (except spaces) with a single space
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    # Replace multiple spaces with a single space
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

# --- New functions for persistence of AMF/CFT levels ---
AMF_CFT_LEVELS_FILE = 'amf_cft_levels.json'

def load_amf_cft_levels():
    """Loads AMF/CFT levels from a JSON file."""
    if os.path.exists(AMF_CFT_LEVELS_FILE):
        try:
            with open(AMF_CFT_LEVELS_FILE, 'r') as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            st.warning(f"Erreur de lecture du fichier AMF/CFT levels: {e}. Les niveaux seront réinitialisés.")
            return {}
    return {}

def save_amf_cft_levels(levels):
    """Saves AMF/CFT levels to a JSON file."""
    try:
        # Convert tuples (for 'Autre') to lists before saving to JSON
        serializable_levels = {}
        for country, level_data in levels.items():
            if isinstance(level_data, tuple):
                serializable_levels[country] = list(level_data)
            else:
                serializable_levels[country] = level_data

        with open(AMF_CFT_LEVELS_FILE, 'w') as f:
            json.dump(serializable_levels, f, indent=4)
    except Exception as e:
        st.error(f"Erreur lors de la sauvegarde des niveaux AMF/CFT: {e}")
        
# Initialisation du session state pour les dossiers
if "download_dir" not in st.session_state:
    st.session_state.download_dir = r"\\panfs001\Syquant Compliance\PG06 - Politique LCB-FT\Screening\1. Listes de sanctions"
if "results_dir" not in st.session_state:
    st.session_state.results_dir = r"\\panfs001\Syquant Compliance\PG06 - Politique LCB-FT\Screening\3. Résultats du screening"

     
# Sidebar avec options
logo_path = r"\\panfs001\Syquant Compliance\PG06 - Politique LCB-FT\Screening\x. codes (old)\Logo.GIF"
try:
    logo = Image.open(logo_path)
    st.sidebar.image(logo, use_container_width=True)
except Exception as e:
    st.sidebar.warning("Logo non trouvé ou chemin incorrect.")

st.sidebar.title("Options")
threshold_nom = st.sidebar.slider("Seuil de correspondance nom (%)",
                                 min_value=70,
                                 max_value=100,
                                 value=85,
                                 step=5,
                                 help="Seuil minimum pour considérer qu'un nom correspond")
threshold_prenom = st.sidebar.slider("Seuil de correspondance prénom (%)", 70, 100, 80,
                                    help="Seuil minimum pour considérer qu'un prénom correspond")
pm_screening_threshold = st.sidebar.slider("Seuil de correspondance PM (%)",
                                         min_value=70,
                                         max_value=100,
                                         value=80,
                                         step=5,
                                         key="pm_screening_threshold", # Store in session state
                                         help="Seuil minimum pour considérer qu'une entité morale correspond")
min_name_length = st.sidebar.slider("Longueur minimale du nom", 2, 5, 3,
                                    help="Ignorer les noms trop courts qui peuvent générer des faux positifs")
require_both_match = st.sidebar.checkbox("Exiger correspondance nom et prénom", value=True,
                                        help="Exiger que le nom et le prénom correspondent pour réduire les faux positifs")
use_intelligent_scoring = st.sidebar.checkbox("Utiliser le scoring intelligent (NLP)", value=True,
                                           help="Utilise des techniques NLP pour réduire les faux positifs")

# Champ pour le dossier de téléchargement des listes
download_dir = st.sidebar.text_input(
    "Dossier de téléchargement des listes",
    key="download_dir",
    help="Chemin du dossier où les listes de sanctions seront téléchargées"
)

# Champ pour le dossier de sauvegarde des résultats
results_dir = st.sidebar.text_input(
    "Dossier de sauvegarde des résultats",
    value=st.session_state.results_dir,
    help="Chemin du dossier où les fichiers de résultats seront enregistrés"
)

# Charger les URLs personnalisées depuis un fichier de configuration
def load_custom_urls():
    urls = DEFAULT_URLS.copy() # Start with default URLs
    try:
        if os.path.exists('sanction_urls.json'):
            with open('sanction_urls.json', 'r') as f:
                custom_urls = json.load(f)
                # Update defaults with custom URLs, potentially overwriting defaults
                urls.update(custom_urls)
    except Exception as e:
        # If there's an error loading or parsing the custom file, log a warning
        st.warning(f"Erreur lors du chargement ou de l'analyse des URLs personnalisées: {e}. Proceeding with available URLs.")


    for key, value in DEFAULT_URLS.items():
        if key not in urls:
            urls[key] = value

    return urls

# Sauvegarder les URLs personnalisées
def save_custom_urls(urls):
    try:
        with open('sanction_urls.json', 'w') as f:
            json.dump(urls, f, indent=4)
    except Exception as e:
        st.error(f"Erreur lors de la sauvegarde des URLs: {e}")

# Create a simple name similarity function for companies (defined globally)
def calculate_company_similarity(name1, name2):
    """
    Calculate similarity between two company names using token_set_ratio.
    """
    if not name1 or not name2:
        return 0
    name1 = str(name1).lower().strip()
    name2 = str(name2).lower().strip()
    # Using token_set_ratio as it handles word order and missing words well for names
    return fuzz.token_set_ratio(name1, name2)

# --- New Sidebar Configuration Section (Dropdown) ---
with st.sidebar.expander("⚙️ Configuration des sources de données"):
    st.subheader("Sources de données")
    urls = load_custom_urls()

    with st.form("url_config_form_sidebar"):
        # Use .get() with DEFAULT_URLS as fallback for robustness
        sdn_url = st.text_input("URL de la liste SDN", value=urls.get('SDN List', DEFAULT_URLS.get('SDN List', '')))
        nonsdn_url = st.text_input("URL de la liste Non-SDN", value=urls.get('Non-SDN List', DEFAULT_URLS.get('Non-SDN List', '')))
        fr_url = st.text_input("URL de la liste française", value=urls.get('French List', DEFAULT_URLS.get('French List', '')))
        iosco_url = st.text_input("URL de la liste OICV", value=urls.get('IOSCO List', DEFAULT_URLS.get('IOSCO List', ''))) # Added IOSCO URL input

        if st.form_submit_button("💾 Sauvegarder les URLs"):
            new_urls = {
                'SDN List': sdn_url,
                'Non-SDN List': nonsdn_url,
                'French List': fr_url,
                'OICV List': iosco_url, # Save IOSCO URL
            }
            save_custom_urls(new_urls)
            st.success("✅ URLs sauvegardées avec succès!")

    st.markdown("___") # Separator

    st.subheader("Téléchargement des listes")
    if st.button("🔄 Vérifier et télécharger les listes (Sidebar)", key="download_button_sidebar"):
        with st.spinner("Téléchargement en cours..."):
            success = download_sanction_lists()
            if success:
                st.success("✅ Listes téléchargées avec succès")
            else:
                st.error("❌ Échec lors du téléchargement")

# Variables globales (après la définition de download_dir)
today = dt.datetime.today().strftime('%Y%m%d')
file_sdn = os.path.join(st.session_state.download_dir, f'{today}_liste_sdn_ofac.xml')
file_nonsdn = os.path.join(st.session_state.download_dir, f'{today}_liste_nonsdn_ofac.xml')
file_fr = os.path.join(st.session_state.download_dir, f'{today}_liste_gel_des_avoirs.json')
file_iosco = os.path.join(st.session_state.download_dir, f'{today}_liste_iosco.csv')

# Fonctions de base
def file_exists(file_path):
    return os.path.isfile(file_path)

import time

def delete_file_with_retries(filepath, retries=5, delay=0.5):
    """
    Attempts to delete a file with retries in case of FileInUse error (WinError 32).
    """
    for i in range(retries):
        try:
            os.remove(filepath)
            return True
        except PermissionError as e: # Catch specifically PermissionError which WinError 32 is
            if "WinError 32" in str(e):
                time.sleep(delay)
            else:
                return False
        except Exception as e:
            return False
    return False

def get_session_with_retries():
    session = requests.Session()
    retries = Retry(
        total=5,
        backoff_factor=2,
        status_forcelist=[500, 502, 503, 504],
        allowed_methods=["GET"]
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount('https://', adapter)
    session.mount('http://', adapter)
    return session

def download_file(url, filename):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) ',
        'Accept': '*/*',
        'Accept-Language': 'fr-FR,fr;q=0.9',
        'Referer': url,
        'Connection': 'keep-alive',
    }

    session = get_session_with_retries()
    session.headers.update(headers)

    try:
        # Délai aléatoire entre 1 et 3 secondes
        time.sleep(random.uniform(1.0, 3.0))

        with session.get(url, stream=True, timeout=90) as response:
            response.raise_for_status()
            with open(filename, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
        print(f"✅ Téléchargement terminé : {filename}")
        return True
    except requests.exceptions.RequestException as e:
        print(f"❌ Erreur lors du téléchargement de {filename} : {e}")
        return False

# Exemple d'utilisation
if __name__ == "__main__":
    url = 'https://home.treasury.gov/system/files/126/sdn.xml'  # lien réel à ajuster si besoin
    filename = '20250515_liste_sdn_ofac.xml'


@st.cache_data(ttl=24*60*60)  # Cache pendant 24h
def download_sanction_lists():
    if not st.session_state.download_dir:
        st.error("Veuillez renseigner le dossier de téléchargement des listes dans la barre latérale.")
        st.stop()
    urls = load_custom_urls()
    files_to_check = {
        'SDN List': (file_sdn, urls['SDN List']),
        'Non-SDN List': (file_nonsdn, urls['Non-SDN List']),
        'French List': (file_fr, urls['French List']),
        'IOSCO List': (file_iosco, urls['IOSCO List'])
    }
    progress_bar = st.progress(0)
    status_text = st.empty()
    total_files = len(files_to_check)
    download_success = True
    for i, (name, (filename, url)) in enumerate(files_to_check.items()):
        status_text.text(f"Vérification de {name}...")
        if file_exists(filename):
            status_text.text(f'{filename} existe déjà.')
        else:
            status_text.text(f'Téléchargement de {filename}...')
            if not download_file(url, filename):
                st.error(f'Erreur lors du téléchargement de {filename}. Veuillez réessayer.')
                download_success = False
                # Continue to check/download other files but mark as overall failure
        progress_bar.progress((i + 1) / total_files)

    if download_success:
        status_text.text('Toutes les listes de sanctions sont prêtes!')
    else:
         status_text.text('Certaines listes n\'ont pas pu être téléchargées.')

    return download_success # Return the overall success status

# Vérification et téléchargement automatique
def check_and_download_sanctions():
    # Check if all expected files exist
    required_files = [file_sdn, file_nonsdn, file_fr, file_iosco] # Include IOSCO file
    all_files_exist = all(file_exists(f) for f in required_files)

    if not all_files_exist:
        with st.spinner("Téléchargement automatique des listes de sanctions..."):
            success = download_sanction_lists()
            if success:
                st.success("Listes de sanctions téléchargées avec succès!")
            else:
                # An error message was already shown inside download_sanction_lists
                pass # Keep the existing error message
    else:
         # Optionally inform user that files are already present
         pass # Keep it silent if files exist

# Fonctions de parsing optimisées
@st.cache_data
def parse_sdn(file):
    try:
        tree = ET.parse(file)
        root = tree.getroot()
        people = []
        for e in root:
            p = {"Nom": "", "Prénom": "", "AKA_Nom": [], "AKA_Prénom": [], "Type": "", "Source": "SDN"}
            for a in e:
                if a.tag.endswith("lastName"):
                    p["Nom"] = a.text if a.text else ""
                if a.tag.endswith("firstName"):
                    p["Prénom"] = a.text if a.text else ""
                if a.tag.endswith("sdnType"):
                    p["Type"] = a.text if a.text else ""
                if a.tag.endswith("akaList"):
                    for b in a:
                        for c in b:
                            if c.tag.endswith("lastName") and c.text:
                                p["AKA_Nom"].append(c.text)
                            if c.tag.endswith("firstName") and c.text:
                                p["AKA_Prénom"].append(c.text)
            # Filtre pour ne garder que les personnes physiques
            if p["Type"] == "Individual":
                # Filtre pour ignorer les noms trop courts (pour réduire les faux positifs)
                if len(p["Nom"]) >= min_name_length:
                    people.append(p)
        return people
    except Exception as e:
        st.error(f"Error parsing {file}: {e}")
        return []

def parse_nonsdn(file):
    """Parses the Non-SDN XML file and extracts/cleans individual names."""
    if not os.path.exists(file):
        st.warning(f"Non-SDN file not found: {file}. Skipping parsing.")
        return []

    try:
        tree = ET.parse(file)
        root = tree.getroot()
        people = []
        # Assuming Non-SDN XML has a similar structure to SDN for individuals
        for entity in root.findall('.//d:entry', {'d': 'http://www.un.org/sanctions/1.0'}): # Adjust namespace if needed
             entity_type = entity.find('.//d:EntityType', {'d': 'http://www.un.org/sanctions/1.0'}) # Adjust tag/namespace
             if entity_type is not None and entity_type.text == "Individual": # Check if it's an individual
                person = {"Nom": "", "Prénom": "", "AKA_Nom": [], "AKA_Prénom": [], "Type": "Individual", "Source": "Non-SDN"}
                
                # Extract names - Adjust tags/namespaces as per Non-SDN XML structure
                # Assuming similar tags for names as SDN for now, adjust if necessary
                last_name_element = entity.find('.//d:LastName', {'d': 'http://www.un.org/sanctions/1.0'})
                first_name_element = entity.find('.//d:FirstName', {'d': 'http://www.un.org/sanctions/1.0'})

                if last_name_element is not None and last_name_element.text:
                     person["Nom"] = last_name_element.text
                if first_name_element is not None and first_name_element.text:
                     person["Prénom"] = first_name_element.text

                # Extract AKAs (aliases) - Adjust tags/namespaces as per Non-SDN XML structure
                # Assuming similar structure for aliases as SDN for now
                aka_elements = entity.findall('.//d:Alias', {'d': 'http://www.un.org/sanctions/1.0'}) # Assuming <Alias> tag for AKAs
                for aka in aka_elements:
                     aka_last_name = aka.find('.//d:LastName', {'d': 'http://www.un.org/sanctions/1.0'})
                     aka_first_name = aka.find('.//d:FirstName', {'d': 'http://www.un.org/sanctions/1.0'})
                     if aka_last_name is not None and aka_last_name.text:
                         person["AKA_Nom"].append(aka_last_name.text)
                     if aka_first_name is not None and aka_first_name.text:
                         person["AKA_Prénom"].append(aka_first_name.text)

                # Filter for minimum name length
                if len(person["Nom"]) >= min_name_length:
                    people.append(person)
        

        return people
    except Exception as e:
        st.error(f"Error parsing {file}: {e}")
        return []

@st.cache_data
def parse_fr(file):
    try:
        with open(file, encoding="utf8") as f:
            data = json.load(f)
        result = []
        for e in data['Publications']['PublicationDetail']:
            if e['Nature'] == "Personne physique":
                d = {"Nom": e['Nom'], "Prénom": "", "AKA_Nom": [], "AKA_Prénom": [], "Source": "FR"}
                for item in e['RegistreDetail']:
                    if item['TypeChamp'] == 'PRENOM':
                        d["Prénom"] = item['Valeur'][0]['Prenom']
                    elif item['TypeChamp'] == 'ALIAS':
                        for alias_item in item['Valeur']:
                            alias = alias_item['Alias']
                            # Essayer de séparer l'alias en prénom/nom si possible
                            parts = alias.split()
                            if len(parts) > 1:
                                d["AKA_Nom"].append(" ".join(parts[1:]))
                                d["AKA_Prénom"].append(parts[0])
                            else:
                                d["AKA_Nom"].append(alias)
                # Filtre pour ignorer les noms trop courts
                if len(d["Nom"]) >= min_name_length:
                    result.append(d)
        return result
    except Exception as e:
        st.error(f"Erreur lors du parsing de {file}: {e}")
        return []

# Fonction pour séparer un nom complet
def split_full_name(full_name):
    if pd.isna(full_name) or full_name == "":
        return "", ""
    parts = str(full_name).strip().split()
    if len(parts) == 1:
        return "", parts[0]
    elif len(parts) == 2:
        return parts[0], parts[1]
    else:
        return parts[0], " ".join(parts[1:])

# Fonctions NLP pour l'analyse intelligente des noms
def normalize_name(name):
    if not name:
        return ""
    # Convertir en minuscules
    name = name.lower()
    # Supprimer les accents
    name = ''.join(c for c in unicodedata.normalize('NFD', name) if unicodedata.category(c) != 'Mn')
    # Supprimer les caractères spéciaux
    name = re.sub(r'[^a-z0-9\s]', '', name)
    return name.strip()

def detect_name_origin(name):
    """Détecte l'origine probable d'un nom basé sur des caractéristiques linguistiques"""
    if not name:
        return "unknown"
    
    name = normalize_name(name)
    
    # Caractéristiques linguistiques simplifiées
    # Ces règles sont très basiques et pourraient être améliorées
    
    # Noms asiatiques (chinois, japonais, coréens)
    asian_chars = set("li zhang wang chen yang liu huang zhao wu zhou xu sun ma zhu ye gao lin he luo tang")
    if len(name) <= 5 and name in asian_chars:
        return "asian"
    
    # Noms arabes/moyen-orientaux
    arabic_patterns = ["al-", "el-", "abdul", "mohammed", "ahmad", "ali", "hassan", "hussein"]
    if any(pattern in name for pattern in arabic_patterns):
        return "arabic"
    
    # Noms slaves
    slavic_endings = ["ov", "ev", "in", "ski", "sky", "ich", "vich", "enko"]
    if any(name.endswith(ending) for ending in slavic_endings):
        return "slavic"
    
    # Noms latins/européens
    latin_patterns = ["de ", "di ", "van ", "von ", "le ", "la "]
    if any(pattern in name for pattern in latin_patterns):
        return "european"
    
    # Par défaut, on suppose européen/occidental
    return "western"

def extract_name_features(first_name, last_name):
    features = {}
    
    # Normaliser les noms
    first_name_norm = normalize_name(first_name)
    last_name_norm = normalize_name(last_name)
    
    # Caractéristiques basiques
    features["first_name_length"] = len(first_name_norm) if first_name_norm else 0
    features["last_name_length"] = len(last_name_norm) if last_name_norm else 0
    
    # Détection d'origine
    features["first_name_origin"] = detect_name_origin(first_name_norm)
    features["last_name_origin"] = detect_name_origin(last_name_norm)
    
    # Caractéristiques de structure
    features["first_name_words"] = len(first_name_norm.split()) if first_name_norm else 0
    features["last_name_words"] = len(last_name_norm.split()) if last_name_norm else 0
    
    # Distribution des caractères (indicateur de l'origine du nom)
    if first_name_norm:
        char_counts = Counter(first_name_norm)
        features["first_name_vowel_ratio"] = sum(char_counts.get(c, 0) for c in 'aeiou') / len(first_name_norm) if first_name_norm else 0
    
    if last_name_norm:
        char_counts = Counter(last_name_norm)
        features["last_name_vowel_ratio"] = sum(char_counts.get(c, 0) for c in 'aeiou') / len(last_name_norm) if last_name_norm else 0
    
    return features

def calculate_cultural_similarity(person1_first, person1_last, person2_first, person2_last):
    """Calcule la similarité culturelle/linguistique entre deux noms"""
    # Extraire les caractéristiques
    features1 = extract_name_features(person1_first, person1_last)
    features2 = extract_name_features(person2_first, person2_last)
    
    # Vérifier si les origines sont compatibles
    origin_match = (features1["first_name_origin"] == features2["first_name_origin"] or 
                   features1["last_name_origin"] == features2["last_name_origin"])
    
    # Comparer les distributions de caractères (indicateur de l'origine ethnique)
    vowel_diff = abs(features1.get("first_name_vowel_ratio", 0) - features2.get("first_name_vowel_ratio", 0)) + \
                 abs(features1.get("last_name_vowel_ratio", 0) - features2.get("last_name_vowel_ratio", 0))
    
    # Score de similarité culturelle (0-100)
    cultural_score = 100
    
    # Pénalité pour différence d'origine
    if not origin_match:
        cultural_score -= 40
    
    # Pénalité pour différence de distribution de caractères
    cultural_score -= vowel_diff * 50
    
    return max(0, cultural_score)

def character_ngram_similarity(name1, name2, n=2):
    """Calcule la similarité basée sur les n-grammes de caractères"""
    if not name1 or not name2:
        return 0
    
    # Normaliser les noms
    name1 = normalize_name(name1)
    name2 = normalize_name(name2)
    
    # Créer des n-grammes de caractères
    vectorizer = CountVectorizer(analyzer='char', ngram_range=(n, n))
    try:
        X = vectorizer.fit_transform([name1, name2])
        # Calculer la similarité cosinus
        similarity = cosine_similarity(X)[0, 1]
        return similarity * 100  # Convertir en pourcentage
    except:
        return 0

def calculate_name_similarity(name1, name2):
    """Calcule la similarité entre deux noms en utilisant plusieurs méthodes."""
    if not name1 or not name2:
        return 0
    
    name1 = name1.lower().strip()
    name2 = name2.lower().strip()
    
    # Ignorer les noms trop courts pour réduire les faux positifs
    if len(name1) < min_name_length or len(name2) < min_name_length:
        return 0
    
    # Calculer plusieurs scores et prendre le meilleur
    ratio = fuzz.ratio(name1, name2)
    partial = fuzz.partial_ratio(name1, name2)
    token_set = fuzz.token_set_ratio(name1, name2)
    
    return max(ratio, partial, token_set)

# Fonction pour dédupliquer les personnes
def deduplicate_people(people_df):
    """
    Déduplique les personnes en fonction de leur nom et prénom.
    Si une pers apparaît plusieurs fois avec des rôles différents, on garde une seule entrée
    avec tous les rôles concaténés.
    """
    if people_df.empty:
        return people_df
    
    # Créer une clé unique pour chaque personne (nom + prénom)
    people_df['person_key'] = people_df['Last Name'].str.lower() + '|' + people_df['First Name'].str.lower()
    
    # Identifier les doublons
    duplicates = people_df.duplicated(subset=['person_key'], keep=False)
    
    if not duplicates.any():
        # Pas de doublons, retourner le DataFrame original sans la colonne person_key
        return people_df.drop(columns=['person_key'])
    
    # Créer un nouveau DataFrame pour les résultats dédupliqués
    unique_people = []
    
    # Grouper par clé de personne
    for key, group in people_df.groupby('person_key'):
        # Prendre la première entrée comme base
        person = group.iloc[0].to_dict()
        
        # Si plusieurs types, les concaténer
        if len(group) > 1:
            types = group['Type'].unique()
            person['Type'] = ' & '.join(types)
            
            # Concaténer les sources si différentes
            sources = group['Source'].unique()
            person['Source'] = ' & '.join(sources)
        
        # Supprimer la clé temporaire
        del person['person_key']
        
        unique_people.append(person)
    
    # Convertir la liste en DataFrame
    result = pd.DataFrame(unique_people)
    
    return result

# Fonction de screening intelligent
def intelligent_screening(person, sanctions, threshold_nom, threshold_prenom):
    first_name = person["First Name"] if not pd.isna(person["First Name"]) else ""
    last_name = person["Last Name"] if not pd.isna(person["Last Name"]) else ""
    
    # Ignorer si le nom est trop court
    if len(last_name) < min_name_length:
        return None
    
    best_match = None
    best_score_total = 0
    
    for sanction in sanctions:
        sanction_nom = sanction.get('Nom', '')
        sanction_prenom = sanction.get('Prénom', '')
        
        # Calcul de similarité traditionnelle
        nom_score = calculate_name_similarity(last_name, sanction_nom)
        prenom_score = calculate_name_similarity(first_name, sanction_prenom) if first_name and sanction_prenom else 0
        
        # Si les scores de base sont trop bas, passer à la sanction suivante
        if nom_score < threshold_nom * 0.7 or (require_both_match and first_name and prenom_score < threshold_prenom * 0.7):
            continue
        
        # Calcul de similarité culturelle/linguistique
        cultural_score = calculate_cultural_similarity(
            first_name, last_name, 
            sanction_prenom, sanction_nom
        )
        
        # Calcul de similarité n-gramme pour détecter les variations orthographiques
        ngram_nom_score = character_ngram_similarity(last_name, sanction_nom)
        ngram_prenom_score = character_ngram_similarity(first_name, sanction_prenom)
        
        # Calcul du score final avec pondération
        # - Score de similarité traditionnelle: 40%
        # - Score de similarité culturelle: 40% 
        # - Score de similarité n-gramme: 20%
        traditional_score = (nom_score * 0.6) + (prenom_score * 0.4) if first_name else nom_score
        ngram_score = (ngram_nom_score * 0.6) + (ngram_prenom_score * 0.4) if first_name else ngram_nom_score
        
        final_score = (traditional_score * 0.4) + (cultural_score * 0.4) + (ngram_score * 0.2)
        
        # Pénalité pour les incompatibilités évidentes
        if first_name and sanction_prenom:
            first_letter_match = first_name[0].lower() == sanction_prenom[0].lower()
            origin_match = (detect_name_origin(first_name) == detect_name_origin(sanction_prenom))
            
            if not first_letter_match:
                final_score *= 0.7  # Pénalité de 30%
            
            if not origin_match:
                final_score *= 0.8  # Pénalité de 20%
        
        # Mettre à jour le meilleur match si nécessaire
        if final_score > best_score_total and final_score >= threshold_nom:
            best_score_total = final_score
            best_match = {
                'Nom': sanction.get('Nom', ''),
                'Prénom': sanction.get('Prénom', ''),
                'Source': sanction.get('Source', ''),
                'Score Nom': nom_score,
                'Score Prénom': prenom_score,
                'Score Cultural': cultural_score,
                'Score Total': round(final_score, 1),
                'Détails': f"Nom: {nom_score:.1f}%, Prénom: {prenom_score:.1f}%, Culturel: {cultural_score:.1f}%"
            }
    
    return best_match

# Fonction pour parser la liste IOSCO
@st.cache_data
@st.cache_data
def parse_iosco_pm_variants(file_path):
    """
    Parse la liste IOSCO pour extraire les noms commerciaux, alias, noms légaux (PM),
    ainsi que la juridiction et le type de société.
    Retourne une liste de dictionnaires contenant les noms nettoyés, la source, la juridiction et le type.
    """
    try:
        # Colonnes à lire : C (2), D (3), G (6), H (7), I (8)
        df = pd.read_csv(file_path, usecols=[2, 3, 6, 7, 8], encoding='utf-8', dtype=str)
        df.columns = ['nca_name', 'nca_jurisdiction', 'commercial_name', 'other_commercial_names', 'corporate_names']
        df.fillna("", inplace=True)

        iosco_entities = []
        for _, row in df.iterrows():
            names = set()
            main_name = ""
            other_names_list = []

            # Collect all names from the row and clean them
            all_row_names = []
            for col in ['commercial_name', 'other_commercial_names', 'corporate_names']:
                cleaned = re.split(r'[;,]', row.get(col, ''))
                all_row_names.extend([p.strip().lower() for p in cleaned if p.strip()])

            # Use the first non-empty cleaned name as the potential main name
            for name in all_row_names:
                if name and not main_name:
                    main_name = name
                elif name:
                    other_names_list.append(name)

            # Ensure main_name is unique among other_names and remove duplicates from other_names
            if main_name and main_name in other_names_list:
                other_names_list.remove(main_name)

            other_names_unique = sorted(list(set(other_names_list)))

            if main_name: # Only add if there's at least one name
                iosco_entities.append({
                    "Nom": main_name,
                    "Autres Noms": other_names_unique,
                    "Source": "IOSCO",
                    "nca_jurisdiction": row.get("nca_jurisdiction", ""),
                    "nca_name": row.get("nca_name", "")
                })

        st.success(f"{len(iosco_entities)} entites sous sanctions trouvées(IOSCO)")
        return iosco_entities
    except FileNotFoundError:
        st.error(f"Error parsing file: The file was not found at {file_path}")
        return []
    except pd.errors.EmptyDataError:
        st.error(f"Error parsing file: The file at {file_path} is empty.")
        return []
    except Exception as e:
        st.error(f"An unexpected error occurred during parsing: {e}")
        return []

# Fonction pour extraire les noms de compte du fichier Account
def extract_account_names(file):
    """Charge et analyse le Report Account Olley avec infos supplémentaires."""
    company_infos = []
    try:
        # Lire les colonnes D (Account Name), N (Company type), P (Jurisdiction)
        df = pd.read_excel(file, header=8, usecols="D,N,P")
        df.columns = df.columns.str.strip()
        for _, row in df.iterrows():
            account_name = str(row.get("Account Name", "")).strip()
            company_type = str(row.get("Company type", "")).strip()
            jurisdiction = str(row.get("Jurisdiction", "")).strip()
            if account_name:
                company_infos.append({
                    "name": account_name,
                    "source": "Account",
                    "jurisdiction_account": jurisdiction,
                    "company_type_account": company_type
                })
        return company_infos
    except Exception as e:
        st.error(f"Erreur lors de l'extraction des noms de compte du fichier Account : {str(e)}")
        return []

# Fonction pour extraire les noms de compte du fichier Contact
def extract_contact_account_names(file):
    company_infos = []
    try:
        df = pd.read_excel(file, header=8, usecols="H")  # H = Account Name
        df.columns = df.columns.str.strip()
        account_name_col = "Account Name"
        if account_name_col in df.columns:
            names = df[account_name_col].dropna().astype(str).str.strip()
            for name in names[names != ''].tolist():
                company_infos.append({
                    "name": name,
                    "source": "Contact"
                })
        return company_infos
    except Exception as e:
        st.error(f"Erreur lors de l'extraction des noms de compte du fichier Contact : {str(e)}")
        return []

# Fonction pour extraire les noms de société/compte du fichier Leads
def extract_leads_company_names(file):
    company_infos = []
    try:
        df = pd.read_excel(file, header=12, usecols="T")  # T = Company/Account
        df.columns = df.columns.str.strip()
        company_account_col = "Company/Account"
        if company_account_col in df.columns:
            names = df[company_account_col].dropna().astype(str).str.strip()
            for name in names[names != ''].tolist():
                company_infos.append({
                    "name": name,
                    "source": "Leads"
                })
        print(f"✅ {len(company_infos)} noms de société/compte extraits du fichier Leads.")
        return company_infos
    except Exception as e:
        return []

# Function to perform fuzzy matching for legal entities
def screen_legal_entities(crm_companies, iosco_entities, threshold, progress_callback=None):
    """
    Compares CRM company names to IOSCO legal entities using a logic where all words
    from the CRM name must be present in the IOSCO name (primary or alias) AND
    the beginning of the names must be similar.
    Returns a DataFrame with the correspondences.
    """
    matches = []
    screened_crm_companies = set()

    if not iosco_entities:
        st.warning("Liste d'entités légales IOSCO vide, screening ignoré.")
        if progress_callback:
             # Report completion for PM even if skipped, assuming total_pm_analyzed was derived
             # Total to analyze is based on unique CRM companies extracted earlier
             total_crm_companies = len(st.session_state.get('unique_crm_companies', [])) # Use count from state
             progress_callback(total_crm_companies, total_crm_companies, "PM") # Report completion
        return pd.DataFrame(matches)

    # Define common legal suffixes and very generic words to remove or de-emphasize
    common_suffixes = [' sa', ' s.a.', ' ltd', ' limited', ' sarl', ' gmbh', ' ag', ' inc', ' corp', ' llc', ' co']
    generic_words = ['the', 'of', 'and', 'or', 'de', 'des', 'du', 'et', 'a', 'à', 'management', 'asset', 'group', 'holdings', 'holding'] # Added common finance/business terms to generic
    common_suffixes.sort(key=len, reverse=True)

    def tokenize_and_clean_name(name):
        if not name:
            return set() # Return an empty set for no name

        cleaned_name = str(name).lower().strip()

        # --- Remove URL patterns from the beginning and optional parentheses ---
        url_pattern = r'^(?:http[s]?://)?(?:www\\.)?([a-z0-9.-]+(?:/[^/\\s]*)?)?\\s*(?:\([^)]*\))?'
        cleaned_name = re.sub(url_pattern, '', cleaned_name).strip()
        # Also handle cases where URL is part of the name but without common prefixes (e.g., just TLD at the end)
        cleaned_name = re.sub(r'\\.com$|\\.org$|\\.net$|\\.int$', '', cleaned_name).strip() # Remove common TLDs at the end

        # --- Remove cloning-related terms ---
        cleaned_name = re.sub(r'\\bcloning(?: of)?(?: entity)?\\b', '', cleaned_name, flags=re.IGNORECASE).strip()


        # Remove common suffixes
        for suffix in common_suffixes:
            if cleaned_name.endswith(suffix):
                cleaned_name = cleaned_name[:-len(suffix)].strip()

        # Remove punctuation and extra spaces, then split into tokens
        # Keep hyphen for now as it can be part of names, might need refinement
        cleaned_name = re.sub(r'[^a-z0-9\\s-]', '', cleaned_name) # Allow hyphens
        tokens = cleaned_name.split()

        # Remove very generic words and single-character tokens
        tokens = [token for token in tokens if token not in generic_words and len(token) > 1]

        return set(tokens) # Return a set of tokens for efficient subset checking

    # Helper function to get the first few significant tokens
    def get_first_n_tokens(token_set, n=3):
        sorted_tokens = sorted(list(token_set))
        return set(sorted_tokens[:n])

    # Prepare IOSCO entities: tokenize names and aliases
    iosco_token_data = [] 
    for iosco_entity in iosco_entities:
        primary_name = iosco_entity.get('Nom', '')
        other_names = iosco_entity.get('Autres Noms', [])

        # Process primary name
        primary_tokens = tokenize_and_clean_name(primary_name)
        if primary_tokens:
            iosco_token_data.append((primary_tokens, iosco_entity, primary_name, "Primary Name"))

        # Process aliases
        for original_alias in other_names:
            if original_alias:
                alias_tokens = tokenize_and_clean_name(original_alias)
                if alias_tokens:
                    iosco_token_data.append((alias_tokens, iosco_entity, original_alias, "Alias"))


    # Iterate through each CRM company name, tokenize it, and compare with IOSCO token data
    unique_crm_companies_list = st.session_state.get('unique_crm_companies', [])
    total_crm_companies = len(unique_crm_companies_list)


    for i, crm_company in enumerate(unique_crm_companies_list):
        crm_company_name_str = str(crm_company.get("name", "")).strip()
        if not crm_company_name_str or crm_company_name_str in screened_crm_companies:
            continue

        screened_crm_companies.add(crm_company_name_str)
        crm_tokens = tokenize_and_clean_name(crm_company_name_str)
        if not crm_tokens:
            continue

        # Get the first significant word from the CRM name tokens
        crm_first_word = sorted(list(crm_tokens))[0] if crm_tokens else None # Take the first word alphabetically for consistency
        if not crm_first_word:
            continue # Skip if no significant first word

        best_match = None

        # Iterate through all prepared IOSCO token data entries to find subset matches first
        potential_matches = [] # Store matches based on subset rule
        for iosco_tokens, original_iosco_entity, original_iosco_name_string, name_type in iosco_token_data:
             # Check if CRM tokens are a subset of IOSCO tokens
             if crm_tokens.issubset(iosco_tokens):
                 # Calculate score for this potential match (using token_set_ratio for now)
                 calculated_score = fuzz.token_set_ratio(crm_company_name_str.lower(), original_iosco_name_string.lower())
                 potential_matches.append({
                     'iosco_tokens': iosco_tokens,
                     'original_iosco_entity': original_iosco_entity,
                     'original_iosco_name_string': original_iosco_name_string,
                     'name_type': name_type,
                     'calculated_score': calculated_score
                 })

        # Now, filter potential matches based on the beginning similarity rule
        filtered_matches = []
        for potential_match in potential_matches:
             iosco_tokens = potential_match['iosco_tokens']
             original_iosco_name_string = potential_match['original_iosco_name_string']
             calculated_score = potential_match['calculated_score']

             # Check beginning similarity
             crm_prefix_tokens = get_first_n_tokens(crm_tokens, 3) # Check similarity of the first 3 tokens or fewer if less than 3 exist
             
             
             # Let's get a cleaned prefix string for both CRM and IOSCO names
             crm_cleaned_prefix = " ".join(sorted(list(crm_prefix_tokens)))
             
             # Get a cleaned prefix from the IOSCO name using its first few tokens
             iosco_prefix_tokens = get_first_n_tokens(iosco_tokens, 3) # Compare with the first 3 IOSCO tokens
             iosco_cleaned_prefix = " ".join(sorted(list(iosco_prefix_tokens)))

             # Calculate fuzzy ratio of the cleaned prefixes
             prefix_fuzzy_score = fuzz.ratio(crm_cleaned_prefix, iosco_cleaned_prefix)

             # Condition pour une correspondance :
             # 1. Les jetons CRM sont un sous-ensemble des jetons IOSCO (déjà filtrés par potential_matches)
             # 2. Le score global token_set_ratio est supérieur ou égal au seuil
             # 3. Le ratio fuzzy des préfixes nettoyés est supérieur ou égal à un certain seuil (par exemple, 70 %)
             prefix_threshold = 70 # Define a threshold for prefix similarity
             
             if calculated_score >= threshold and prefix_fuzzy_score >= prefix_threshold:
                  filtered_matches.append(potential_match)


        # Select the best match from the filtered list (e.g., highest calculated_score)
        if filtered_matches:
            best_filtered_match = max(filtered_matches, key=lambda x: x['calculated_score'])
            original_iosco_entity = best_filtered_match['original_iosco_entity']
            original_iosco_name_string = best_filtered_match['original_iosco_name_string']
            name_type = best_filtered_match['name_type']
            calculated_score = best_filtered_match['calculated_score']

            matches.append({
                'Entity Name (CRM)': crm_company_name_str,
                'Source (CRM)': crm_company.get("source", ""),
                'Jurisdiction (Account)': crm_company.get("jurisdiction_account", ""),
                'Company Type (Account)': crm_company.get("company_type_account", ""),
                'Matched Name (IOSCO)': original_iosco_name_string,
                'Jurisdiction (IOSCO)': original_iosco_entity.get('nca_jurisdiction', ''),
                'Company Type (IOSCO)': original_iosco_entity.get('nca_name', ''),
                'Source (IOSCO)': original_iosco_entity.get('Source', 'IOSCO'),
                'Score': calculated_score,
                'Type': 'Legal Entity',
                'Regulatory Authority': original_iosco_entity.get('nca_name', ''), # Added Regulatory Authority
                'Commentaire': '',
                'Status': ''
            })


    print(f"✅ Screening des personnes morales terminé. {len(matches)} correspondances trouvées.")

    # Call progress callback one last time to show 100% completion for PM
    if progress_callback:
        progress_callback(total_crm_companies, total_crm_companies, "PM")


    return pd.DataFrame(matches)

# Fonction de screening principale
def process_all_reports(account_file, contact_file, leads_file, sanctions, threshold_nom, threshold_prenom, progress_callback):
    # Traiter chaque rapport
    account_people = process_account_report(account_file)
    contact_people = process_contact_report(contact_file)
    leads_people = process_leads_report(leads_file)
    
    # Combiner tous les résultats
    all_people = pd.concat([account_people, contact_people, leads_people], ignore_index=True)
    
    # Initialiser les colonnes de résultats avec les bons types
    all_people["Screening Result"] = "OK - Aucune correspondance trouvée"
    all_people["Liste de sanctions"] = ""
    all_people["Score Nom"] = 0.0
    all_people["Score Prénom"] = 0.0
    all_people["Score Total"] = 0.0
    all_people["Détails"] = ""
    
    # Filtrer les données aberrantes
    abnormal_data = filter_abnormal_data(all_people)
    normal_data = all_people[~all_people.index.isin(abnormal_data.index)].copy() # Use .copy() to avoid SettingWithCopyWarning
    
    # Filtrer les personnes qui doivent être analysées
    normal_data['should_analyze'] = normal_data.apply(should_analyze_person, axis=1)
    to_analyze = normal_data[normal_data['should_analyze']].copy() # Use .copy()
    
    # Effectuer le screening pour chaque personne
    total_to_analyze = len(to_analyze)
    for i, (idx, person) in enumerate(to_analyze.iterrows()):
        # Update progress via callback.
        # The callback now expects current index, total for THIS stage, and stage name
        progress_callback(i, total_to_analyze, "PP")

        if use_intelligent_scoring:
            best_match = intelligent_screening(person, sanctions, threshold_nom, threshold_prenom)
        else:
            best_match = traditional_screening(person, sanctions, threshold_nom, threshold_prenom)
        
        if best_match:
            # Use .loc for safe assignment
            normal_data.loc[idx, "Screening Result"] = f"Correspondance avec {best_match['Prénom']} {best_match['Nom']}"
            normal_data.loc[idx, "Liste de sanctions"] = best_match['Source']
            normal_data.loc[idx, "Score Nom"] = float(best_match['Score Nom'])
            normal_data.loc[idx, "Score Prénom"] = float(best_match['Score Prénom'])
            normal_data.loc[idx, "Score Total"] = float(best_match['Score Total'])
            normal_data.loc[idx, "Détails"] = best_match['Détails']
            normal_data.loc[idx, "Last Analysis"] = pd.Timestamp.now()
    
    # Update the original DataFrame 'all_people' with results from 'normal_data'
    all_people.update(normal_data)
    
    return all_people, abnormal_data

# Fonction de screening traditionnel (pour comparaison)
def traditional_screening(person, sanctions, threshold_nom, threshold_prenom):
    first_name = person["First Name"] if not pd.isna(person["First Name"]) else ""
    last_name = person["Last Name"] if not pd.isna(person["Last Name"]) else ""
    
    # Ignorer si le nom est trop court
    if len(last_name) < min_name_length:
        return None
    
    best_match = None
    best_score_total = 0
    
    for sanction in sanctions:
        sanction_nom = sanction.get('Nom', '')
        sanction_prenom = sanction.get('Prénom', '')
        
        # Calcul de similarité traditionnelle
        nom_score = calculate_name_similarity(last_name, sanction_nom)
        prenom_score = calculate_name_similarity(first_name, sanction_prenom) if first_name and sanction_prenom else 0
        
        # Si les scores de base sont trop bas, passer à la sanction suivante
        if nom_score < threshold_nom or (require_both_match and first_name and prenom_score < threshold_prenom):
            continue
        
        # Calculer le score total
        if first_name and require_both_match:
            # Moyenne pondérée: le nom a plus de poids que le prénom
            score_total = (nom_score * 0.6) + (prenom_score * 0.4)
        else:
            score_total = nom_score
        
        # Mettre à jour le meilleur match si nécessaire
        if score_total > best_score_total:
            best_score_total = score_total
            best_match = {
                'Nom': sanction.get('Nom', ''),
                'Prénom': sanction.get('Prénom', ''),
                'Source': sanction.get('Source', ''),
                'Score Nom': nom_score,
                'Score Prénom': prenom_score,
                'Score Total': round(score_total, 1),
                'Détails': f"Nom: {nom_score:.1f}%, Prénom: {prenom_score:.1f}%"
            }
    
    return best_match

# Fonctions de traitement des rapports
def process_account_report(file):
    try:
        # Lire le fichier Excel en sautant les lignes d'en-tête
        # Read columns T to AB which should include "Account Name", "UBO name", "Signatory"
        df = pd.read_excel(file, header=8, usecols="T:AB")
        people = []

        # Colonnes UBO et Signatory and Account Name
        ubo_columns = ["UBO name 1", "UBO name 2", "UBO name 3", "UBO name 4"]
        signatory_columns = ["Signatory 1", "Signatory 2", "Signatory 3", "Signatory 4"]
        # Assuming "Account Name" is within the T:AB range
        account_name_column = "Account Name"

        for _, row in df.iterrows():
            account_name = row.get(account_name_column, "") if pd.notna(row.get(account_name_column)) else ""

            # Traiter les UBOs
            for col in ubo_columns:
                if pd.notna(row.get(col)) and row.get(col) != "":
                    first_name, last_name = split_full_name(row.get(col))
                    if len(last_name) >= min_name_length:
                        people.append({
                            "First Name": first_name,
                            "Last Name": last_name,
                            "Type": "UBO",
                            "Source": "Account Report",
                            "Title": "", # No Title for Account UBOs
                            "Company/Account": account_name, # Use Account Name as Company/Account
                            "Commentaire": "",
                            "Status": "",
                            "Last Analysis": None
                        })

            # Traiter les Signatories
            for col in signatory_columns:
                if pd.notna(row.get(col)) and row.get(col) != "":
                    first_name, last_name = split_full_name(row.get(col))
                    if len(last_name) >= min_name_length:
                        people.append({
                            "First Name": first_name,
                            "Last Name": last_name,
                            "Type": "Signatory",
                            "Source": "Account Report",
                            "Title": "", # No Title for Account Signatories
                            "Company/Account": account_name, # Use Account Name as Company/Account
                            "Commentaire": "",
                            "Status": "",
                            "Last Analysis": None
                        })

        return pd.DataFrame(people)
    except Exception as e:
        st.error(f"Erreur lors du traitement du fichier Account: {str(e)}")
        return pd.DataFrame()

def process_contact_report(file):
    try:
        # Lire le fichier Excel en sautant les lignes d'en-tête
        # Read columns E to H to get First Name, Last Name, Title, Account Name
        df = pd.read_excel(file, header=8, usecols="E:H")
        people = []

        for _, row in df.iterrows():
            first_name = row.get("First Name")
            last_name = row.get("Last Name")
            title = row.get("Title", "") if pd.notna(row.get("Title")) else ""
            account_name = row.get("Account Name", "") if pd.notna(row.get("Account Name")) else ""

            if pd.notna(first_name) and pd.notna(last_name):
                 last_name_str = str(last_name)
                 if len(last_name_str) >= min_name_length:
                    # Removed special handling for Xavier Morin, treat like any other contact
                    company_account_value = account_name # Use original account_name

                    people.append({
                        "First Name": first_name,
                        "Last Name": last_name_str,
                        "Type": "Contact",
                        "Source": "Contact Report",
                        "Title": title, # Use Title from column G
                        "Company/Account": company_account_value, # Use the determined value
                        "Commentaire": "",
                        "Status": "",
                        "Last Analysis": None
                    })

        return pd.DataFrame(people)
    except Exception as e:
        st.error(f"Erreur lors du traitement du fichier Contact: {str(e)}")
        return pd.DataFrame()

def process_leads_report(file):
    try:
        # Lire le fichier Excel en sautant les lignes d'en-tête
        df = pd.read_excel(file, header=12, usecols="C,D,H,T")
        people = []

        for _, row in df.iterrows():
            first_name = row.get("First Name")
            last_name = row.get("Last Name")
            title = row.get("Title", "") if pd.notna(row.get("Title")) else ""
            company_account = row.get("Company/Account", "") if pd.notna(row.get("Company/Account")) else ""
        

            if pd.notna(first_name) and pd.notna(last_name):
                 last_name_str = str(last_name)
                 if len(last_name_str) >= min_name_length:
                    people.append({
                        "First Name": first_name,
                        "Last Name": last_name_str,
                        "Type": "Lead",
                        "Source": "Leads Report",
                        "Title": title, # Use Title from column H
                        "Company/Account": company_account,
                        "Commentaire": "",
                        "Status": "",
                        "Last Analysis": None
                    })

        return pd.DataFrame(people)
    except Exception as e:
        st.error(f"Erreur lors du traitement du fichier Leads: {str(e)}")
        return pd.DataFrame()

def export_pm_excel_styled(df, filename):
    display_columns = [
        'Entity Name (CRM)', 'Source (CRM)', 'Jurisdiction (Account)', 'Company Type (Account)',
        'Matched Name (IOSCO)', 'Source (IOSCO)', 'Jurisdiction (IOSCO)', 'Regulatory Authority',
        'Score', 'Type', 'Commentaire', 'Status'
    ]
    df_export = df[[col for col in display_columns if col in df.columns]].copy()
    if 'Score' in df_export.columns:
        df_export['Score'] = df_export['Score'].apply(lambda x: f"{float(str(x).replace('%', '').strip()):.1f} %" if pd.notna(x) and str(x).strip() != '' else '')
    wb = Workbook()
    ws = wb.active
    ws.title = "Correspondances PM"
    ws.append(["Liste des correspondances personnes morales"])
    # Check if the DataFrame is not empty before merging cells
    if not df_export.empty:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_export.columns))
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append(df_export.columns.tolist())
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    for row in dataframe_to_rows(df_export, index=False, header=False):
        ws.append(row)
    # Largeur auto (corrigé)
    for i, col in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row), 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2
    wb.save(filename)

def export_pp_excel_styled(df, filename):
    display_columns = [
        'First Name', 'Last Name', 'Type', 'Source', 'Title', 'Company/Account',
        'Screening Result', 'Liste de sanctions', 'Score Nom', 'Score Prénom',
        'Score Total', 'Détails', 'Commentaire', 'Status'
    ]
    df_export = df[[col for col in display_columns if col in df.columns]].copy()
    for col in ['Score Nom', 'Score Prénom', 'Score Total']:
        if col in df_export.columns:
            df_export[col] = df_export[col].apply(lambda x: f"{float(str(x).replace('%', '').strip()):.1f} %" if pd.notna(x) and str(x).strip() != '' else '')
    wb = Workbook()
    ws = wb.active
    ws.title = "Correspondances PP"
    ws.append(["Liste des correspondances personnes physiques"])
    # Check if the DataFrame is not empty before merging cells
    if not df_export.empty:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_export.columns))
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append(df_export.columns.tolist())
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    for row in dataframe_to_rows(df_export, index=False, header=False):
        ws.append(row)
    # Largeur auto (corrigé)
    for i, col in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row), 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2
    wb.save(filename)

def generate_outlook_email_both(pm_df, pp_df, output_file_pm, output_file_pp):
    try:
        today = dt.datetime.today().strftime("%Y%m%d")
        mail_subject = f"{today} - Contrôle PG06 - LCB-FT - Résultat screening CRM"

        # Génère les deux tableaux HTML
        pm_table_html = dataframe_to_html_table(pm_df, display_columns=[
            'Entity Name (CRM)', 'Source (CRM)', 'Jurisdiction (Account)', 'Company Type (Account)',
            'Matched Name (IOSCO)', 'Source (IOSCO)', 'Jurisdiction (IOSCO)', 'Regulatory Authority',
            'Score', 'Type', 'Commentaire', 'Status'
        ])
        pp_table_html = dataframe_to_html_table(pp_df, display_columns=[
            'First Name', 'Last Name', 'Type', 'Source', 'Title', 'Company/Account',
            'Screening Result', 'Liste de sanctions', 'Score Nom', 'Score Prénom',
            'Score Total', 'Détails', 'Commentaire', 'Status'
        ])
        # Ajoute les liens vers les fichiers s'ils existent
        pm_link = f'<p><a href="file://{os.path.abspath(output_file_pm)}">📥 Télécharger PM</a></p>' if os.path.exists(output_file_pm) else ""
        pp_link = f'<p><a href="file://{os.path.abspath(output_file_pp)}">📥 Télécharger PP</a></p>' if os.path.exists(output_file_pp) else ""
        # Corps du mail
        html_body = f'''
        <html>
        <head>
        <style>
        table {{ border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; font-size: 13px; }}
        th {{ background-color: #2F5496; color: #fff; padding: 8px 10px; border: 1px solid #b6c2d2; font-size: 14px; }}
        td {{ padding: 8px 10px; border: 1px solid #b6c2d2; vertical-align: top; }}
        </style>
        </head>
        <body>
            <p>Bonjour,</p>
            <p>Veuillez trouver ci-dessous la liste des correspondances détectées lors du contrôle LCB-FT.</p>
            <h2>1- Listes des correspondances personnes morales</h2>
            {pm_link}
            {pm_table_html}
            <h2>2- Listes des correspondances personnes physiques</h2>
            {pp_link}
            {pp_table_html}
            <p style="margin-top:20px;">Bien cordialement,<br></p>
        </body>
        </html>
        '''

        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.SentOnBehalfOfName = "rcci@syquant.com"
        mail.To = "rcci@syquant.com"
        mail.CC = " jeremy.saus@syquant.com"
        mail.Subject = mail_subject
        mail.HTMLBody = html_body

        # Ajoute les deux fichiers en pièce jointe
        if os.path.exists(output_file_pm):
            mail.Attachments.Add(os.path.abspath(output_file_pm))
        if os.path.exists(output_file_pp):
            mail.Attachments.Add(os.path.abspath(output_file_pp))

        mail.Display()
        st.success("✅ Email Outlook généré et affiché avec succès.")

    except Exception as e:
        st.error(f"❌ Erreur Outlook : {e}")
        
def generate_outlook_email_pp(df, output_file_matches_only):
    try:
        # Filtrer les correspondances si la colonne existe
        if 'Screening Result' in df.columns:
            matches_df = df[df['Screening Result'] != "OK - Aucune correspondance trouvée"]
        else:
            matches_df = df

        if matches_df.empty:
            st.info("Aucune correspondance trouvée, aucun mail généré.")
            return False

        today = dt.datetime.today()
        formatted_date = today.strftime("%Y%m%d")
        mail_subject = f"{formatted_date} - Contrôle PG06 - LCB-FT - Résultat screening CRM (PP)"

        # Colonnes à afficher pour PP
        pp_display_columns = [
            'First Name', 'Last Name', 'Type', 'Source', 'Title', 'Company/Account',
            'Screening Result', 'Liste de sanctions', 'Score Nom', 'Score Prénom',
            'Score Total', 'Détails', 'Commentaire', 'Status'
        ]
        table_html = dataframe_to_html_table(matches_df, display_columns=pp_display_columns)

        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; font-size: 13px;">
            <p>Bonjour,</p>
            <p>
                Veuillez trouver ci-dessous la liste des correspondances détectées lors du contrôle LCB-FT.<br>
            </p>
            {table_html}
            <p style="margin-top:20px;">Bien cordialement,<br></p>
        </body>
        </html>
        """

        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = "rcci@syquant.com"
        mail.CC = "jeremy.saus@syquant.com"
        mail.Subject = mail_subject
        mail.SentOnBehalfOfName = "rcci@syquant.com"
        mail.HTMLBody = html_body

        abs_path = os.path.abspath(output_file_matches_only)
        if os.path.exists(abs_path):
            mail.Attachments.Add(abs_path)
        else:
            st.warning(f"La pièce jointe n'a pas été trouvée : {abs_path}")

        try:
            mail.Display()
        except Exception as e:
            st.error(f"Erreur lors de l'affichage du mail : {e}")

        st.success("✅ Email Outlook généré avec succès.")
        return True

    except Exception as e:
        st.error(f"❌ Erreur Outlook : {e}")
        return False

# Fonction pour générer un email Outlook
def generate_outlook_email_custom(df, output_file_matches_only):
    try:
        # Check if 'Screening Result' column exists
        if 'Screening Result' in df.columns:
            matches_df = df[df['Screening Result'] != "OK - Aucune correspondance trouvée"]
        else:
            # If the column doesn't exist, assume the input dataframe already contains the matches
            matches_df = df

        if matches_df.empty:
            st.info("Aucune correspondance trouvée, aucun mail généré.")
            return False

        today = dt.datetime.today()
        formatted_date = today.strftime("%Y%m%d")
        mail_subject = f"{formatted_date} - Contrôle PG06 - LCB-FT - Résultat screening CRM"

        # Use the custom function for the table
        # Define columns to display for PM results
        # START: Added PM specific display columns
        pm_display_columns = [
            'Entity Name (CRM)', 'Source (CRM)', 'Jurisdiction (Account)', 'Company Type (Account)',
            'Matched Name (IOSCO)', 'Source (IOSCO)', 'Jurisdiction (IOSCO)', 'Regulatory Authority',
            'Score', 'Type', 'Commentaire', 'Status'
        ]
        # END: Added PM specific display columns
        # Use pm_display_columns for the PM email
        table_html = dataframe_to_html_table(matches_df, display_columns=pm_display_columns)

        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; font-size: 13px;">
            <p>Bonjour,</p>
            <p>
                Veuillez trouver ci-dessous la liste des correspondances détectées lors du contrôle LCB-FT.<br>
            </p>
            {table_html}
            <p style="margin-top:20px;">Bien cordialement,<br></p>
        </body>
        </html>
        """

        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = "rcci@syquant.com"
        mail.CC = "jeremy.saus@syquant.com"
        mail.Subject = mail_subject
        mail.SentOnBehalfOfName = "rcci@syquant.com"
        mail.HTMLBody = html_body

        abs_path = os.path.abspath(output_file_matches_only)
        if os.path.exists(abs_path):
            mail.Attachments.Add(abs_path)
        else:
            st.warning(f"La pièce jointe n'a pas été trouvée : {abs_path}")

        try:
            mail.Display()
        except Exception as e:
            st.error(f"Erreur lors de l'affichage du mail : {e}")

        st.success("✅ Email Outlook généré  avec succès.")
        return True

    except Exception as e:
        st.error(f"❌ Erreur Outlook : {e}")
        return False

def filter_abnormal_data(df):
    """Filtre les données aberrantes (noms/prénoms trop courts)"""
    abnormal = df[
        (df['First Name'].str.len() <= 1) |
        (df['Last Name'].str.len() <= 1)
    ].copy()
    return abnormal

def should_analyze_person(person, days_threshold=3):
    """Vérifie si une personne doit être analysée en fonction de son statut et de la date de dernière analyse"""
    if not person.get('Last Analysis'):
        return True
    
    if person.get('Status') in ["", "Résolu", "En cours d'investigation", "Sous sanction"]:
        last_analysis = pd.to_datetime(person['Last Analysis'])
        days_since_analysis = (pd.Timestamp.now() - last_analysis).days
        return days_since_analysis >= days_threshold
    
    return True

def save_excel_with_style(df, filename):
    """Sauvegarde un DataFrame dans un fichier Excel avec un style personnalisé"""
    try:
        # Essayer d'abord de sauvegarder directement
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # Accéder à la feuille de calcul
        worksheet = writer.sheets['Sheet1']
        
        # Définir les styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Appliquer le style aux en-têtes
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Ajuster la largeur des colonnes
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)
        
        writer.close()
        return True
    except PermissionError:
        try:
            # Si erreur de permission, essayer de sauvegarder dans un fichier temporaire
            temp_dir = tempfile.gettempdir()
            temp_filename = os.path.join(temp_dir, os.path.basename(filename))
            
            writer = pd.ExcelWriter(temp_filename, engine='openpyxl')
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # Accéder à la feuille de calcul
            worksheet = writer.sheets['Sheet1']
            
            # Définir les styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Appliquer le style aux en-têtes
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # Ajuster la largeur des colonnes
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(str(col))
                )
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)
            
            writer.close()
            
            # Copier le fichier temporaire vers la destination finale
            import shutil
            shutil.copy2(temp_filename, filename)
            
            # Supprimer le fichier temporaire
            os.remove(temp_filename)
            return True
        except Exception as e:
            st.error(f"Erreur lors de la sauvegarde du fichier : {e}")
            return False
    except Exception as e:
        st.error(f"Erreur lors de la sauvegarde du fichier : {e}")
        return False

def send_mail(Subject, Body, To, displayOnly=False, sign=False, CC=None):
    try:
        ol = win32.GetActiveObject('Outlook.Application')
        olmailitem = 0x0
        newmail = ol.CreateItem(olmailitem)
        signature = ""
        if sign:
            try:
                newmail.Display()
            except:
                pass
            signature = newmail.HTMLBody
        newmail.Subject = Subject
        newmail.To = To
        if CC is not None and CC.strip() != "":
            newmail.CC = CC
        html_text = '<html><body>'
        html_text += Body.replace('\n', '<br>')
        html_text += '</body></html>'        
        newmail.HTMLBody = html_text + signature
        if displayOnly:
            try:
                newmail.Display()
                print("Mail affiché !")
            except Exception as e:
                print(f"Erreur Display : {e}")
                st.error(f"Erreur lors de l'affichage du mail : {e}")
        else:
            newmail.Send()
    except Exception as e:
        print(f"Erreur lors de l'envoi du mail : {e}")

def dataframe_to_html_table(df, display_columns=None):
    # Define styles
    table_style = "border-collapse:collapse;width:100%;font-family:Arial,sans-serif;font-size:13px;"
    th_style = "background-color:#2F5496;color:#fff;padding:8px 10px;border:1px solid #b6c2d2;font-size:14px;"
    td_style = "padding:8px 10px;border:1px solid #b6c2d2;vertical-align:top;word-wrap:break-word;"
    tr_even = "background-color:#f6f6f6;"
    tr_odd = "background-color:#ffffff;"
    # Columns to format as percentage
    percent_cols = ["Score Nom", "Score Prénom", "Score Total", "Score"] # Added "Score" for PM
    # Define the order and presence of columns (Default for PP)
    default_display_columns = [
        'First Name', 'Last Name', 'Type', 'Source', 'Title', 'Company/Account',
        'Screening Result', 'Liste de sanctions', 'Score Nom', 'Score Prénom',
        'Score Total', 'Détails', 'Commentaire', 'Status'
    ]

    # Use provided display_columns if available, otherwise use default
    cols_to_include_order = display_columns if display_columns is not None else default_display_columns

    # Filter the DataFrame to only include columns that exist in both the df and the desired display_columns, in the specified order
    cols_to_include = [col for col in cols_to_include_order if col in df.columns]
    df_display = df[cols_to_include]

    html = f'<table style="{table_style}">'
    # Headers
    html += "<tr>"
    for col in df_display.columns: # Iterate through the columns in the display order
        html += f'<th style="{th_style}">{col}</th>'
    html += "</tr>"
    # Rows
    for i, row in df_display.iterrows():
        row_style = tr_even if i % 2 == 0 else tr_odd
        html += f'<tr style="{row_style}">'
        for col in df_display.columns: # Iterate through the columns in the display order
             val = row[col]
             
             # Apply specific formatting for 'Matched Name (IOSCO)' column
             if col == 'Matched Name (IOSCO)' and pd.notna(val) and isinstance(val, str):
                 original_val = str(val).strip()
                 cleaned_val = original_val

                 match_url_with_paren = re.match(r'^(?:http[s]?://)?(?:www\.)?[a-z0-9.-]+(?:/\S*)?\s*\(([^)]*)\)', cleaned_val, re.IGNORECASE)

                 if match_url_with_paren:
                     # If pattern matches, use the captured content from parentheses
                     content_in_paren = match_url_with_paren.group(1).strip()
                     display_val = content_in_paren.title() # Capitalize words
                 else:
                     # If pattern doesn't match, apply the previous cleaning (remove all parentheses content and basic URL parts, then capitalize)
                     cleaned_val = re.sub(r'\s*\([^)]*\)', '', cleaned_val).strip()
                     cleaned_val = re.sub(r'^(http[s]?://)?(www\.)?', '', cleaned_val).strip()
                     cleaned_val = re.sub(r'\.co\.uk$|\.com$|\.org$|\.net$|\\.int$', '', cleaned_val).strip() # Remove common TLDs
                     display_val = cleaned_val.title() # Capitalize words

             elif col in percent_cols and pd.notna(val):
                 try:
                     display_val = f"{float(val):.1f} %"
                 except:
                     display_val = val # Fallback if conversion fails
             else:
                 # Handle None/NaN values and other columns
                 display_val = val if pd.notna(val) else ""
                 
             html += f'<td style="{td_style}">{display_val}</td>'
        html += "</tr>"
    html += "</table>"
    return html

def load_screening_history(results_dir):
    """Charge l'historique des correspondances PP et PM avec Types + Statuts"""
    import os
    import pandas as pd    
    if not (results_dir and os.path.exists(results_dir) and os.path.isdir(results_dir)):
        st.warning(f"Dossier des résultats invalide : '{results_dir}'")
        return pd.DataFrame()
        
    files = [f for f in os.listdir(results_dir)
             if f.endswith(".xlsx") and (
                 "screening des correspondances PP.xlsx" in f or
                 "screening des résultats PM.xlsx" in f or
                 "screening des correspondances.xlsx" in f
             )]
    files = sorted(files)
    
    # Debug: Afficher les fichiers trouvés
    # st.write(f"Fichiers trouvés: {files}")
    
    grouped_by_date = {}
    for file in files:
        try:
            date_str = file.split(" - ")[0]
            date_obj = pd.to_datetime(date_str, format="%Y%m%d")
            file_path = os.path.join(results_dir, file)
            
            # Essayer différentes configurations d'en-tête
            try:
                df = pd.read_excel(file_path, header=1)
            except:
                try:
                    df = pd.read_excel(file_path, header=0)
                except:
                    continue
            
            df.columns = df.columns.str.strip()
            
            entry = grouped_by_date.setdefault(date_obj, {
                "Correspondances PP": 0,
                "Correspondances PM": 0,
                "Types": {},
                "Statuts": {}
            })
            
            # Déterminer le type d'entité correctement
            if "résultats PM" in file:
                entity_type = "PM"
            else:
                # Si c'est "correspondances PP" ou simplement "correspondances", c'est PP
                entity_type = "PP"
            
            # Filtrer les correspondances réelles
            if "Screening Result" in df.columns:
                df = df[df["Screening Result"] != "OK - Aucune correspondance trouvée"]
                
            # Compte total
            count = len(df)
            if entity_type == "PP":
                entry["Correspondances PP"] += count
            else:
                entry["Correspondances PM"] += count
                
            # Compter les types
            if 'Type' in df.columns:
                for t, nb in df['Type'].value_counts().items():
                    if pd.notna(t):  # Ignorer les valeurs NaN
                        entry["Types"][t] = entry["Types"].get(t, 0) + nb
                        
            # Compter les statuts - Vérifier différentes orthographes possibles
            status_columns = [col for col in df.columns if col.lower() in ['status', 'statut', 'état']]
            
            if status_columns:
                status_col = status_columns[0]
                
                # Si le statut n'est pas renseigné dans les anciens fichiers, attribuer "En cours d'investigation"
                df[status_col] = df[status_col].fillna("En cours d'investigation")
                
                for s, nb in df[status_col].value_counts().items():
                    if pd.notna(s):  # Ignorer les valeurs NaN
                        status_label = f"{s} ({entity_type})"
                        entry["Statuts"][status_label] = entry["Statuts"].get(status_label, 0) + nb
            else:
                # Si pas de colonne statut, considérer tous comme "En cours d'investigation"
                status_label = f"En cours d'investigation ({entity_type})"
                entry["Statuts"][status_label] = entry["Statuts"].get(status_label, 0) + count
                
        except Exception as e:
            print(f"Erreur dans {file} : {e}")
    
    # Finaliser le tableau
    history = []
    for date, val in grouped_by_date.items():
        pp = val.get("Correspondances PP", 0)
        pm = val.get("Correspondances PM", 0)
        total = pp + pm
        history.append({
            "Date": date,
            "Correspondances PP": pp,
            "Correspondances PM": pm,
            "Total": total,
            "Types": val.get("Types", {}),
            "Statuts": val.get("Statuts", {})
        })
    
    # Créer et trier le DataFrame
    history_df = pd.DataFrame(history)
    if not history_df.empty:
        history_df = history_df.sort_values("Date")
    
    return history_df
def list_screening_history(results_dir):
    """
    Liste l'historique des screenings dans le dossier donné, en séparant PP, PM et correspondances.
    Retourne un DataFrame avec les colonnes : date, type (PP/PM/correspondance), fichier.
    Ajoute une colonne 'missing' pour indiquer les dates où il manque un screening PP ou PM.
    """
    import os
    import re
    import pandas as pd
    from datetime import datetime, timedelta

    # Liste des fichiers
    files = os.listdir(results_dir)
    pattern = re.compile(r"(\d{8})\s*[-_].*(PP|PM|correspondance)", re.IGNORECASE)
    records = []
    for f in files:
        m = pattern.search(f)
        if m:
            date_str = m.group(1)
            type_str = m.group(2).upper()
            try:
                date = datetime.strptime(date_str, "%Y%m%d").date()
            except Exception:
                continue
            if "CORRESPONDANCE" in type_str:
                type_str = "correspondance"
            elif "PP" in type_str:
                type_str = "PP"
            elif "PM" in type_str:
                type_str = "PM"
            else:
                type_str = "autre"
            records.append({"date": date, "type": type_str, "file": f})

    df = pd.DataFrame(records)
    if df.empty:
        return pd.DataFrame(columns=["date", "type", "file", "missing"])

    # Générer la liste de toutes les dates entre min et max
    all_dates = pd.date_range(df["date"].min(), df["date"].max()).date
    types = ["PP", "PM", "correspondance"]
    full_index = pd.MultiIndex.from_product([all_dates, types], names=["date", "type"])
    df_full = pd.DataFrame(index=full_index).reset_index()
    df = pd.merge(df_full, df, on=["date", "type"], how="left")
    df["missing"] = df["file"].isna()
    return df

def load_account_report(file):
    """Charge et analyse le Report Account Olley"""
    try:
        # Lire le fichier Excel en sautant les lignes d'en-tête
        df = pd.read_excel(file, header=8)       
       
        # Nettoyer les noms des colonnes
        df.columns = df.columns.str.strip()
        
        return df
    except Exception as e:
        st.error(f"Erreur lors du chargement du fichier Account Report : {e}")
        return None

def create_dashboard(df, column, title):
    """Crée un tableau de bord pour une colonne spécifique"""
    try:
        # Vérifier si la colonne existe
        if column not in df.columns:
            st.warning(f"La colonne '{column}' n'existe pas dans le fichier. Colonnes disponibles : {', '.join(df.columns)}")
            return None, None
            
        # Compter les occurrences
        counts = df[column].value_counts()
        
        # --- Handle 'Source of regulation' specifics ---
        if column == 'Source of regulation':
            # Extract initials or use full name if no initials found
            def get_initials_or_name(value):
                if pd.isna(value) or value == "":
                    return ""
                match = re.search(r'\((.*?)\)', str(value))
                if match:
                    return f"({match.group(1)})"
                return str(value)

            counts.index = counts.index.map(get_initials_or_name)
            # Recalculate counts after mapping initials (important if multiple sources map to same initial)
            counts = counts.groupby(counts.index).sum().sort_values(ascending=False)
        else:
            # For other columns, ensure index is just the value and sort
            counts = counts.sort_values(ascending=False)

        # Define a consistent discrete color palette
        # You can modify these hex codes to change the colors
        discrete_color_palette = [
            '#0068C9', '#83C9FF', '#FF9500', '#FF5733', '#05a33a', # Example colors
            '#6a329f', '#b4a7d6', '#a64d79', '#d5a6bd', '#e06666', # More example colors
            '#f6b26b', '#ffd966', '#93c47d', '#674ea7', '#c27ba0', # And more
            '#a4c2f4', '#c9daf8', '#ebf1de', '#f4cccc', '#fce5cd'  # Continue as needed
        ]
        
        # Define specific color maps for certain columns
        color_map = None
        if column == 'Vigilance':
            color_map = {
                'Simplified': '#05a33a',  # Vert
                'Reinforced': '#e60000',  # Rouge
                'Standard vigilance': '#FF9500' # Orange (using a color from the general palette as a default example)
            }
        elif column == 'Rebate agreement':
             color_map = {
                 'No': '#05a33a', # Vert
                 'Yes': '#e60000' # Rouge
             }

        # Create the chart based on the column type
        if column == 'Account Name':
             # Use horizontal bar chart for Account Name for better readability
             fig = px.bar(
                 x=counts.values,
                 y=counts.index,
                 title=title,
                 labels={'x': 'Nombre', 'y': column},
                 color=counts.values, # Color based on count
                 color_continuous_scale='Viridis', # Use a sequential color scale
                 orientation='h'
             )
             # Improve layout for horizontal chart
             fig.update_layout(
                 xaxis_title="Nombre",
                 yaxis_title=column,
                 showlegend=False,
                 height=min(max(400, len(counts) * 15), 600), # Adjust height based on number of accounts
                 yaxis={'categoryorder': 'total ascending'} # Order bars by count (bottom to top)
             )
        else:
            # Create a standard vertical bar chart for other columns
            fig = px.bar(
                x=counts.index,
                y=counts.values,
                title=title,
                labels={'x': column, 'y': 'Nombre'},
                # Use discrete_color_sequence or color_discrete_map
                color=counts.index, # Use the category itself for coloring
                color_discrete_sequence=discrete_color_palette[:len(counts.index)], # Use only as many colors as needed
                color_discrete_map=color_map # Apply specific color map if defined
            )
            # Personnaliser le layout for vertical chart
            fig.update_layout(
                xaxis_title=column,
                yaxis_title="Nombre",
                showlegend=True if color_map is None else False, # Show legend only for discrete_color_sequence
                height=400
            )
            # Ensure legend is shown for specific color maps if needed (e.g. Vigilance)
            if column == 'Vigilance' or column == 'Rebate agreement':
                 fig.update_layout(showlegend=True)

        
        # Créer un tableau de données
        table_df = pd.DataFrame({
            column: counts.index,
            'Nombre': counts.values,
            'Pourcentage': (counts.values / len(df) * 100).round(2)
        })
        
        return fig, table_df
    except Exception as e:
        st.error(f"Erreur lors de la création du tableau de bord pour {column}: {e}")
        return None, None

def analyze_inconsistencies(df):
    """Analyse les incohérences dans les données"""
    inconsistencies = []
    
    # 1. Vérifier les différences entre Billing Country et Jurisdiction
    country_jurisdiction_mismatch = df[df['Billing Country (text only)'] != df['Jurisdiction']][
        ['Account Name', 'Billing Country (text only)', 'Jurisdiction']
    ].copy()
    
    # 2. Vérifier les Business relationships avec Source of regulation ou Agreement number manquants
    business_relationships = df[df['Relationship'] == 'Business relationship'].copy()
    missing_info = business_relationships[
        (business_relationships['Source of regulation'].isna()) | 
        (business_relationships['Source of regulation'] == '') |
        (business_relationships['Agreement number'].isna()) |
        (business_relationships['Agreement number'] == '')
    ][['Account Name', 'Relationship', 'Source of regulation', 'Agreement number']]
    
    # 3. Vérifier les accords de remise actifs
    active_rebates = df[
        (df['Rebate agreement'] == 'Yes') & 
        (df['Rebate status'] == 'Active')
    ][[
        'Account Name', 'Relationship', 'Company type', 'Company activity',
        'Jurisdiction', 'Source of regulation', 'Agreement number', 'Vigilance',
        'UBO name 1', 'UBO name 2', 'UBO name 3', 'UBO name 4',
        'Signatory 1', 'Signatory 2', 'Signatory 3', 'Signatory 4'
    ]].copy()
    
    return country_jurisdiction_mismatch, missing_info, active_rebates

def display_matches_interactive(matches_df, type_pp_pm="PP"):
    if matches_df is None or matches_df.empty:
        st.info("Aucune correspondance trouvée.")
        return

    st.subheader(f"Liste des {len(matches_df)} correspondances trouvées")
    show_expanders = st.checkbox("Afficher le menu déroulant des correspondances", value=False, key=f"show_expanders_{type_pp_pm}")

    if show_expanders:
        for idx, row in matches_df.iterrows():
            # Titre du menu déroulant
            if type_pp_pm == "PP":
                title = f"{row['First Name']} {row['Last Name']} | {row['Type']} | {row['Source']} | {row['Screening Result']}"
            else:  # PM
                title = f"{row['Entity Name (CRM)']} | {row['Source (CRM)']} | Correspondance avec {row['Matched Name (IOSCO)']}"

            with st.expander(title, expanded=False):
                # Affichage du tableau récapitulatif
                if type_pp_pm == "PP":
                    pp_display_cols_subset = [
                        'First Name', 'Last Name', 'Type', 'Source', 'Title','Company/Account','Screening Result',
                        'Liste de sanctions','Score Total', 'Commentaire', 'Status'
                    ]
                    cols_to_show_in_table = [col for col in pp_display_cols_subset if col in row.index]
                    
                    # Create a single-row DataFrame for horizontal display without the index
                    display_df_pp = pd.DataFrame([row[cols_to_show_in_table]])
                    st.dataframe(display_df_pp, hide_index=True, use_container_width=True)
                else:  # PM
                    # Define the columns to attempt to display for PM
                    pm_display_cols_subset = [
                        'Entity Name (CRM)', 'Source (CRM)', 'Jurisdiction (Account)', 'Company Type (Account)',
                        'Matched Name (IOSCO)', 'Source (IOSCO)', 'Jurisdiction (IOSCO)', 'Regulatory Authority',
                        'Score', 'Commentaire', 'Status'
                    ]
                    # Filter the columns to include only those present in the current row's index
                    cols_to_show_in_table = [col for col in pm_display_cols_subset if col in row.index]
                    
                    # Create a single-row DataFrame for horizontal display without the index
                    display_df_pm = pd.DataFrame([row[cols_to_show_in_table]])
                    st.dataframe(display_df_pm, hide_index=True, use_container_width=True)

                # Commentaire (avec stockage session_state)
                comment_key = f"comment_{type_pp_pm}_{idx}"
                status_key = f"status_{type_pp_pm}_{idx}"

                # Valeur par défaut
                default_comment = row.get("Commentaire", "")
                default_status = row.get("Status", "")

                comment = st.text_area("Commentaire", value=st.session_state.get(comment_key, default_comment), key=comment_key)
                status_options = ["", "Résolu", "En cours d'investigation", "Sous sanction"]
                default_status = row.get("Status", "")
                if pd.isna(default_status):
                    default_status = ""
                else:
                    default_status = str(default_status)
                status = st.selectbox(
                    "Statut",
                    status_options,
                    index=status_options.index(default_status) if default_status in status_options else 0,
                    key=status_key
                )
                # Mettre à jour le DataFrame (en mémoire) à chaque modification
                matches_df.at[idx, "Commentaire"] = comment
                matches_df.at[idx, "Status"] = status

    # Optionnel : retour du DataFrame modifié
    return matches_df
def get_latest_file(directory, pattern):
    files = glob.glob(os.path.join(directory, pattern))
    if not files:
        return None
    return max(files, key=os.path.getmtime)
# Interface principale
def main():
    st.markdown(
        """
        <h1 style='text-align: center; color: #1F4E79; font-family: "Segoe UI", sans-serif;'>🕵️‍♂️ Anti-Money Laundering System</h1>
        <br>
        """,
        unsafe_allow_html=True
    )
    # Initialize session state variables if they don't exist
    if 'screening_results' not in st.session_state:
        st.session_state.screening_results = None
    if 'abnormal_data' not in st.session_state:
        st.session_state.abnormal_data = None
    if 'screening_completed' not in st.session_state:
        st.session_state.screening_completed = False
    if 'comments' not in st.session_state:
        st.session_state.comments = {}
    if 'statuses' not in st.session_state: #
        st.session_state.statuses = {}
    if 'run_screening_command' not in st.session_state:
        st.session_state.run_screening_command = False
    # Initialize AMF/CFT levels at the very beginning of main()
    if 'amf_cft_levels' not in st.session_state:
        st.session_state.amf_cft_levels = load_amf_cft_levels()
    # Initialize session state for screening status
    if 'screening_status' not in st.session_state:
        st.session_state.screening_status = 'initial' 

    # Onglets
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "⚙️ Screening automatique",
        "📊 Suivi des analyses",
        "📈 Analyse des comptes",
        "📄 Analyse du registre",
        "🔎 Recherche manuelle",
    ])
# ------------------------------------------  SCREENING DES SANCTIONS --------------------------------------------------------------------
    with tab1:

        # Vérification et téléchargement automatique des listes de sanctions
        check_and_download_sanctions()

        # Interface de téléchargement des fichiers
        st.subheader("Étape 1: Téléchargement des fichiers")         
        # Zone de drag and drop unique pour tous les fichiers
        uploaded_files = st.file_uploader(
            "Glissez et déposez vos fichiers Excel (Account, Contact, Leads)",
            type=["xlsx"],
            accept_multiple_files=True,
            key="all_files"
        )
        upload_dir = r"\\panfs001\Syquant Compliance\PG06 - Politique LCB-FT\Screening\2. Liste de screening"       
        # Store uploaded files in session state
        if uploaded_files:
            st.session_state['uploaded_files'] = uploaded_files

        # Verification des fichiers telecharges
        if st.session_state.get('uploaded_files'):
            contact_file = next((f for f in st.session_state['uploaded_files'] if "contact" in f.name.lower()), None)
        else:
            contact_file_path = get_latest_file(upload_dir, "*contact*.xlsx")
            contact_file = contact_file_path if contact_file_path else None

        if st.session_state.get('uploaded_files'):
            account_file = next((f for f in st.session_state['uploaded_files'] if "account" in f.name.lower()), None)
        else:
            account_file_path = get_latest_file(upload_dir, "*account*.xlsx")
            account_file = account_file_path if account_file_path else None
      
        for file in uploaded_files:
            save_path = os.path.join(upload_dir, file.name)
            if not os.path.exists(save_path):
                with open(save_path, "wb") as f:
                    f.write(file.getbuffer())
                st.success(f"Les fichiers ont été enregistrés dans le dossier CRM ")
            else:
                st.info(f"Le fichier : {file.name} existe déjà dans le dossier" )

            # Attribution des fichiers selon leur nom
            if "account" in file.name.lower():
                account_file = file
            elif "contact" in file.name.lower():
                contact_file = file
            elif "leads" in file.name.lower():
                leads_file = file
                
            leads_file = None
            account_file = None
            contact_file = None                   
            # Vérifier si tous les fichiers requis sont présents
            if not all([account_file, contact_file, leads_file]):
                missing_files = []
                if not account_file:
                    missing_files.append("Account")
                if not contact_file:
                    missing_files.append("Contact")
                if not leads_file:
                    missing_files.append("Leads")

        
        # Étape 2: Lancement du screening
        st.subheader("Étape 2: Lancement du screening")

        today_str = dt.datetime.today().strftime('%Y%m%d')
        existing_files_for_today = check_for_existing_screening_files(results_dir, today_str)

        if st.session_state.screening_status == 'initial' or st.session_state.screening_status == 'display_results':
            with st.form("launch_screening_form"):
                launch_button_clicked = st.form_submit_button("▶️ Lancer le screening", help="Lance le processus complet de screening")
            
            if launch_button_clicked:
                st.session_state.screening_completed = False # Reset completion status for new run
                st.session_state.overwrite_confirmed = False # Reset for new click
                st.session_state.should_load_existing_results = False # Reset

                if existing_files_for_today:
                    st.session_state.screening_status = 'prompt_overwrite'
                else:
                    st.session_state.screening_status = 'running_new_pp'
                st.rerun()

        elif st.session_state.screening_status == 'prompt_overwrite':
            st.warning(f"⚠️ Un screening existe déjà pour aujourd'hui ({today_str}). Voulez-vous écraser les fichiers existants ?")
            col_overwrite1, col_overwrite2 = st.columns(2)
            with col_overwrite1:
                if st.button("Oui, écraser", key="overwrite_yes"):
                    st.session_state.overwrite_confirmed = True
                    st.session_state.screening_status = 'running_new_pp'
                    st.rerun()
            with col_overwrite2:
                if st.button("Non", key="overwrite_no"):
                    st.session_state.should_load_existing_results = True
                    st.session_state.screening_status = 'loading_existing'
                    st.rerun()

        # --- Main screening/loading logic ---
        # Check for uploaded files before proceeding with screening or loading
        current_uploaded_files = st.session_state.get('uploaded_files', [])
        
        if st.session_state.screening_status.startswith('running_new') or st.session_state.screening_status == 'loading_existing':
            # This block will be executed if the user initiated a screening (new or load)
            if not current_uploaded_files and st.session_state.screening_status != 'loading_existing':
                st.error("❌ Veuillez télécharger les fichiers Excel requis (Account, Contact, Leads) pour lancer un nouveau screening.")
                st.session_state.screening_status = 'initial' # Reset status if files are missing
                st.stop() # Stop execution to prevent further errors
            else:
                # Only define file paths if files are uploaded (for new screening)
                account_file_path = None
                contact_file_path = None
                leads_file_path = None

                if st.session_state.screening_status != 'loading_existing':
                    account_file_path = next((f for f in current_uploaded_files if "account" in f.name.lower()), None)
                    contact_file_path = next((f for f in current_uploaded_files if "contact" in f.name.lower()), None)
                    leads_file_path = next((f for f in current_uploaded_files if "leads" in f.name.lower()), None)

                    if not all([account_file_path, contact_file_path, leads_file_path]):
                        st.error("❌ Fichiers de rapport manquants pour le screening.")
                        st.session_state.screening_status = 'initial' # Reset status
                        st.stop() # Stop execution


                # Proceed with deletion if overwrite confirmed (only for new screening)
                if st.session_state.get('overwrite_confirmed', False):
                    deleted_files = delete_current_day_screening_files(results_dir, today_str)
                    if deleted_files:
                        st.success("Anciens fichiers supprimés, screening relancé ")
                    else:
                        st.info("Aucun fichier de screening existant trouvé à supprimer.")
                    st.session_state.overwrite_confirmed = False # Reset after deletion

                # Reset comment/status states for interactive display only when running a new screening
                if st.session_state.screening_status.startswith('running_new') and 'comments' in st.session_state:
                    for key in list(st.session_state.keys()):
                        if key.startswith(('comment_', 'status_')):
                            del st.session_state[key]
                    del st.session_state['comments'] # Ensure these are also deleted
                    del st.session_state['statuses'] # Ensure these are also deleted

                # Prepare progress bars (moved outside the individual screening blocks for clarity)
                progress_bar_pp = st.progress(0, text="Progression screening PP")
                status_text_pp = st.empty()
                progress_bar_pm = st.progress(0, text="Progression screening PM")
                status_text_pm = st.empty()

                # --- Screening Personnes Physiques ---
                if st.session_state.screening_status == 'running_new_pp':
                    with st.spinner("Screening des Personnes Physiques en cours..."):
                        all_sanctions = parse_sdn(file_sdn) + parse_nonsdn(file_nonsdn) + parse_fr(file_fr)
                        def pp_progress_callback(current_idx, total, stage_name):
                            if total > 0:
                                progress_bar_pp.progress(current_idx / total)
                                status_text_pp.text(f"Screening Personnes Physiques : {current_idx}/{total} analysées")
                            else:
                                progress_bar_pp.progress(0)
                                status_text_pp.text("Screening PP : aucune personne à analyser.")

                        results_df, abnormal_data = process_all_reports(
                            account_file_path, contact_file_path, leads_file_path,
                            all_sanctions, threshold_nom, threshold_prenom, 
                            pp_progress_callback
                        )
                        if "Commentaire" not in results_df.columns:
                            results_df["Commentaire"] = ""
                        if "Status" not in results_df.columns:
                            results_df["Status"] = ""

                        progress_bar_pp.progress(1.0)
                        status_text_pp.text("Screening PP terminé !")
                        st.session_state.screening_results = results_df
                        st.session_state.abnormal_data = abnormal_data

                        pp_history_df = load_previous_screening_data(results_dir, 'PP', today_str)
                        if pp_history_df is not None and not pp_history_df.empty:
                            pp_history_map = create_history_map(pp_history_df, is_pm=False)
                            if pp_history_map:
                                st.session_state.screening_results = apply_history_to_current_screening(st.session_state.screening_results, pp_history_map, is_pm=False)
                                st.info("Historique PP appliqué avec succès.")
                            else:
                                st.info("Aucun historique PP pertinent trouvé pour application.")
                        else:
                            st.info("Aucun fichier historique PP trouvé ou il est vide.")

                    st.session_state.screening_status = 'running_new_pm' # Move to PM screening
                    st.rerun() # Rerun to display PM progress

                # --- Screening Personnes Morales ---
                if st.session_state.screening_status == 'running_new_pm':
                    with st.spinner("Screening des Personnes Morales en cours..."):
                        iosco_entities = parse_iosco_pm_variants(file_iosco)
                        all_crm_companies_temp = []
                        all_crm_companies_temp.extend(extract_account_names(account_file_path))
                        all_crm_companies_temp.extend(extract_contact_account_names(contact_file_path))
                        all_crm_companies_temp.extend(extract_leads_company_names(leads_file_path))
                        seen = set()
                        unique_crm_companies = []
                        for d in all_crm_companies_temp:
                            name = d.get("name", "").strip().lower()
                            if name and name not in seen:
                                seen.add(name)
                                unique_crm_companies.append(d)
                        st.session_state.unique_crm_companies = unique_crm_companies
                        pm_threshold_value = st.session_state.get('pm_screening_threshold', 80)

                        def pm_progress_callback(current_idx, total, stage_name):
                            if total > 0:
                                progress_bar_pm.progress(current_idx / total)
                                status_text_pm.text(f"Screening PM : {current_idx}/{total} personnes morales analysées")
                            else:
                                progress_bar_pm.progress(0)
                                status_text_pm.text("Screening PM : aucune personne morale à analyser.")

                        df_pm_matches = screen_legal_entities(unique_crm_companies, iosco_entities, pm_threshold_value, 
                                                            progress_callback=pm_progress_callback)
                        if "Commentaire" not in df_pm_matches.columns:
                            df_pm_matches["Commentaire"] = ""
                        if "Status" not in df_pm_matches.columns:
                            df_pm_matches["Status"] = ""
                        if "Entity Name (CRM)" not in df_pm_matches.columns:
                            df_pm_matches["Entity Name (CRM)"] = ""

                        progress_bar_pm.progress(1.0)
                        status_text_pm.text("Screening PM terminé !")
                        st.session_state.pm_matches = df_pm_matches

                        pm_history_df = load_previous_screening_data(results_dir, 'PM', today_str)
                        if pm_history_df is not None and not pm_history_df.empty:
                            pm_history_map = create_history_map(pm_history_df, is_pm=True)
                            if pm_history_map:
                                st.session_state.pm_matches = apply_history_to_current_screening(st.session_state.pm_matches, pm_history_map, is_pm=True)
                                st.info("Historique PM appliqué avec succès.")
                            else:
                                st.info("Aucun historique PM pertinent trouvé pour application.")
                        else:
                            st.info("Aucun fichier historique PM trouvé ou il est vide.")

                    st.session_state.screening_completed = True # Mark screening as completed
                    st.session_state.screening_status = 'display_results' # Move to display results
                    st.rerun() # Rerun to display results and reset UI

                elif st.session_state.screening_status == 'loading_existing':
                    st.info(f"Chargement des résultats de screening existants pour aujourd'hui ({today_str})...")
                    try:
                        # Load PP results
                        output_file_pp_all = os.path.join(results_dir, f"{today_str} - screening des résultats PP.xlsx")
                        if os.path.exists(output_file_pp_all):
                            st.session_state.screening_results = pd.read_excel(output_file_pp_all, header=1)
                            st.session_state.screening_results.columns = st.session_state.screening_results.columns.str.strip()
                            st.success("Résultats PP existants chargés.")
                        else:
                            st.warning("Aucun fichier de résultats PP existant trouvé.")
                            st.session_state.screening_results = pd.DataFrame() # Initialize empty DF

                        # Load PM results
                        output_file_pm_all = os.path.join(results_dir, f"{today_str} - screening des résultats PM.xlsx")
                        if os.path.exists(output_file_pm_all):
                            st.session_state.pm_matches = pd.read_excel(output_file_pm_all, header=1)
                            st.session_state.pm_matches.columns = st.session_state.pm_matches.columns.str.strip()
                            st.success("Résultats PM existants chargés.")
                        else:
                            st.warning("Aucun fichier de résultats PM existant trouvé.")
                            st.session_state.pm_matches = pd.DataFrame() # Initialize empty DF

                        # Load all analyzed PM entities
                        output_file_pm_all = os.path.join(results_dir, f"{today_str} - screening des résultats PM.xlsx")
                        if os.path.exists(output_file_pm_all):
                            temp_df_all_entities = pd.read_excel(output_file_pm_all, header=0) # Read with header
                            if not temp_df_all_entities.empty:
                                # No need to manually assign columns if header=0 is used, just strip whitespace
                                temp_df_all_entities.columns = temp_df_all_entities.columns.str.strip()
                                # Explicitly convert relevant columns to string type
                                if 'name' in temp_df_all_entities.columns:
                                    temp_df_all_entities['name'] = temp_df_all_entities['name'].astype(str)
                                if 'source' in temp_df_all_entities.columns:
                                    temp_df_all_entities['source'] = temp_df_all_entities['source'].astype(str)
                                # Ensure we only take relevant columns if they exist
                                relevant_cols = [col for col in ['name', 'source', 'jurisdiction_account', 'company_type_account'] if col in temp_df_all_entities.columns]
                                if relevant_cols:
                                    st.session_state.unique_crm_companies = temp_df_all_entities[relevant_cols].to_dict(orient='records')
                                    st.success("Toutes les entités PM analysées existantes chargées.")
                                else:
                                    st.warning("Le fichier 'toutes entités PM' ne contient pas les colonnes attendues.")
                                    st.session_state.unique_crm_companies = []
                            else:
                                st.warning("Le fichier 'toutes entités PM' est vide.")
                                st.session_state.unique_crm_companies = []
                        else:
                            st.warning("Aucun fichier 'toutes entités PM' existant trouvé.")
                            st.session_state.unique_crm_companies = []

                        # Also load abnormal data if it exists
                        output_file_pp_abn = os.path.join(results_dir, f"{today_str} - screening des valeurs aberrantes PP.xlsx")
                        if os.path.exists(output_file_pp_abn):
                            st.session_state.abnormal_data = pd.read_excel(output_file_pp_abn, header=0) # Assuming this one has header=0
                            st.session_state.abnormal_data.columns = st.session_state.abnormal_data.columns.str.strip()
                        else:
                            st.session_state.abnormal_data = pd.DataFrame()

                        st.session_state.screening_completed = True # Mark as completed after loading
                        st.session_state.should_load_existing_results = False # Reset
                        st.session_state.screening_status = 'display_results' # Move to display results
                        st.rerun() # Rerun to display results

                    except Exception as e:
                        st.error(f"Erreur lors du chargement des fichiers de screening existants : {e}. Veuillez relancer un nouveau screening.")
                        st.session_state.screening_status = 'initial' # Reset status on error
                        st.rerun() # Rerun to clear error

        # --- Display results (this part should always be evaluated if screening_completed is True) ---
        if st.session_state.get('screening_completed', False) and st.session_state.screening_status == 'display_results' or st.session_state.screening_status == 'initial':
            results_df = st.session_state.get('screening_results')
            abnormal_data = st.session_state.get('abnormal_data')
            pm_results_df = st.session_state.get('pm_matches')
            unique_crm_companies = st.session_state.get('unique_crm_companies', [])
            today = dt.datetime.today().strftime('%Y%m%d')
            output_file_pp = os.path.join(results_dir, f"{today} - screening des correspondances PP.xlsx")
            output_file_pp_all = os.path.join(results_dir, f"{today} - screening des résultats PP.xlsx")
            output_file_pp_abn = os.path.join(results_dir, f"{today} - screening des valeurs aberrantes PP.xlsx")
            output_file_pm = os.path.join(results_dir, f"{today} - screening des correspondances PM.xlsx")
            output_file_pm_all = os.path.join(results_dir, f"{today} - screening des résultats PM.xlsx")

            # Filtrage des correspondances - MOVED AFTER HISTORY APPLICATION
            pp_matches = results_df[results_df['Screening Result'] != "OK - Aucune correspondance trouvée"].copy() if results_df is not None else pd.DataFrame()
            pm_matches = pm_results_df if pm_results_df is not None else pd.DataFrame()

            # ENREGISTREMENT AUTOMATIQUE
            # Perform automatic saving only if screening results are available
            if results_df is not None and isinstance(results_df, pd.DataFrame):
                print("Columns in results_df (PP) before saving:", results_df.columns.tolist())
                export_pp_excel_styled(pp_matches, output_file_pp)         # correspondances PP
                export_pp_excel_styled(results_df, output_file_pp_all)     # tous les résultats PP
                if abnormal_data is not None and isinstance(abnormal_data, pd.DataFrame) and not abnormal_data.empty:
                    abnormal_data.to_excel(output_file_pp_abn, index=False)
                else:
                    # Crée un fichier vide si pas de données aberrantes
                    pd.DataFrame().to_excel(output_file_pp_abn, index=False)

            if pm_results_df is not None and isinstance(pm_results_df, pd.DataFrame):
                print("Columns in pm_results_df (PM) before saving:", pm_results_df.columns.tolist())
                export_pm_excel_styled(pm_matches, output_file_pm)         # correspondances PM
                export_pm_excel_styled(pm_results_df, output_file_pm_all)  # tous les résultats PM

                # Automatic save for all analyzed PM entities
                if st.session_state.get('unique_crm_companies') is not None and not st.session_state.unique_crm_companies == []:
                    total_entities_df = pd.DataFrame(st.session_state.unique_crm_companies)
                    total_entities_df.to_excel(os.path.join(results_dir, f"{today} - screening toutes entités PM.xlsx"), index=False)
                else:
                    print("")

            # --- Affichage des résultats Personnes Physiques ---
            # Check if results_df is a valid DataFrame before processing/displaying PP results
            if results_df is not None and isinstance(results_df, pd.DataFrame) and not results_df.empty:
                st.subheader("📊 Résultats du screening Personnes Physiques")
                matches = results_df[results_df['Screening Result'] != "OK - Aucune correspondance trouvée"].copy()
                total_analyzed = len(results_df)
                total_matches = len(matches)
                sdn_count = matches[matches['Liste de sanctions'].str.contains("SDN", case=False, na=False)].shape[0]
                nonsdn_count = matches[matches['Liste de sanctions'].str.contains("Non-SDN", case=False, na=False)].shape[0]
                fr_count = matches[matches['Liste de sanctions'].str.contains("FR", case=False, na=False)].shape[0]

                col1, col2, col3, col4 = st.columns(4)
                col1.metric("👥 Individus analysés", total_analyzed)
                col2.metric("⚠️ Correspondances", total_matches)
                col3.metric("🗂️ SDN / Non-SDN", sdn_count + nonsdn_count)
                col4.metric("Gel des avoirs FR", fr_count)
                
                # Anneau taux de correspondance PP
                pp_match_count = total_matches
                pp_no_match_count = total_analyzed - pp_match_count

                pp_rate_labels = ['Correspondances', 'Sans correspondance']
                pp_rate_values = [pp_match_count, pp_no_match_count]
                pp_rate_colors = ['#e60000', '#05a33a']

                # Créer un graphique en anneau pour les statistiques
                col_chart1, col_chart2, col_chart3 = st.columns(3)
                 
                with col_chart1:
                    st.subheader("Taux de correspondance (PP)")
                    
                    # Données pour le graphique en anneau
                    pp_rate_labels = ['Correspondances', 'Sans correspondance']
                    pp_rate_values = [total_matches, total_analyzed - total_matches]
                    pp_rate_colors = ['#FF5733', '#05a33a']
                    
                    # Créer le graphique en anneau
                    fig = go.Figure(data=[go.Pie(
                        labels=pp_rate_labels,
                        values=pp_rate_values,
                        hole=.6,
                        marker_colors=pp_rate_colors,
                        textinfo='label', # Keep label and percentage
                        insidetextorientation='radial',
                        texttemplate='%{label}: %{percent:.2%}' # Format percentage to 2 decimal places
                    )])

                    # Personnaliser l'apparence
                    fig.update_layout(
                        showlegend=True,
                        legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
                        annotations=[dict(text=f'{total_analyzed}', x=0.5, y=0.5, font_size=20, showarrow=False)]
                    )
                    
                    st.plotly_chart(fig, use_container_width=True)
                               
                with col_chart2:
                    st.subheader("Répartition des correspondances (PP)")
                    
                    # Données pour le graphique en anneau
                    pp_rate_labels = ['SDN', 'Non-SDN', 'Gel des avoirs FR']
                    pp_rate_values = [sdn_count, nonsdn_count, fr_count]
                    pp_rate_colors = ['#FF9500', '#0068C9', '#83C9FF']
                    
                    # Créer le graphique en anneau
                    fig = go.Figure(data=[go.Pie(
                        labels=pp_rate_labels,
                        values=pp_rate_values,
                        hole=.6,
                        marker_colors=pp_rate_colors
                    )])
                    
                    # Personnaliser l'apparence
                    fig.update_layout(
                        showlegend=True,
                        legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
                        annotations=[dict(text=f'{total_matches}', x=0.5, y=0.5, font_size=20, showarrow=False)]
                    )
                    
                    st.plotly_chart(fig, use_container_width=True)

                with col_chart3:
                    st.subheader("Résultats de l'analyse (PP)")

                    # Compter les statuts dans la colonne "Status"
                    status_counts = matches["Status"].value_counts(dropna=False)
                    status_labels = status_counts.index.tolist()
                    status_values = status_counts.values.tolist()

                    # Définir des couleurs pour les statuts principaux
                    status_colors = []
                    for label in status_labels:
                        if label == "Résolu":
                            status_colors.append("#05a33a")  # Vert
                        elif label == "En cours d'investigation":
                            status_colors.append("#FF9500")  # Orange
                        elif label == "Sous sanction":
                            status_colors.append("#FF5733")  # Rouge
                        else:
                            status_colors.append("#808080")  # Gris pour les autres/vides

                    fig_status_pp = go.Figure(data=[go.Pie(
                        labels=status_labels,
                        values=status_values,
                        hole=.6,
                        marker_colors=status_colors,
                        textinfo='percent+label'
                    )])
                    fig_status_pp.update_layout(
                        showlegend=True,
                        legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
                        annotations=[dict(text=f'{sum(status_values)}', x=0.5, y=0.5, font_size=20, showarrow=False)]
                    )
                    st.plotly_chart(fig_status_pp, use_container_width=True, key="pp_status_chart")
                if total_matches > 0:
                    matches = display_matches_interactive(matches, type_pp_pm="PP")
                    merge_interactive_changes_back('screening_results', matches, is_pm=False)
                else:
                    st.info("✅ Aucune correspondance Personne Physique trouvée.")
                if abnormal_data is not None and not abnormal_data.empty:
                    st.subheader("⚠️ Valeurs aberrantes (Personnes Physiques)")
                    st.warning(f"{len(abnormal_data)} entrées identifiées comme aberrantes (noms/prénoms trop courts). Elles n'ont pas été incluses dans le screening.")
                    abnormal_display_columns = [
                        'First Name', 'Last Name', 'Type', 'Source', 'Title', 'Company/Account'
                    ]
                    abnormal_display = abnormal_data[[col for col in abnormal_display_columns if col in abnormal_data.columns]]
                    st.dataframe(abnormal_display, use_container_width=True)


                # --- Boutons de téléchargement SS2 pour PP ---
                st.subheader("🖨️ Téléchargement des résultats Personnes Physiques")
                today = dt.datetime.today().strftime('%Y%m%d')
                output_file_pp = os.path.join(results_dir, f"{today} - screening des correspondances PP.xlsx")
                output_file_pp_all = os.path.join(results_dir, f"{today} - screening des résultats PP.xlsx")
                output_file_pp_abn = os.path.join(results_dir, f"{today} - screening des valeurs aberrantes PP.xlsx")
                col_dl1, col_dl2, col_dl3 = st.columns(3)
                with col_dl1:
                    excel_buffer = io.BytesIO()
                    results_df.to_excel(excel_buffer, index=False)
                    excel_buffer.seek(0)
                    st.download_button(
                        "📥 Télécharger tous les résultats",
                        data=excel_buffer,
                        file_name=os.path.basename(output_file_pp_all),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                with col_dl2:
                    excel_buffer = io.BytesIO()
                    matches.to_excel(excel_buffer, index=False)
                    excel_buffer.seek(0)
                    st.download_button(
                        "📥 Télécharger uniquement les correspondances",
                        data=excel_buffer,
                        file_name=os.path.basename(output_file_pp),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                with col_dl3:
                    if abnormal_data is not None and not abnormal_data.empty:
                        excel_buffer = io.BytesIO()
                        abnormal_data.to_excel(excel_buffer, index=False)
                        excel_buffer.seek(0)
                        st.download_button(
                            "📥 Télécharger les valeurs aberrantes",
                            data=excel_buffer,
                            file_name=os.path.basename(output_file_pp_abn),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.info("Aucune donnée aberrante détectée, fichier vide généré.")

            # --- PERSONNES MORALES : Affichage des résultats Personnes Morales ---
            if pm_results_df is not None:
                st.subheader("📊 Résultats du screening Personnes Morales (IOSCO)")
                unique_crm_companies = st.session_state.get('unique_crm_companies', []) # Retrieve it just before use
                total_pm_analyzed = len(unique_crm_companies) if unique_crm_companies else 0
                total_pm_matches = len(pm_results_df)
                col_pm_stats1, col_pm_stats2 = st.columns(2)
                col_pm_stats1.metric("👥 Entités analysées", total_pm_analyzed)
                col_pm_stats2.metric("⚠️ Correspondances PM", total_pm_matches)
                
                # Create columns for the PM charts
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("#### Taux de correspondance (PM)")
                    # Anneau taux de correspondance PM
                    pm_match_count = total_pm_matches
                    pm_no_match_count = total_pm_analyzed - pm_match_count

                    pm_rate_labels = ['Correspondances PM', 'Sans correspondance PM']
                    pm_rate_values = [pm_match_count, pm_no_match_count]
                    pm_rate_colors = ['#e60000', '#05a33a']

                    fig_pm_rate = go.Figure(data=[go.Pie(
                            labels=pm_rate_labels,
                            values=pm_rate_values,
                        hole=.6,
                        marker_colors=pm_rate_colors,
                        textinfo='label+percent',
                        insidetextorientation='radial',
                    )])
                    fig_pm_rate.update_layout(
                        showlegend=True,
                        legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
                        annotations=[dict(text=f'{total_pm_analyzed}', x=0.5, y=0.5, font_size=20, showarrow=False)]
                    )
                    st.plotly_chart(fig_pm_rate, use_container_width=True, key="pm_rate_chart") # Changed key to be unique for PM rate chart
                with col2:
                    st.markdown("#### Répartition des statuts (PM)")
                                # Compter les statuts dans la colonne "Status" du DataFrame pm_results_df
                    status_counts = pm_results_df["Status"].value_counts(dropna=False)
                    status_labels = status_counts.index.tolist()
                    status_values = status_counts.values.tolist()

                        # Définir des couleurs pour les statuts principaux
                    status_colors = []
                    for label in status_labels:
                            if label == "Résolu":
                                status_colors.append("#05a33a")  # Vert
                            elif label == "En cours d'investigation":
                                status_colors.append("#FF9500")  # Orange
                            elif label == "Sous sanction":
                                status_colors.append("#FF5733")  # Rouge
                            else:
                                status_colors.append("#808080")  # Gris pour les autres/vides
             
                    fig_status_pm = go.Figure(data=[go.Pie(
                            labels=status_labels,
                            values=status_values,
                            hole=.6,
                            marker_colors=status_colors
                        )])
                    fig_status_pm.update_layout(
                            showlegend=True,
                            legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
                            annotations=[dict(text=f'{sum(status_values)}', x=0.5, y=0.5, font_size=20, showarrow=False)]
                        )

                    st.plotly_chart(fig_status_pm, use_container_width=True, key="suivi_chart_status_analyzed")

                with st.container():
                    if total_pm_matches > 0:
                        pm_results_df = display_matches_interactive(pm_results_df, type_pp_pm="PM")
                        merge_interactive_changes_back('pm_matches', pm_results_df, is_pm=True)
                    else:
                        st.info("✅ Aucune correspondance Personne Morale trouvée.")

                    st.subheader("🖨️ Téléchargement des résultats Personnes Morales")
                    col_dlpm1, col_dlpm2 = st.columns(2)
                                # --- Boutons de téléchargement SS2 pour PM ---
                    
                    today = dt.datetime.today().strftime('%Y%m%d')
                    output_file_pm = os.path.join(results_dir, f"{today} - screening des correspondances PM.xlsx")
                    with col_dlpm1:
                        excel_buffer = io.BytesIO()
                        pm_results_df.to_excel(excel_buffer, index=False)
                        excel_buffer.seek(0)
                        st.download_button(
                            "📥 Télécharger toutes les correspondances PM",
                            data=excel_buffer,
                            file_name=os.path.basename(output_file_pm),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    with col_dlpm2:
                        excel_buffer = io.BytesIO()
                        pd.DataFrame(unique_crm_companies).to_excel(excel_buffer, index=False)
                        excel_buffer.seek(0)
                        st.download_button(
                            "📥 Télécharger toutes les entités analysées",
                            data=excel_buffer,
                            file_name=f"{today} - screening toutes entités PM.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                # --- Bouton pour générer l'email global (PM + PP) ---
                today = dt.datetime.today().strftime('%Y%m%d')
                output_file_pm = os.path.join(results_dir, f"{today} - screening des correspondances PM.xlsx")
                output_file_pp = os.path.join(results_dir, f"{today} - screening des correspondances PP.xlsx")

                st.subheader("📧 E-Mail de Contrôle")
                if st.button("📧 Générer l'email Outlook ", key="generate_email_both"):
                    with st.spinner("Génération de l'email global..."):
                        # On filtre pour ne garder que les correspondances
                        pm_matches = pm_results_df if pm_results_df is not None else None
                        pp_matches = results_df[results_df['Screening Result'] != "OK - Aucune correspondance trouvée"] if results_df is not None else None
                        if (pm_matches is None or pm_matches.empty) and (pp_matches is None or pp_matches.empty):
                            st.info("Aucune correspondance trouvée, aucun mail généré.")
                        else:
                            # Juste avant d'appeler generate_outlook_email_both
                            output_file_pm_all = os.path.join(results_dir, f"{today} - screening des résultats PM.xlsx")
                            output_file_pp_all = os.path.join(results_dir, f"{today} - screening des résultats PP.xlsx") # Corrected to output_file_pp_all

                            # Sauvegarder toutes les lignes (pas seulement les correspondances)
                            export_pm_excel_styled(pm_results_df, output_file_pm_all) # Save full PM results
                            export_pp_excel_styled(results_df, output_file_pp_all) # Save full PP results
                            # Appeler la fonction d'email avec ces fichiers
                            generate_outlook_email_both(
                                pm_matches if pm_matches is not None else pd.DataFrame(),
                                pp_matches if pp_matches is not None else pd.DataFrame(),
                                output_file_pm, # Use the matches file for email attachment
                                output_file_pp # Use the matches file for email attachment
                            )

    # ------------------------------------- SUIVI DES ANALYSES LCBFT -------------------------------------------
    # ------------------------------------- SUIVI DES ANALYSES LCBFT -------------------------------------------
    with tab2:
        st.header("📊 Suivi des analyses")
        
        # Charger l'historique
        history_df = load_screening_history(results_dir)
        
        if not history_df.empty:
            # Graphique 1: Évolution par type
            st.subheader("1. Évolution du nombre d'individus par type")

            type_data = []
            for _, row in history_df.iterrows():
                for type_name, count in row['Types'].items():
                    type_data.append({
                        'Date': row['Date'],
                        'Type': type_name,
                        'Nombre': count
                    })

            type_df = pd.DataFrame(type_data)
            type_df["Date"] = pd.to_datetime(type_df["Date"])  # ✅ Conversion
            type_df = type_df.sort_values("Date")              # ✅ Tri

            fig_types = px.line(type_df, x='Date', y='Nombre', color='Type', markers=True)

            fig_types.update_layout(
                xaxis_title="Date",
                yaxis_title="Nombre d'individus",
                legend_title="Type",
                legend=dict(
                    orientation="h",
                    yanchor="bottom",
                    y=-0.3,
                    xanchor="center",
                    x=0.5
                ),
                xaxis=dict(
                    range=[type_df["Date"].min(), type_df["Date"].max()],  # ✅ Afficher toutes les dates
                    tickformat="%d %b"
                )
            )

            st.plotly_chart(fig_types, use_container_width=True, key="suivi_chart_types_1")
#------------------------------------------------------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------------------             
            # Graphique 2: Évolution des correspondances
            st.subheader("2.Évolution des correspondances (PP et PM)")

            df_melted = history_df.melt(
                id_vars="Date",
                value_vars=["Correspondances PP", "Correspondances PM"],
                var_name="Type",
                value_name="Correspondances"
            )

            fig1 = px.line(
                df_melted,
                x="Date",
                y="Correspondances",
                color="Type",
                markers=True,
            )

            fig1.update_layout(
                xaxis_title="Date",
                yaxis_title="Nombre de correspondances",
                legend_title="Type",
                legend=dict(orientation="h", x=0.5, xanchor="center", y=-0.3, yanchor="bottom")
            )

            st.plotly_chart(fig1, use_container_width=True)

#---------------------------------------------            # Graphique 3: Evolution des Résultats du screening-------------------------------------------------------

# --- GRAPH 2 : Résultats du screening par statut + type (PP vs PM) ---
            st.subheader(" 3.Evolution des Résultats du screening ")
            # Étape 1 : Consolider les données
            status_data = []
            for _, row in history_df.iterrows():
                for status_key, count in row.get("Statuts", {}).items():
                    if "(" in status_key:
                        statut, origine = status_key.rsplit("(", 1)
                        statut = statut.strip()
                        origine = origine.replace(")", "").strip()
                    else:
                        statut = status_key.strip()
                        origine = "Inconnu"
                    if statut and origine and count > 0:
                        status_data.append({
                            "Date": pd.to_datetime(row["Date"]).date(),
                            "Statut": statut,
                            "Origine": origine,
                            "Nombre": count,
                            "Statut+Origine": f"{statut} ({origine})"
                        })

            # Pour les dates antérieures où il n'y a pas de statut défini, ajouter manuellement
            # des entrées avec statut "Résolu" pour les correspondances PP
            for _, row in history_df.iterrows():
                date = pd.to_datetime(row["Date"]).date()
                pp_count = row["Correspondances PP"]
                pm_count = row["Correspondances PM"]
                
                # Si aucun statut n'a été défini pour cette date mais il y a des correspondances
                if not any(item["Date"] == date for item in status_data) and (pp_count > 0 or pm_count > 0):
                    # Ajouter des entrées par défaut
                    if pp_count > 0:
                        status_data.append({
                            "Date": date,
                            "Statut": "Résolu",
                            "Origine": "PP",
                            "Nombre": pp_count,
                            "Statut+Origine": "Résolu (PP)"
                        })
                    if pm_count > 0:
                        status_data.append({
                            "Date": date,
                            "Statut": "Résolu",
                            "Origine": "PM",
                            "Nombre": pm_count,
                            "Statut+Origine": "Résolu (PM)"
                        })

            status_df = pd.DataFrame(status_data)

            # Convertir la colonne Date en datetime si ce n'est pas déjà fait
            status_df["Date"] = pd.to_datetime(status_df["Date"])

            # Étape 2 : Couleurs personnalisées par statut + origine
            couleurs_personnalisees = {
                "Résolu (PP)": "#4CAF50",        # Vert clair
                "Résolu (PM)": "#087f23",        # Vert foncé
                "En cours d'investigation (PP)": "#FFC107",  # Jaune clair
                "En cours d'investigation (PM)": "#FF8F00",  # Jaune foncé
                "Sous sanction (PP)": "#EF5350", # Rouge clair
                "Sous sanction (PM)": "#B71C1C", # Rouge foncé
            }

            # Étape 3 : Forcer l'ordre d'apparition dans les groupes (PM avant PP)
            ordre_personnalise = [
                "Résolu (PP)","Résolu (PM)", 
                "En cours d'investigation (PM)", "En cours d'investigation (PP)",
                "Sous sanction (PM)", "Sous sanction (PP)"
            ]

            # Vérifier si le DataFrame n'est pas vide
            if not status_df.empty:
                # Filtrer l'ordre pour n'inclure que les catégories présentes dans les données
                categories_existantes = [cat for cat in ordre_personnalise if cat in status_df["Statut+Origine"].unique()]
                
                if categories_existantes:  # S'assurer qu'il y a des catégories à afficher
                    status_df["Statut+Origine"] = pd.Categorical(
                        status_df["Statut+Origine"],
                        categories=categories_existantes,
                        ordered=True
                    )
                
                    # Étape 4 : Générer le graphique Plotly (barmode = group)
                    fig = px.bar(
                        status_df,
                        x="Date",
                        y="Nombre",
                        color="Statut+Origine",
                        barmode="group",  # ✅ Groupé : barres côte à côte
                        color_discrete_map={k: v for k, v in couleurs_personnalisees.items() if k in categories_existantes},
                        labels={
                            "Date": "Date",
                            "Nombre": "Nombre d'individus",
                            "Statut+Origine": "Statut / Origine"
                        },
                    )
                    
                    # Définir les limites de l'axe des x pour exclure le 10 mai
                    # Trouver la date minimale et maximale dans les données
                    min_date = status_df["Date"].min()
                    max_date = status_df["Date"].max()
                    
                    # Si le 10 mai est présent, définir la date minimale au 11 mai
                    if pd.Timestamp(2023, 5, 10).date() in status_df["Date"].dt.date.unique():
                        min_date = pd.Timestamp(2023, 5, 11)
                    # Calculer la fenêtre initiale de 5 jours (à partir de la date maximale)
                    window_start = max_date - pd.Timedelta(days=5)
                    if window_start < min_date:
                        window_start = min_date
                    
                    fig.update_layout(
                        xaxis_title="Date",
                        yaxis=dict(range=[0, 40]),
                        bargap=0.25,
                        legend_title="Statut / Type",
                        legend=dict(
                            orientation="h",
                            yanchor="bottom",
                            y=-0.3,
                            xanchor="center",
                            x=0.6,
                        ),
                        # Définir les limites de l'axe des x et masquer les weekends
                        xaxis=dict(
                            # Afficher initialement les 5 derniers jours
                            range=[max_date - pd.Timedelta(days=30), max_date],
                            # Masquer les weekends
                            rangebreaks=[
                                dict(bounds=["sat", "mon"])  # Masquer samedi et dimanche
                            ],
                            # Personnaliser le format des dates
                            tickformat="%d %b",
                            # Afficher tous les jours
                            dtick="D1"
                        ),
                        # Ajouter des boutons de navigation et de zoom
                        updatemenus=[
                            dict(
                                type="buttons",
                                direction="right",
                                x=0.1,
                                y=1.1,
                                showactive=False,
                                buttons=[
                                    dict(
                                        label="Tout afficher",
                                        method="relayout",
                                        args=[{"xaxis.range": [min_date, max_date]}]
                                    ),
                                    dict(
                                        label="5 derniers jours",
                                        method="relayout",
                                        args=[{"xaxis.range": [window_start, max_date]}]
                                    ),
                                    dict(
                                        label="Mois précédent",
                                        method="relayout",
                                        args=[{"xaxis.range": [max_date - pd.Timedelta(days=30), max_date]}]
                                    ),
                                ]
                            )
                        ]
                    )
                    
                    # Activer les outils de zoom et de navigation
                    config = {
                        'scrollZoom': True,
                        'displayModeBar': True,
                        'modeBarButtonsToAdd': ['select2d', 'lasso2d'],
                        'modeBarButtonsToRemove': ['autoScale2d']
                    }
                    
                    st.plotly_chart(fig, use_container_width=True, config=config, height=300)
                else:
                    st.info("Aucune catégorie à afficher dans le graphique.")
            else:
                st.info("Aucune donnée disponible pour générer ce graphique.")


        #__Afficher le tableau récapitulatif__

            st.subheader("4.Tableau récapitulatif")
            if st.checkbox("Afficher le tableau récapitulatif des correspondances", value=False):
                if not history_df.empty:
                    df_sorted = history_df.sort_values("Date", ascending=False)  
                    st.dataframe(df_sorted[["Date", "Correspondances PP", "Correspondances PM", "Total"]])
                else:
                    st.info("Aucune donnée d'historique disponible.")
        
        palette = px.colors.qualitative.Pastel + px.colors.qualitative.Set2
        # ... Visualiser les  Screening Precedents ...

        RESULTS_DIR = r"\\panfs001\Syquant Compliance\PG06 - Politique LCB-FT\Screening\3. Résultats du screening"


        df_history = list_screening_history(RESULTS_DIR)
        dates_disponibles = sorted(df_history[~df_history['missing']]['date'].unique(), reverse=True)

        # 2. Sélecteur de date
        st.subheader(f"5.Résultats détaillés des Screening")
        if st.checkbox("Afficher le Résultats détaillés des Screening", value=False):
            date_selection = st.date_input("Choisir une date de screening :", value=max(dates_disponibles), min_value=min(dates_disponibles), max_value=max(dates_disponibles))
            if date_selection:  
                files = os.listdir(RESULTS_DIR)
                date_str = date_selection.strftime('%Y%m%d')
                files_pp = [f for f in files if date_str in f and 'PP' in f and f.endswith('.xlsx')]
                files_pm = [f for f in files if date_str in f and 'PM' in f and f.endswith('.xlsx')]
                file_correspondance = [f for f in files if date_str in f and 'correspondance' in f.lower() and f.endswith('.xlsx')]

                st.markdown("### 1- Personnes physiques (PP)")
                if files_pp:
                        df_pp = pd.read_excel(os.path.join(RESULTS_DIR, files_pp[0]), header=1)
                        st.dataframe(df_pp, use_container_width=True, hide_index=True)
                elif file_correspondance:
                        df_corr = pd.read_excel(os.path.join(RESULTS_DIR, file_correspondance[0]))
                        st.dataframe(df_corr, use_container_width=True, hide_index=True)
                else:
                        st.info("Aucune correspondance PP trouvée pour cette date.")

                st.markdown("### 2- Personnes morales (PM)") 
                if files_pm:
                        df_pm = pd.read_excel(os.path.join(RESULTS_DIR, files_pm[0]), header=1)
                        st.dataframe(df_pm, use_container_width=True, hide_index=True)
                else:
                        st.info("Aucune correspondance PM trouvée pour cette date.")
#______________________________________________________Statistiques des CRM_______________________________________________________
        leads_file = get_latest_file(upload_dir, "*leads*.xlsx")
        contact_file = get_latest_file(upload_dir, "*contact*.xlsx")
        account_file = get_latest_file(upload_dir, "*account*.xlsx")

        st.subheader("🟣 Distribution des clients -Leads")
        try:
            df_leads = pd.read_excel(leads_file, header=12)
            df_leads.columns = df_leads.columns.str.strip()
            lead_owner_col = "Lead Owner" if "Lead Owner" in df_leads.columns else df_leads.columns[11]  # souvent colonne L
            df_leads = df_leads[df_leads[lead_owner_col].notna() & (df_leads[lead_owner_col].astype(str).str.strip() != "")]
            owners_leads = df_leads[lead_owner_col].value_counts().reset_index()
            owners_leads.columns = ["Lead Owner", "Nombre de leads"]
            fig_leads = px.bar(
                owners_leads,
                x='Lead Owner',
                y='Nombre de leads',
                color='Lead Owner',
                color_discrete_sequence=palette,
                title="Distribution des clients par 'Lead Owner'",
                text='Nombre de leads'
            )
            fig_leads.update_traces(textposition='outside')
            fig_leads.update_layout(
                xaxis_title="Lead Owner",
                yaxis_title="Nombre de leads",
                showlegend=False,
                bargap=0.25
            )
            st.plotly_chart(fig_leads, use_container_width=True)
            st.dataframe(owners_leads, use_container_width=True, hide_index=True)
        except Exception as e:
            st.warning(f"Veuillez charger le fichier Leads dans le screening")

        st.markdown("---")

        # Analyse CONTACTS
        st.subheader("🔵 Distribution des clients -Contact")
        try:
            df_contact = pd.read_excel(contact_file, header=8)
            df_contact.columns = df_contact.columns.str.strip()
            contact_owner_col = "Contact Owner" if "Contact Owner" in df_contact.columns else df_contact.columns[17]  # souvent colonne R
            df_contact = df_contact[df_contact[contact_owner_col].notna() & (df_contact[contact_owner_col].astype(str).str.strip() != "")]
            owners_contacts = df_contact[contact_owner_col].value_counts().reset_index()
            owners_contacts.columns = ["Contact Owner", "Nombre de contacts"]
            fig_contacts = px.bar(
                owners_contacts,
                x='Contact Owner',
                y='Nombre de contacts',
                color='Contact Owner',
                color_discrete_sequence=palette,
                title="Distribution des clients par 'Contact Owner'",
                text='Nombre de contacts'
            )
            fig_contacts.update_traces(textposition='outside')
            fig_contacts.update_layout(
                xaxis_title="Contact Owner",
                yaxis_title="Nombre de contacts",
                showlegend=False,
                bargap=0.25
            )
            st.plotly_chart(fig_contacts, use_container_width=True)
            st.dataframe(owners_contacts, use_container_width=True, hide_index=True)
        except Exception as e:
            st.warning(f"Veuillez charger le fichier Contacts dans le screening")

        st.markdown("---")

        # Analyse ACCOUNT
        st.subheader("🟢 Distribution des clients -Account")
        try:
            df_account = pd.read_excel(account_file, header=8)
            df_account.columns = df_account.columns.str.strip()
            account_owner_col = "Account Owner" if "Account Owner" in df_account.columns else df_account.columns[28]  # colonne AC
            df_account = df_account[df_account[account_owner_col].notna() & (df_account[account_owner_col].astype(str).str.strip() != "")]
            owners_accounts = df_account[account_owner_col].value_counts().reset_index()
            owners_accounts.columns = ["Account Owner", "Nombre de comptes"]
            fig_accounts = px.bar(
                owners_accounts,
                x='Account Owner',
                y='Nombre de comptes',
                color='Account Owner',
                color_discrete_sequence=palette,
                title="Distribution des clients par 'Account Owner'",
                text='Nombre de comptes'
            )
            fig_accounts.update_traces(textposition='outside')
            fig_accounts.update_layout(
                xaxis_title="Account Owner",
                yaxis_title="Nombre de comptes",
                showlegend=False,
                bargap=0.25
            )
            st.plotly_chart(fig_accounts, use_container_width=True)
            st.dataframe(owners_accounts, use_container_width=True, hide_index=True)
        except Exception as e:
            st.warning(f"Veuillez charger le fichier Account dans le screening")
        
                # --- COMPARAISON ENTRE DEUX DATES : COMPTES AJOUTÉS/SUPPRIMÉS ---
        st.markdown("_______________")
        st.subheader("6.Suivi des évolutions des comptes")
        account_files = glob.glob(os.path.join(upload_dir, 'Report Account Olley-*.xlsx'))
        account_files = sorted(account_files)

        import re
        from datetime import datetime

        def extract_date_from_filename(file_path):
            basename = os.path.basename(file_path)
            match = re.search(r'(\d{4})-(\d{2})-(\d{2})', basename)
            if match:
                try:
                    return datetime.strptime(match.group(0), "%Y-%m-%d").date()  # <-- .date() ici
                except Exception:
                    return None
            return None

        files_and_dates = [(f, extract_date_from_filename(f)) for f in account_files]
        files_and_dates = [(f, d) for f, d in files_and_dates if d is not None]
        files_and_dates.sort(key=lambda x: x[1])

        dates_sorted = [d for f, d in files_and_dates]
        if len(dates_sorted) < 2:
            st.info("Il faut au moins deux fichiers Account pour comparer les évolutions.")
        else:
            min_date, max_date = dates_sorted[0], dates_sorted[-1]
            col1, col2 = st.columns(2)
            with col1:
                date_debut = st.date_input("Date de début", min_value=min_date, max_value=max_date, value=min_date, key="date_debut")
            with col2:
                date_fin = st.date_input("Date de fin", min_value=min_date, max_value=max_date, value=max_date, key="date_fin")

            # Filtrer les fichiers dans l'intervalle
            files_interval = [(f, d) for f, d in files_and_dates if date_debut <= d <= date_fin]
            if len(files_interval) < 2:
                st.warning("Sélectionnez un intervalle contenant au moins deux dates différentes.")
            else:
                # --- Graphe évolution (AVANT les tableaux) ---
                ajouts_par_date = []
                suppressions_par_date = []
                labels = []

                for i in range(1, len(files_interval)):
                    df_prev = load_account_report(files_interval[i-1][0])
                    df_curr = load_account_report(files_interval[i][0])
                    label = files_interval[i][1].strftime("%Y-%m-%d")
                    labels.append(label)
                    if df_prev is not None and df_curr is not None and 'Account Name' in df_prev.columns and 'Account Name' in df_curr.columns:
                        ajoutes = df_curr[~df_curr['Account Name'].isin(df_prev['Account Name'])]
                        supprimes = df_prev[~df_prev['Account Name'].isin(df_curr['Account Name'])]
                        ajouts_par_date.append(len(ajoutes))
                        suppressions_par_date.append(len(supprimes))
                    else:
                        ajouts_par_date.append(0)
                        suppressions_par_date.append(0)

                fig = go.Figure()
                fig.add_trace(go.Bar(
                    x=labels,
                    y=ajouts_par_date,
                    name="Comptes ajoutés",
                    marker_color='green',
                    customdata=ajouts_par_date,
                    hovertemplate='%{customdata} personnes<extra></extra>'
                ))
                fig.add_trace(go.Bar(
                    x=labels,
                    y=[-x for x in suppressions_par_date],
                    name="Comptes supprimés",
                    marker_color='red',
                    customdata=suppressions_par_date,
                    hovertemplate='%{customdata} personnes<extra></extra>'
                ))
                fig.add_hline(y=0, line_width=2, line_dash="solid", line_color="black")
                fig.update_layout(
                    title="Évolution des comptes (ajouts/suppressions) par date",
                    xaxis_title="Date",
                    yaxis_title="Nombre de comptes",
                    legend=dict(x=0.01, y=0.99),
                    template="simple_white",
                    barmode='group',  # ou 'relative' selon ton besoin
                )
                st.plotly_chart(fig, use_container_width=True)

                # --- Sélection de deux dates précises pour afficher les tableaux ---
                st.markdown("### Détail des ajouts/suppressions entre deux dates")
                show_details = st.checkbox("Afficher le détail entre deux dates",value=False)

                if show_details:
                    # Proposer la sélection de deux dates dans l'intervalle
                    options = [d for f, d in files_interval]
                    min_opt, max_opt = options[0], options[-1]
                    col_date1, col_date2 = st.columns(2)
                    with col_date1:
                        date1 = st.date_input("Sélectionnez une date de début", min_value=min_opt, max_value=max_opt, value=min_opt, key="date1")
                    with col_date2:
                        date2 = st.date_input("Sélectionnez une date de fin", min_value=min_opt, max_value=max_opt, value=max_opt, key="date2")

                    if date1 != date2:
                        # Trouver les index correspondants
                        try:
                            i1 = options.index(date1)
                            i2 = options.index(date2)
                        except ValueError:
                            st.warning("Dates sélectionnées non valides.")
                            i1, i2 = None, None

                        if i1 is not None and i2 is not None:
                            file1, file2 = files_interval[i1][0], files_interval[i2][0]
                            df1 = load_account_report(file1)
                            df2 = load_account_report(file2)
                            if df1 is not None and df2 is not None and 'Account Name' in df1.columns and 'Account Name' in df2.columns:
                                ajoutes = df2[~df2['Account Name'].isin(df1['Account Name'])]
                                supprimes = df1[~df1['Account Name'].isin(df2['Account Name'])]
                                for col in ['Unnamed: 0', 'Unnamed: 2', 'A']:
                                    if col in ajoutes.columns:
                                        ajoutes = ajoutes.drop(columns=[col])
                                    if col in supprimes.columns:
                                        supprimes = supprimes.drop(columns=[col])
                                st.markdown(f"**🟢 Comptes ajoutés entre {date1} et {date2} : {len(ajoutes)}**")
                                if not ajoutes.empty:
                                    st.dataframe(ajoutes, use_container_width=True, hide_index=True)
                                else:
                                    st.success("Aucun compte ajouté sur la période.")
                                st.markdown(f"**🔴 Comptes supprimés entre {date1} et {date2} : {len(supprimes)}**")
                                if not supprimes.empty:
                                    st.dataframe(supprimes, use_container_width=True, hide_index=True)
                                else:
                                    st.success("Aucun compte supprimé sur la période.")
                            else:
                                st.warning("Impossible de lire les deux fichiers Account ou la colonne 'Account Name' est manquante.")
                    else:
                        st.info("Veuillez sélectionner deux dates différentes pour la comparaison.")   


#--------------------------------------------    ANALYSE DU FICHIER ACCOUNT (CRM) ---------------------------------------------------------

    with tab3:
        st.header("📈 Analyse des comptes")

        account_file = get_latest_file(upload_dir, "*account*.xlsx")
        # Interface de téléchargement du fichier Account Report
        uploaded_file = st.file_uploader(
            "Téléchargez le fichier Report Account Olley",
            type=["xlsx"],
            key="account_report"
        )
    
        if uploaded_file:
            df = load_account_report(uploaded_file)       
        else:
            df = load_account_report(account_file)   

            if df is not None:
                # Afficher les colonnes disponibles
                st.subheader("📊 Statistiques globales")
                total_accounts = len(df)

                # Compter les pays et juridictions si les colonnes existent
                total_countries = df['Billing Country (text only)'].nunique() if 'Billing Country (text only)' in df.columns else 0
                total_juridictions = df['Jurisdiction'].nunique() if 'Jurisdiction' in df.columns else 0

                col_stats1, col_stats2, col_stats3 = st.columns(3)
                col_stats1.metric("Nombre total de comptes", total_accounts)
                if total_countries > 0:
                    col_stats2.metric("Nombre de pays", total_countries)
                if total_juridictions > 0:
                    col_stats3.metric("Nombre de juridictions", total_juridictions)

                # Créer des colonnes pour les graphiques
                col1, col2 = st.columns(2)

                # List of columns to analyze with their titles and whether to show initials
                columns_to_analyze_crm = [
                    ('Account Name', 'Distribution des comptes'),
                    ('Billing Country (text only)', 'Distribution par pays de facturation'),
                    ('Jurisdiction', 'Distribution par juridiction'),
                    ('Rebate agreement', 'Distribution des accords de remise'),
                    ('Rebate status', 'Conventions de distributions'),
                    ('Relationship', 'Distribution des types de relation'),
                    ('Company type', 'Distribution des types de société'),
                    ('Source of regulation', 'Distribution des sources de régulation'), # Show initials for this one
                    ('Vigilance', 'Distribution des niveaux de vigilance'),
                    ('Company activity', 'Distribution des activités'),
                    ('MIFID II Classification', 'Classification MIFID II')
                ]

                # Create charts for each column
                for i, (column, title) in enumerate(columns_to_analyze_crm):
                    # --- Start of modification ---
                    df_to_pass = df # Use the original df by default

                    # If the column is 'Rebate status', filter out rows with missing values in this column
                    if column == 'Rebate status' and df is not None:
                         # Filter out None, NaN, and empty strings for this specific column
                         # Use .loc to avoid SettingWithCopyWarning if the original df is a slice
                         df_to_pass = df.loc[df[column].notna() & (df[column].astype(str).str.strip() != 'n.a.')].copy()
                    # --- End of modification ---

                    # Pass the potentially filtered DataFrame to create_dashboard
                    fig, table = create_dashboard(df_to_pass, column, title) # Use df_to_pass here
                    if fig is not None and table is not None:
                        # Place 'Source of regulation' in a specific spot if desired, otherwise use columns
                        # For now, let's just use the columns
                        col = col1 if i % 2 == 0 else col2
                        with col:
                            st.plotly_chart(fig, use_container_width=True, key=f"account_chart_{column}_{i}")
                            show_table = st.checkbox(f"Afficher le tableau pour {title}", value=False, key=f"show_table_{column}_{i}")
                            if show_table:
                                # Format the table for better readability
                                styled_table = table.style.format({'Pourcentage': '{:.2f} %'})
                                st.dataframe(styled_table, use_container_width=True, hide_index=True, height=min(len(table) * 35 + 30, 400)) # Adjust height based on rows


                # Analysis of inconsistencies - Keep this section as is, it's text/tables
                st.markdown ('___________')
                st.subheader("🔍 Analyse des incohérences")

                # Create three columns for the inconsistency tables
                col_inc1, col_inc2 = st.columns(2)

                with col_inc1:
                    st.markdown("#### Différences Pays/Juridiction")
                    country_jurisdiction_mismatch, missing_info, active_rebates = analyze_inconsistencies(df)
                    
                    if not country_jurisdiction_mismatch.empty:
                        st.dataframe(
                            country_jurisdiction_mismatch,
                            use_container_width=True,
                            hide_index=True
                        )
                        st.markdown(f"**Nombre d'incohérences entre le pays de facturation et la juridiction :** {len(country_jurisdiction_mismatch)}")
                    else:
                        st.success("✅ Aucune incohérence trouvée entre les pays de facturation et les juridictions")
                
                with col_inc2:
                    st.markdown("#### Business Relationships incomplètes")
                    if not missing_info.empty:
                        st.dataframe(
                            missing_info,
                            use_container_width=True,
                            hide_index=True
                        )
                        st.markdown(f"**Nombre de Business Relationships incomplètes :** {len(missing_info)}")
                    else:
                        st.success("✅ Toutes les Business Relationships sont complètes")
                
                # Ajouter le tableau des contrat de distribution actifs
                st.markdown ('___________')
                st.markdown("#### Contrat de distribution actifs")
                if not active_rebates.empty:
                    st.dataframe(
                        active_rebates,
                        use_container_width=True,
                        hide_index=True
                    )
                    st.markdown(f"**Nombre de contrat de distribution actifs :** {len(active_rebates)}")
                else:
                    st.success("✅ Aucun contrat de distribution actif trouvé")
                
                # Ajouter un résumé des incohérences
                total_inconsistencies = len(country_jurisdiction_mismatch) + len(missing_info)
                if total_inconsistencies > 0:
                    st.warning(f"⚠️ **Total des incohérences détectées :** {total_inconsistencies}")
                else:
                    st.success("✅ Aucune incohérence détectée dans les données")
                
                st.markdown("---")  # Séparateur
                # -- Stocke tous les tableaux et images générés lors de l'analyse
                df_dict, img_paths = {}, {}

                # Exemple de statistiques globales
                stats_glob_df = pd.DataFrame([
                    {"Statistique": "Nombre total de comptes", "Valeur": total_accounts},
                    {"Statistique": "Nombre de pays", "Valeur": total_countries},
                    {"Statistique": "Nombre de juridictions", "Valeur": total_juridictions},
                ])

                df_dict["Statistiques globales"] = stats_glob_df

                # Pour chaque colonne analysée, tu choisis celles pertinentes pour le mail
                for i, (column, title) in enumerate(columns_to_analyze_crm):
                    fig, table = create_dashboard(df, column, title)
                    if table is not None and not table.empty:
                        df_dict[title] = table
                        # Sauver éventuellement l'image du graphe
                        file_img = f"/tmp/{column}_tab4_mail.png"
                        try:
                            fig.write_image(file_img, width=640, height=400)
                            img_paths[f"{title}_img"] = file_img
                        except Exception:
                            pass  # ignore les erreurs image

                # Différences Pays/Juridiction
                if not country_jurisdiction_mismatch.empty:
                    df_dict["Incohérences Pays/Juridiction"] = country_jurisdiction_mismatch

                if not missing_info.empty:
                    df_dict["Business Relationships incomplètes"] = missing_info

                if not active_rebates.empty:
                    df_dict["Contrats de distribution actifs"] = active_rebates
                def send_analyse_comptes_mail(subject, to, cc, df_dict, img_paths, excel_attachment=None):
                    from win32com.client import Dispatch
                    html = """
                    <html>
                    <head>
                    <style>
                      body { font-family: Arial,sans-serif; margin:10px;}
                      .dash-row { display:flex; align-items:flex-start; margin-bottom:28px; }
                      .dash-table { flex:1; max-width:60%; }
                      .dash-img { flex:1; margin-left:32px; }
                      table { border-collapse:collapse; width:100%; }
                      th,td {border:1px solid #b6c2d2;padding:7px 12px;text-align:left;}
                      th {background-color:#2F5496;color:#fff;}
                      h3 {margin-bottom:7px;}
                      hr { margin:30px 0; }
                    </style>
                    </head>
                    <body>
                    <h2>Dashboard Analyse des comptes</h2>
                    """
                    for titre, df in df_dict.items():
                        if df is not None and not df.empty:
                            html += f"<h3>{titre}</h3><div class='dash-row'>"
                            html += f"<div class='dash-table'>{df.head(10).to_html(index=False)}</div>"
                            img_key = f"{titre}_img"
                            if img_key in img_paths and os.path.exists(img_paths[img_key]):
                                with open(img_paths[img_key], "rb") as f:
                                    b64img = base64.b64encode(f.read()).decode()
                                html += f"<div class='dash-img'><img src='data:image/png;base64,{b64img}' width='350'></div>"
                            html += "</div><hr/>"
                    html += "<p style='margin-top:30px;'>Bien cordialement,</body></html>"
                    outlook = Dispatch("Outlook.Application")
                    mail = outlook.CreateItem(0)
                    mail.Subject = subject
                    mail.To = to
                    mail.CC = cc
                    mail.HTMLBody = html
                    if excel_attachment and os.path.exists(excel_attachment):
                        mail.Attachments.Add(excel_attachment)
                    mail.Display()

                # -- Ajoute ce bouton à la fin de la page (après tous les calculs !)
                if st.button("📧 Générer mail Analyse 'Account'", key="gen_mail_tab4"):
                    send_analyse_comptes_mail(
                        subject=f"{dt.datetime.today().strftime('%Y%m%d')} - Dashboard Analyse comptes",
                        to="rcci@syquant.com",
                        cc="jeremy.saus@syquant.com",
                        df_dict=df_dict,
                        img_paths=img_paths,
                        excel_attachment=excel_path if 'excel_path' in locals() else None
                    )
                    st.success("✅ Email dashboard généré, prêt à envoyer !")


#--------------------------------------------    ANALYSE DU REGISTRE CACEIS ---------------------------------------------------------
    
    with tab4:
        st.header("🏦 Analyse du registre CACEIS")

        uploaded_csv = st.file_uploader(
            "Déposez ici votre fichier 'Investor Status Report.csv'",
            type=["csv"],
            key="registrefile"
        )

        if uploaded_csv is not None:
            # Lire la première ligne pour les entêtes et nettoyer les noms de colonnes
            uploaded_csv.seek(0)
            header = uploaded_csv.readline().decode("utf-8").strip().replace(",,", ",").split(";")
            header = [col.replace(",,", ",").strip(",") for col in header]  # Nettoyage supplémentaire

            # Revenir au début du fichier pour la lecture complète
            uploaded_csv.seek(0)
            df = pd.read_csv(uploaded_csv, sep=";", names=header, header=0, encoding="utf-8", low_memory=False)
            df.columns = [col.replace(",,", ",").replace(",", "").strip() for col in df.columns]
            df = df.applymap(lambda x: str(x).replace(",,", ",") if isinstance(x, str) else x)

            # Explicitly convert 'Balance Amount in Euro' to numeric, coercing errors
            balance_col_name = "Balance Amount in Euro"
            if balance_col_name in df.columns:
                df[balance_col_name] = pd.to_numeric(df[balance_col_name], errors='coerce').fillna(0)
            else:
                st.warning(f"Colonne d'encours ('{balance_col_name}') non trouvée dans le fichier.")


            # 3. Ajouter la colonne Pays selon la nationalité
            nat_to_country = {
                "LUX": "Luxembourg", "FRA": "France", "AUT": "Autriche", "ESP": "Espagne", "CHE": "Suisse",
                "ITA": "Italie", "SWE": "Suède", "GBR": "Royaume-Uni", "DEU": "Allemagne", "FIN": "Finlande",
                "NLD": "Pays-Bas", "MCO": "Monaco", "BEL": "Belgique", "IMN": "Île de Man", "USA": "États-Unis",
                "IRL": "Irlande", "DNK": "Danemark", "MLT": "Malte", "GGY": "Guernesey"
            }
            nat_col = "Country of Tax Residence 1"  # adapte si besoin
            if nat_col in df.columns:
                df["Pays"] = df[nat_col].map(nat_to_country)
            else:
                st.warning("Colonne de nationalité non trouvée dans le fichier.")

            # 4. Ajouter la colonne Investor Risk Indicator (logique Excel demandée)
            col_T = "Vigilance Level (Deduced)"
            col_U = "Vigilance Level (Entered)"
            if col_U in df.columns and col_T in df.columns:
                df["Investor Risk Indicator"] = df.apply(
                    lambda row: row[col_U] if pd.notna(row[col_U]) and row[col_U] != "" else row[col_T],
                    axis=1
                )
            else:
                st.warning(
                    f"Colonnes pour la formule non trouvées. "
                    f"Vérifie bien le nom exact de la colonne T (actuellement '{col_T}') et U (actuellement '{col_U}')."
                )

            # Ajouter les drapeaux
            country_to_flag = {
                "Luxembourg": "",
                "France": "",
                "Autriche": "",
                "Espagne": "",
                "Suisse": "",
                "Italie": "",
                "Suède": "",
                "Royaume-Uni": "",
                "Allemagne": "",
                "Finlande": "",
                "Pays-Bas": "",
                "Monaco": "",
                "Belgique": "",
                "Île de Man": "",
                "États-Unis": "",
                "Irlande": "",
                "Danemark": "",
                "Malte": "",
                "Guernesey": ""
            }
            if "Pays" in df.columns:
                df["Pays"] = df["Pays"].apply(
                    lambda x: f"{country_to_flag.get(x, '')} {x}" if pd.notna(x) and x in country_to_flag else x
                )

                # ----------------------------------------------------
                # PART 1: Sidebar Input for AMF/CFT Levels
                # ----------------------------------------------------
                if "amf_cft_levels" not in st.session_state:
                    st.session_state["amf_cft_levels"] = {}

                with st.sidebar:
                    with st.popover("🛡️ Remplir matrice AMF/CFT par pays"):
                        st.markdown("Assignez un niveau de risque AMF/CFT pour chaque pays détecté.")

                        if "Pays" in df.columns:
                            unique_countries_df = df["Pays"].dropna().unique().tolist()
                            unique_countries_df.sort()  # Sort countries alphabetically
                            amf_cft_options = ["Non défini", "Low", "Medium-Low", "Medium", "High", "Autre"]

                            # Ensure amf_cft_levels in session state is initialized for all countries
                            for country in unique_countries_df:
                                if country not in st.session_state["amf_cft_levels"]:
                                    st.session_state["amf_cft_levels"][country] = "Non défini"  # Default value
                            for country in unique_countries_df:
                                current_level = st.session_state["amf_cft_levels"].get(country, "Non défini")
                                
                                # If the current level is a tuple (Autre, custom_text), extract the custom text
                                custom_text_value = ""
                                if isinstance(current_level, tuple) and current_level[0] == "Autre":
                                    current_level_display = current_level[0] # Display "Autre" in selectbox
                                    custom_text_value = current_level[1] # Actual custom text
                                else:
                                    current_level_display = current_level # Display normal level

                                selected_level = st.selectbox(
                                    f"Niveau pour {country}:",
                                    amf_cft_options,
                                    index=amf_cft_options.index(current_level_display) if current_level_display in amf_cft_options else 0,
                                    key=f"amf_cft_selectbox_{country}"
                                )
                                
                                custom_risk_text_key = f"amf_cft_custom_text_{country}"
                                if selected_level == "Autre":
                                    custom_risk_text = st.text_input(
                                        f"Spécifiez le niveau pour {country} (Autre):",
                                        value=st.session_state.get(custom_risk_text_key, custom_text_value), # Use current_level[1] for default
                                        key=custom_risk_text_key
                                    )
                                    # Store as a tuple (Autre, custom_text)
                                    st.session_state["amf_cft_levels"][country] = ("Autre", custom_risk_text)
                                else:
                                    st.session_state["amf_cft_levels"][country] = selected_level

                                # Save levels immediately after selection
                                save_amf_cft_levels(st.session_state["amf_cft_levels"])

            # 5. Afficher le tableau propre
            st.dataframe(df.sort_values(by="Balance Amount in Euro", ascending=False))

            # --- TCD : Liste des investisseurs filtrés ---
            st.subheader("Liste des High Risk Investors")

            fund_col = "Fund Name"
            risk_col = "Investor Risk Indicator"

            fund_options = ["Tous"]
            if fund_col in df.columns:
                unique_funds = df[fund_col].dropna().unique().tolist()
                fund_options += unique_funds
                fund_options.append("Vide")
            risk_options = ["Tous"]
            if risk_col in df.columns:
                unique_risks = df[risk_col].dropna().unique().tolist()
                risk_options += unique_risks

            col1, col2 = st.columns(2)
            with col1:
                selected_fund = st.selectbox("Filtrer par Fund Name :", fund_options)
            with col2:
                selected_risk = st.selectbox("Filtrer par niveau de vigilance :", risk_options)

            # Appliquer les filtres sur le DataFrame principal
            df_filtered = df.copy()
            if selected_fund != "Tous" and fund_col in df_filtered.columns:
                if selected_fund == "Vide":
                    df_filtered = df_filtered[df_filtered[fund_col].isna() | (df_filtered[fund_col] == "")]
                else:
                    df_filtered = df_filtered[df_filtered[fund_col] == selected_fund]
            if selected_risk != "Tous" and risk_col in df_filtered.columns:
                df_filtered = df_filtered[df_filtered[risk_col] == selected_risk]

            # Colonnes à afficher
            cols_to_show = [
                "Investor Account Identifier",  # A
                "Investor Identifier",          # G
                "Investor Account Name",        # AJ
                "Official Name",                # S
                "Vigilance Level (Deduced)",    # T
                "Vigilance Level (Entered)",    # U
                "Nationality",                  # Q
                "Balance Amount in Euro"        # K
            ]
            cols_to_show = [col for col in cols_to_show if col in df_filtered.columns]

            # Affichage compact du tableau avec case à cocher
            show_table = st.checkbox("Afficher le tableau des High Risk Investors", value=True)
            if show_table and not df_filtered.empty and cols_to_show:
                # Vérifier et trier par encours
                if "Balance Amount in Euro" in df_filtered.columns:
                    df_filtered["Balance Amount in Euro"] = pd.to_numeric(
                        df_filtered["Balance Amount in Euro"], errors="coerce"
                    )
                    df_filtered = df_filtered.sort_values(by="Balance Amount in Euro", ascending=False)

                st.dataframe(df_filtered[cols_to_show].head(5), use_container_width=True, height=150)
                st.caption(f"Affichage des 5 premières lignes sur {len(df_filtered)} investisseurs filtrés (triés par encours ↓).")
            elif not show_table:
                st.caption("Cochez la case pour afficher le tableau des investisseurs filtrés.")
            # --- Statistiques par pays (donut sur le filtré) ---
            st.subheader("Répartition des investisseurs et encours par pays ")
            if not df_filtered.empty and "Pays" in df_filtered.columns:
                pays_stats = df_filtered["Pays"].value_counts().reset_index()
                pays_stats.columns = ["Pays", "Nombre d'investisseurs"]

                # Calculate percentages for grouping
                total_investors = pays_stats["Nombre d'investisseurs"].sum()
                pays_stats["Percentage"] = (pays_stats["Nombre d'investisseurs"] / total_investors) * 100

                # Identify countries less than 2% by investor count
                countries_to_group_investors = pays_stats[pays_stats["Percentage"] < 2]["Pays"].tolist()

                # Donut 1 : nombre d'investisseurs
                col_donut1, col_donut2 = st.columns(2)
                with col_donut1:
                    # Create data for the investor donut chart, grouping small countries
                    pays_stats_grouped_investors = pays_stats.copy()
                    if countries_to_group_investors:
                        autres_investors_count = pays_stats_grouped_investors[pays_stats_grouped_investors["Pays"].isin(countries_to_group_investors)]["Nombre d'investisseurs"].sum()
                        pays_stats_grouped_investors = pays_stats_grouped_investors[~pays_stats_grouped_investors["Pays"].isin(countries_to_group_investors)]
                        if autres_investors_count > 0:
                             autres_row = pd.DataFrame([{"Pays": "Autres", "Nombre d'investisseurs": autres_investors_count}])
                             pays_stats_grouped_investors = pd.concat([pays_stats_grouped_investors, autres_row], ignore_index=True)


                    fig_pays_donut = px.pie(
                        pays_stats_grouped_investors,
                        names="Pays",
                        values="Nombre d'investisseurs",
                        hole=0.5,
                        title="Répartition nombres d'investisseurs par pays ",
                        color_discrete_sequence=px.colors.qualitative.Pastel
                    )
                    st.plotly_chart(fig_pays_donut, use_container_width=True, key="registre_chart_investors_count") # Changed key to be unique
                     # --- TCD : Encours par pays ---
            st.subheader("Répartition des encours par pays")
            if not df_filtered.empty and "Pays" in df_filtered.columns and "Balance Amount in Euro" in df_filtered.columns:
                tcd_pays = (
                    df_filtered.groupby("Pays")["Balance Amount in Euro"]
                    .sum()
                    .reset_index()
                    .rename(columns={"Pays": "Nationalité/Pays", "Balance Amount in Euro": "Encours (€)"})
                )
                tcd_pays["Encours (%)"] = (tcd_pays["Encours (€)"] / tcd_pays["Encours (€)"].sum() * 100).round(2)
                tcd_pays = tcd_pays.sort_values("Encours (€)", ascending=False)
                # Ajout du total général
                total_row_pays = pd.DataFrame({
                    "Nationalité/Pays": ["Total général"],
                    "Encours (€)": [tcd_pays["Encours (€)"].sum()],
                    "Encours (%)": [100.0]
                })
                tcd_pays = pd.concat([tcd_pays, total_row_pays,], ignore_index=True)

                # Format the 'Encours (€)' column with thousand separators for display
                tcd_pays['Encours (€)'] = tcd_pays['Encours (€)'].apply(lambda x: f"{x:,.0f} €" if pd.notna(x) else "")

                # Donut 2 : somme des montants par pays
                if "Balance Amount in Euro" in df_filtered.columns:
                    pays_balance = df_filtered.groupby("Pays")["Balance Amount in Euro"].sum().reset_index()
                    pays_balance = pays_balance.sort_values("Balance Amount in Euro", ascending=False)
                    # Calculate percentages for grouping based on balance
                    total_balance = pays_balance["Balance Amount in Euro"].sum()
                    pays_balance["Percentage"] = (pays_balance["Balance Amount in Euro"] / total_balance) * 100
                    # Identify countries less than 2% by balance
                    countries_to_group_balance = pays_balance[pays_balance["Percentage"] < 2]["Pays"].tolist()
                    with col_donut2:
                        # Create data for the balance donut chart, grouping small countries
                        pays_balance_grouped = pays_balance.copy()
                        if countries_to_group_balance:
                            autres_balance_amount = pays_balance_grouped[pays_balance_grouped["Pays"].isin(countries_to_group_balance)]["Balance Amount in Euro"].sum()
                            pays_balance_grouped = pays_balance_grouped[~pays_balance_grouped["Pays"].isin(countries_to_group_balance)]
                            if autres_balance_amount > 0:
                                autres_balance_row = pd.DataFrame([{"Pays": "Autres", "Balance Amount in Euro": autres_balance_amount}])
                                pays_balance_grouped = pd.concat([pays_balance_grouped, autres_balance_row], ignore_index=True)
                        fig_balance_donut = px.pie(
                            pays_balance_grouped,
                            names="Pays",
                            values="Balance Amount in Euro",
                            hole=0.5,
                            title="Répartition des encours par pays",
                            color_discrete_sequence=px.colors.qualitative.Pastel # Add color sequence here
                        )
                        st.plotly_chart(fig_balance_donut, use_container_width=True, key="registre_chart_pays_balance") # Changed key to be unique
                else:
                    st.info("Pas de données pour les statistiques par pays.")
                # Option d'affichage du tableau
                show_tcd_pays = st.checkbox("Afficher le tableau des encours par pays", value=True)
                if show_tcd_pays:
                    st.dataframe(tcd_pays, use_container_width=True)
                    csv_pays = tcd_pays.to_csv(index=False).encode("utf-8")
#-------------------------------------------------------------------------------------------------------------
        # Préparation du DataFrame pour la carte
                st.subheader("Carte de répartition des encours en Europe ")
                df_map = tcd_pays.copy()
                # Nettoie la ligne "Total général"
                df_map = df_map[~df_map["Nationalité/Pays"].str.lower().str.contains("total")]

                # Mapping des pays vers codes ISO3
                iso_map = {
                    "France": "FRA", "Royaume-Uni": "GBR", "Belgique": "BEL", "Suisse": "CHE",
                    "Luxembourg": "LUX", "Espagne": "ESP", "Allemagne": "DEU", "Pays-Bas": "NLD",
                    "Italie": "ITA", "Danemark": "DNK", "Finlande": "FIN", "Suède": "SWE",
                    "États-Unis": "USA", "Irlande": "IRL", "Monaco": "MCO", "Malte": "MLT",
                    "Guernesey": "GGY", "Île de Man": "IMN"
                }
                df_map["Pays_clean"] = df_map["Nationalité/Pays"].apply(lambda x: x.strip().split(" ", 1)[-1])
                df_map["iso_alpha"] = df_map["Pays_clean"].map(iso_map)
                df_map["Encours_num"] = pd.to_numeric(df_map["Encours (%)"], errors="coerce")

                df_map_valid = df_map.dropna(subset=["iso_alpha", "Encours_num"])
                if not df_map_valid.empty:
                    fig_europe = px.choropleth(
                        df_map_valid,
                        locations="iso_alpha",
                        color="Encours_num",
                        hover_name="Pays_clean",
                        color_continuous_scale=px.colors.sequential.Blues,
                        scope="europe",
                        labels={"Encours_num": "Encours (%)"},
                    )
                    fig_europe.update_coloraxes(colorbar_title="Encours (%)")
                    fig_europe.update_layout(margin=dict(l=90, r=90, t=15, b=10))
                    st.plotly_chart(fig_europe, use_container_width=True)
                else:
                    st.info("Aucune donnée exploitable pour afficher la carte.")

       
            # --- Répartition par pays et niveau de vigilance (bar chart empilé, filtré) ---
            st.subheader("Répartition par pays et niveau de vigilance ")
            if not df_filtered.empty and "Pays" in df_filtered.columns and risk_col in df_filtered.columns:
                stacked_stats_all = df_filtered.groupby(["Pays", risk_col]).size().reset_index(name="Nombre d'investisseurs")
                # Calculate total investors per country for sorting
                country_totals = stacked_stats_all.groupby("Pays")["Nombre d'investisseurs"].sum().reset_index()
                # Sort countries by total investors descending
                sorted_countries = country_totals.sort_values("Nombre d'investisseurs", ascending=False)["Pays"]
                # Reorder the stacked_stats_all DataFrame based on sorted countries
                stacked_stats_all["Pays"] = pd.Categorical(stacked_stats_all["Pays"], categories=sorted_countries, ordered=True)
                stacked_stats_all = stacked_stats_all.sort_values("Pays")

                color_map_stacked = { # Use a different color map name for clarity
                    "Light vigilance": "#05a33a",       # Vert un peu plus foncé (Medium Aquamarine)
                    "Standard vigilance": "#FFA500",   # Orange
                    "Reinforced vigilance": "#FF0000" # Rouge
                }
                fig_stacked_all = px.bar(
                    stacked_stats_all,
                    x="Pays",
                    y="Nombre d'investisseurs",
                    color=risk_col,
                    barmode="stack",
                    color_discrete_map=color_map_stacked # Use the new color map
                )
                fig_stacked_all.update_layout(xaxis={'categoryorder':'array', 'categoryarray':sorted_countries})
                st.plotly_chart(fig_stacked_all, use_container_width=True, key="registre_chart_stacked_all") # SUPPRIMEZ , ascending=False

                try:
                    stacked_img_path = os.path.join(tempfile.gettempdir(), "stacked_bar_chart.png")
                    fig_stacked_all.write_image(stacked_img_path, width=800, height=500)
                except Exception as e:
                    print(f"Error saving stacked bar chart: {e}")
                    stacked_img_path = None

            else:
                st.info("Pas de données pour la répartition par pays et niveau de vigilance.")
            # --- TCD : Encours par niveau de vigilance ---
            st.subheader("Indicateurs de risque : Encours par niveau de vigilance")
            # Calcul du TCD
            if not df_filtered.empty and "Investor Risk Indicator" in df_filtered.columns and "Balance Amount in Euro" in df_filtered.columns:
                tcd_risk = (
                    df_filtered.groupby("Investor Risk Indicator")["Balance Amount in Euro"]
                    .sum()
                    .reset_index()
                    .rename(columns={"Investor Risk Indicator": "Niveau de vigilance", "Balance Amount in Euro": "Encours (€)"})
                )
                tcd_risk["Encours (%)"] = (tcd_risk["Encours (€)"] / tcd_risk["Encours (€)"].sum() * 100).round(2)
                tcd_risk = tcd_risk.sort_values("Encours (€)", ascending=False)
                # Ajout du total général
                total_row = pd.DataFrame({
                    "Niveau de vigilance": ["Total général"],
                    "Encours (€)": [tcd_risk["Encours (€)"].sum()],
                    "Encours (%)": [100.0]
                })
                tcd_risk = pd.concat([tcd_risk, total_row], ignore_index=True)
                # Option d'affichage du tableau
                show_tcd = st.checkbox("Afficher le tableau des encours par niveau de vigilance", value=True)
                if show_tcd:
                    st.dataframe(tcd_risk, use_container_width=True)
                    # Option de téléchargement
                    csv = tcd_risk.to_csv(index=False).encode("utf-8")

                # Donut chart interactif
                fig_risk = px.pie(
                    tcd_risk[tcd_risk["Niveau de vigilance"] != "Total général"],
                    names="Niveau de vigilance",
                    values="Encours (%)",
                    hole=0.5,
                    color="Niveau de vigilance",
                    color_discrete_map={
                        "Light vigilance": "#05a33a",
                        "Standard vigilance": "#FFA500",
                        "Reinforced vigilance": "#FF0000"
                    },
                    custom_data=["Encours (€)"]
                )
                fig_risk.update_traces(
                    textinfo="percent+label",
                    hovertemplate="<b>%{label}</b><br>Encours: %{customdata[0]:,.0f} €<br>Part: %{percent}"
                )
                fig_risk.update_layout(
                    showlegend=True
                )

                # Affichage côte à côte des deux donuts
                col_risk_list = st.columns(1)
                with col_risk_list[0]:
                    st.plotly_chart(fig_risk, use_container_width=True, key="donut_risk")
                # Sauvegarde des visuels pour l'export Excel
                import tempfile
 # --- TCD : Encours par Investor Type et Economic Nature Group ---
            st.subheader("Répartition des encours par Type d'investisseur et nature économique")

            investor_type_col = "Investor Type"  # Adjust column name if needed
            economic_nature_col = "Economic Nature Group"  # Adjust column name if needed

            # New filters for this section
            col_filter_type1, col_filter_type2 = st.columns(2)
            with col_filter_type1:
                investor_type_options = ["Tous"] + (df[investor_type_col].dropna().unique().tolist() if investor_type_col in df.columns else [])
                selected_investor_type = st.selectbox("Filtrer par Investor Type :", investor_type_options)
            with col_filter_type2:
                economic_nature_options = ["Tous"] + (df[economic_nature_col].dropna().unique().tolist() if economic_nature_col in df.columns else [])
                selected_economic_nature = st.selectbox("Filtrer par Groupe de nature économique :", economic_nature_options)

            # Apply filters for this specific TCD
            df_tcd_investor = df.copy()
            if selected_investor_type != "Tous" and investor_type_col in df_tcd_investor.columns:
                df_tcd_investor = df_tcd_investor[df_tcd_investor[investor_type_col] == selected_investor_type]
            if selected_economic_nature != "Tous" and economic_nature_col in df_tcd_investor.columns:
                df_tcd_investor = df_tcd_investor[df_tcd_investor[economic_nature_col] == selected_economic_nature]

            # Calculate the TCD
            if (
                not df_tcd_investor.empty
                and investor_type_col in df_tcd_investor.columns
                and economic_nature_col in df_tcd_investor.columns
                and "Balance Amount in Euro" in df_tcd_investor.columns
            ):
                # Fill NaN in economic_nature_col, and re-categorize 'Individual person' if investor type is 'Standard'
                df_tcd_investor[economic_nature_col] = df_tcd_investor.apply(
                    lambda row:
                        'Autre' if pd.isna(row[economic_nature_col]) or row[economic_nature_col] == ''
                        else row[economic_nature_col],
                    axis=1
                )

                # Group by both Investor Type and Economic Nature Group and sum the balance
                tcd_investor_group = (
                    df_tcd_investor.groupby([investor_type_col, economic_nature_col])["Balance Amount in Euro"]
                    .sum()
                    .reset_index()
                    .rename(columns={
                        investor_type_col: "Type d'investisseur",
                        economic_nature_col: "Groupe nature économique",
                        "Balance Amount in Euro": "Encours (€)"
                    })
                )

                # Explicitly correct "Individual person" under "Standard" type after grouping
                # REMOVED: This logic was incorrect and has been removed.

                # Calculate percentage
                total_encours_investor = tcd_investor_group["Encours (€)"].sum()
                if total_encours_investor > 0:
                    tcd_investor_group["Encours (%)"] = (tcd_investor_group["Encours (€)"] / total_encours_investor * 100).round(2)
                else:
                    tcd_investor_group["Encours (%)"] = 0.0

                # Add total row
                total_row_investor = pd.DataFrame({
                    "Type d'investisseur": ["Total général"],
                    "Groupe nature économique": [""],
                    "Encours (€)": [total_encours_investor],
                    "Encours (%)": [100.0 if total_encours_investor > 0 else 0.0]
                })
                tcd_investor_group = pd.concat([tcd_investor_group, total_row_investor], ignore_index=True)

                # Format Encours column
                tcd_investor_group["Encours (€)"] = tcd_investor_group["Encours (€)"].apply(lambda x: f"{x:,.0f} €" if pd.notna(x) else "")

                # Display table
                cols_to_display = ["Type d'investisseur", "Groupe nature économique", "Encours (€)", "Encours (%)"]
                cols_to_display = [col for col in cols_to_display if col in tcd_investor_group.columns]
                if cols_to_display:
                    st.dataframe(tcd_investor_group[cols_to_display], use_container_width=True)
                else:
                    st.info("Les colonnes nécessaires pour afficher le tableau ne sont pas disponibles.")

                # Create and display chart
                st.subheader("Répartition des encours par type et nature d'investisseurs ")
                chart_data = tcd_investor_group[tcd_investor_group["Type d'investisseur"] != "Total général"].copy()

                # Make sure each economic nature only appears with its corresponding investor type
                # This is the key change needed
                chart_data = chart_data.copy()  # Create a clean copy to work with
                if chart_data["Encours (€)"].dtype == object:
                    chart_data["Encours (€)"] = chart_data["Encours (€)"].str.replace(' €', '').str.replace(',', '').astype(float)
                if not chart_data.empty:
                    fig_investor_bar = px.bar(
                        chart_data,
                        x='Type d\'investisseur',
                        y='Encours (€)',
                        color='Groupe nature économique',
                        labels={'Type d\'investisseur': 'Type d\'investisseur', 'Encours (€)': 'Encours (€)', 'Groupe nature économique': 'Groupe économique'},
                        hover_data={'Encours (%)': ':.2f%'}
                    )
                    fig_investor_bar.update_yaxes(tickprefix="€", tickformat=",.0f")
                    fig_investor_bar.update_traces(
                        hovertemplate="<b>%{x}</b><br>%{color}: %{y:,.0f} € (%{customdata[0]:.2f}%)<extra></extra>"
                    )
                    fig_investor_bar.update_layout(
                        xaxis_tickangle=-45,
                        xaxis_automargin=True,
                        bargap=0.2
                    )
                    st.plotly_chart(fig_investor_bar, use_container_width=True, key="registre_chart_investor_group")

                # Extraire uniquement les initiales entre parenthèses pour la légende, ou utiliser une liste d'autorités en dur
                authorities = ["AMF", "FCA", "CSSF", "BaFin", "FINMA", "CONSOB", "CNMV", "CBFA", "FMA", "FSA", "SEC", "DFSA", "MFSA", "AFM", "CMVM", "FSMA", "CBM", "FIN-FSA", "FI", "FMA-AT", "FMA-LI", "FMA-NZ", "FMA-IS", "FMA-EE", "FMA-LU", "FMA-NO", "FMA-SE", "FMA-DK", "FMA-FI", "FMA-IE", "FMA-MT", "FMA-GG", "FMA-JE", "FMA-IM", "FMA-MC", "FMA-MO", "FMA-SM", "FMA-VA"]
                def authority_initial(val):
                    import re
                    # Cherche une autorité connue dans la valeur
                    for auth in authorities:
                        if auth in str(val):
                            return f"({auth})"
                    # Sinon, essaie de trouver des initiales entre parenthèses
                    match = re.search(r'\((.*?)\)', str(val))
                    if match and match.group(1) in authorities:
                        return f"({match.group(1)})"
                    return str(val) # Return original value if no specific authority is found
                chart_data['Groupe nature économique'] = chart_data['Groupe nature économique'].apply(authority_initial)

            # import tempfile # Moved to the beginning of main

            network_path = r"\\panfs001\Syquant Compliance\PG06 - Politique LCB-FT\Screening\6. Analyse du registre CACEIS"
            if not os.path.exists(network_path):
                try:
                    os.makedirs(network_path, exist_ok=True)
                    print(f"Dossier créé: {network_path}")
                except Exception as e:
                    print(f"Impossible d'accéder ou de créer le dossier réseau: {e}")
                    # Fallback sur un dossier temporaire local
                    network_path = tempfile.gettempdir()
                    print(f"Utilisation du dossier temporaire local: {network_path}")

            # Créer le nom du fichier avec la date
            today = dt.datetime.today().strftime("%Y%m%d")
            excel_filename =  f"{today_str} - Analyse du Registre CACEIS.xlsx"
            excel_path = os.path.join(r"\\panfs001\Syquant Compliance\PG06 - Politique LCB-FT\Screening\6. Analyse du registre CACEIS", excel_filename)

            # Définir les chemins des images dans le même dossier que l'Excel
            images_dir = os.path.dirname(excel_path)
            # Save the plots as images for Excel export
            risk_img_path = os.path.join(images_dir, "donut_risk.png")
            stacked_img_path = os.path.join(images_dir, "stacked_bar_chart.png")
            investors_count_donut_img_path = os.path.join(images_dir, "investors_count_donut.png")
            balance_donut_img_path = os.path.join(images_dir, "balance_donut.png")
            investor_account_bar_img_path = os.path.join(images_dir, "investor_account_bar.png")
            investor_type_economic_nature_bar_img_path = os.path.join(images_dir, "investor_type_economic_nature_bar.png")

            # Export new charts if they were generated
            if 'fig_individual_bar' in locals():
                try:
                    fig_individual_bar.write_image(individual_person_img_path, width=800, height=500)
                    print(f"DEBUG: Saved individual person chart to {individual_person_img_path}")
                except Exception as e:
                    print(f"Error saving individual person chart: {e}")
                    individual_person_img_path = None

            if 'fig_standard_bar' in locals():
                try:
                    fig_standard_bar.write_image(standard_bar_img_path, width=800, height=500)
                    print(f"DEBUG: Saved standard bar chart to {standard_bar_img_path}")
                except Exception as e:
                    print(f"Error saving standard bar chart: {e}")
                    standard_bar_img_path = None

            # ----------------------------------------------------
            # PART 2: Final Summary Table with AMF/CFT Levels
            # ----------------------------------------------------
            st.subheader("Tableau récapitulatif AMF/CFT par pays")
            if not df_filtered.empty and "Pays" in df_filtered.columns and "Balance Amount in Euro" in df_filtered.columns:
                # Aggregate data by country
                final_summary_df = (
                    df_filtered.groupby("Pays")["Balance Amount in Euro"]
                    .sum()
                    .reset_index()
                    .rename(columns={"Pays": "Pays", "Balance Amount in Euro": "Encours (€)"})
                )
                
                # Calculate percentages
                total_encours = final_summary_df["Encours (€)"].sum()
                final_summary_df["Encours (%)"] = (final_summary_df["Encours (€)"] / total_encours * 100).round(2)
                
                # Add AMF/CFT Level from session state
                final_summary_df["Niveau AMF/CFT"] = final_summary_df["Pays"].apply(
                    lambda country: st.session_state["amf_cft_levels"].get(country, "Non défini")
                )

                # Format Encours (€) column
                final_summary_df['Encours (€)'] = final_summary_df['Encours (€)'].apply(lambda x: f"{x:,.0f} €" if pd.notna(x) else "")

                # Sort the final table by Encours (€) in descending order
                final_summary_df = final_summary_df.sort_values(by="Encours (€)", ascending=False) # Sorting here ensures the format doesn't break sort

                # Define coloring function for AMF/CFT Level
                def color_amf_cft_level(val):
                    if val == "Low":
                        color = "#4CBB17"  # Light Green
                    elif val == "Medium-Low":
                        color = "#ff9f00 " # Light Orange
                    elif val == "Medium":
                        color = "#FCE5CD" # Orange
                    elif val == "High":
                        color = "#ff2e2e"  # Light Red
                    elif val == "Autre":
                        color = "#CFE2F3" # Light Blue
                    else:
                        color = "#F3F3F3" # Greyish
                    return f'background-color: {color}'

                # Display the styled DataFrame
                st.dataframe(
                    final_summary_df.style.applymap(color_amf_cft_level, subset=['Niveau AMF/CFT']),
                    use_container_width=True,
                    hide_index=True
                )
            else:
                st.info("Aucune donnée filtrée disponible pour le tableau récapitulatif AMF/CFT.")

            # --- New Section: Encours by Official Name / Last Name and Investor Account Identifier ---
            st.subheader("Encours par Nom et identifiants des investisseurs")
            
            investor_account_id_col = "Investor Account Identifier"
            official_name_col = "Official Name / Last Name"
            balance_col = "Balance Amount in Euro"

            if not df_filtered.empty and investor_account_id_col in df_filtered.columns and \
               official_name_col in df_filtered.columns and balance_col in df_filtered.columns:

                tcd_investors_combined = (
                    df_filtered.groupby([investor_account_id_col, official_name_col])[balance_col]
                    .sum()
                    .reset_index()
                    .rename(columns={balance_col: "Encours (€)"})
                )
                
                total_encours_combined = tcd_investors_combined["Encours (€)"].sum()
                if total_encours_combined > 0:
                    tcd_investors_combined["Encours (%)"] = (tcd_investors_combined["Encours (€)"] / total_encours_combined * 100).round(2)
                else:
                    tcd_investors_combined["Encours (%)"] = 0.0
                
                tcd_investors_combined = tcd_investors_combined.sort_values("Encours (€)", ascending=False)

                # Add total row
                total_row_combined = pd.DataFrame({
                    investor_account_id_col: ["Total général"],
                    official_name_col: [""],
                    "Encours (€)": [total_encours_combined],
                    "Encours (%)": [100.0 if total_encours_combined > 0 else 0.0]
                })
                tcd_investors_combined = pd.concat([tcd_investors_combined, total_row_combined], ignore_index=True)

                tcd_investors_combined["Encours (€)"] = tcd_investors_combined["Encours (€)"].apply(lambda x: f"{x:,.0f} €" if pd.notna(x) else "")

                st.dataframe(tcd_investors_combined, use_container_width=True)

                csv_combined = tcd_investors_combined.to_csv(index=False).encode("utf-8")

                # Exclude "Total général" for charting and take top 10
                chart_data_top_investors = tcd_investors_combined[tcd_investors_combined[investor_account_id_col] != "Total général"].head(15).copy()
                # Create a combined label for the x-axis if needed, or just use Investor Account Identifier
                chart_data_top_investors['Investor Label'] = chart_data_top_investors[investor_account_id_col] + ' - ' + chart_data_top_investors[official_name_col].fillna('')

                if not chart_data_top_investors.empty:
                    fig_top_investors = px.bar(
                        chart_data_top_investors,
                        x='Investor Label', # Use the combined label
                        y="Encours (€)",
                        title="Top 10 Investisseurs par Encours",
                        color="Encours (€)",
                        color_continuous_scale=px.colors.sequential.Plasma,
                        text="Encours (€)"
                    )

                    fig_top_investors.update_layout(
                        title_font_size=18,
                        xaxis_title="Investisseur (ID - Nom)", 
                        yaxis_title="Encours (€)",
                        xaxis_tickangle=-20, # Rotate labels for better readability
                        xaxis=dict(tickfont=dict(size=8), title_font_size=12),
                        yaxis=dict(tickfont=dict(size=9), title_font_size=12)
                    )
                    st.plotly_chart(fig_top_investors, use_container_width=True)
                    # Save the chart image for Excel export
                    try:
                        fig_top_investors.write_image(investor_account_bar_img_path, width=800, height=500)
                        print(f"DEBUG: Saved top investors bar chart to {investor_account_bar_img_path}")
                    except Exception as e:
                        print(f"Error saving top investors bar chart: {e}")
                        investor_account_bar_img_path = None
                else:
                    st.info("Pas assez de données pour afficher le graphique des top investisseurs.")
            else:
                st.info("Pas de données pour l'analyse des investisseurs (ID/Nom).")

            # --- TCD combiné avec AML Investor-Fund Operability AU DÉBUT ---
            st.subheader("Encours par AML Investor-Fund Operability")
                        
            investor_account_id_col = "Investor Account Identifier"
            official_name_col = "Official Name / Last Name"
            balance_col = "Balance Amount in Euro"
            AML_investor_col = "AML Investor-Fund Operability"

            # --- FILTRE sur AML Investor-Fund Operability ---
            if AML_investor_col in df_filtered.columns:
                aml_options = sorted(df_filtered[AML_investor_col].dropna().unique())
                default_aml_selection = [aml for aml in ["BLS", "BLR", "BLO"] if aml in aml_options]
                selected_aml = st.multiselect("Filtrer par AML Investor-Fund Operability :", aml_options, default=default_aml_selection)
                df_filtered_aml = df_filtered[df_filtered[AML_investor_col].isin(selected_aml)]
            else:
                df_filtered_aml = df_filtered

            if not df_filtered_aml.empty and investor_account_id_col in df_filtered_aml.columns and \
            official_name_col in df_filtered_aml.columns and balance_col in df_filtered_aml.columns and AML_investor_col in df_filtered_aml.columns:
                tcd_investors_combined = (
                    df_filtered_aml.groupby([AML_investor_col, investor_account_id_col, official_name_col])[balance_col]
                    .sum()
                    .reset_index()
                    .rename(columns={balance_col: "Encours (€)"})
                )
                total_encours_combined = tcd_investors_combined["Encours (€)"].sum()
                if total_encours_combined > 0:
                    tcd_investors_combined["Encours (%)"] = (tcd_investors_combined["Encours (€)"] / total_encours_combined * 100).round(2)
                else:
                    tcd_investors_combined["Encours (%)"] = 0.0

                tcd_investors_combined = tcd_investors_combined.sort_values("Encours (€)", ascending=False)
                # Add total row
                total_row_combined = pd.DataFrame({
                    AML_investor_col: ["Total général"],
                    investor_account_id_col: [""],
                    official_name_col: [""],
                    "Encours (€)": [total_encours_combined],
                    "Encours (%)": [100.0 if total_encours_combined > 0 else 0.0]
                })
                tcd_investors_combined = pd.concat([tcd_investors_combined, total_row_combined], ignore_index=True)
                tcd_investors_combined["Encours (€)"] = tcd_investors_combined["Encours (€)"].apply(lambda x: f"{x:,.0f} €" if pd.notna(x) else "")
                st.dataframe(tcd_investors_combined, use_container_width=True)
                csv_combined = tcd_investors_combined.to_csv(index=False).encode("utf-8")
            else:
                st.info("Pas de données ou colonnes manquantes pour l'analyse AML.")
#-----------------------------------------------------------------EXCEL REGISTRE CACEIS_______________________________________________________
            # Create a new workbook
            wb = Workbook()

            # Sheet 1: Raw Data
            ws_raw = wb.active
            ws_raw.title = "Investor Status Report" # Rename sheet

            # Write the full DataFrame to the first sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_raw.append(r)

            # Apply basic formatting to the raw data header (optional, but good practice)
            header_font_raw = Font(bold=True)
            for cell in ws_raw[1]:
                 cell.font = header_font_raw

            # Sheet 2: Summary and Visuals
            ws_summary = wb.create_sheet("Synthèse Analyse Registre")

            # Define styles for summary tables (re-defined here to ensure scope)
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            header_font_summary = Font(bold=True, color="FFFFFF") # White text
            header_fill_summary = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Blue fill
            alignment_center = Alignment(horizontal='center', vertical='center')

            # Helper function to add a section (title, table, image) to the summary sheet
            def add_section_to_excel(
                worksheet, current_row_offset, section_title,
                dataframe_to_add=None, image_path=None,
                table_title_font_size=14, table_header_columns=None,
                image_anchor_col=None, image_width=None, image_height=None,
                add_table_borders=True, style_alternating_rows=True
            ):
                # Add section title with better styling (HTML-like header)
                worksheet.append([section_title])
                title_row = current_row_offset + 1
                worksheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=12)  # Merge across more columns
                
                # Style the title like an HTML header
                title_cell = worksheet[f'A{title_row}']
                title_cell.font = Font(bold=True, size=table_title_font_size, color="FFFFFF")
                title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Blue background
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                title_cell.border = Border(
                    bottom=Side(style='medium', color="000000"),
                    top=Side(style='medium', color="000000"),
                    left=Side(style='medium', color="000000"),
                    right=Side(style='medium', color="000000")
                )
                
                # Set row height for title
                worksheet.row_dimensions[title_row].height = 30
                
                current_row_offset += 1  # For the title row
                
                # Add a blank row for spacing with lighter background
                worksheet.append([])
                spacing_row = current_row_offset + 1
                for col_idx in range(1, 13):  # Apply to all merged columns
                    cell = worksheet.cell(row=spacing_row, column=col_idx)
                    cell.fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")  # Light green background
                
                current_row_offset += 1
                start_row_for_content = current_row_offset + 1  # Where table or image will start
                
                # If there's a DataFrame, add it with enhanced styling
                if dataframe_to_add is not None and not dataframe_to_add.empty:
                    # Write header with better styling
                    header_row = current_row_offset + 1
                    worksheet.append(dataframe_to_add.columns.tolist())
                    
                    # Style header row
                    for col_idx, cell in enumerate(worksheet[header_row], 1):
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Darker blue for header
                        if add_table_borders:
                            cell.border = Border(
                                bottom=Side(style='thin'),
                                top=Side(style='thin'),
                                left=Side(style='thin'),
                                right=Side(style='thin')
                            )
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Set row height for header
                    worksheet.row_dimensions[header_row].height = 24
                    
                    # Write data rows with alternating colors and better formatting
                    for r_idx, r in enumerate(dataframe_to_rows(dataframe_to_add, index=False, header=False)):
                        data_row = current_row_offset + 2 + r_idx
                        worksheet.append(r)
                        
                        # Apply alternating row colors and formatting
                        row_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") if style_alternating_rows and r_idx % 2 == 0 else None
                        
                        for c_idx, cell in enumerate(worksheet[data_row], 1):
                            if add_table_borders:
                                cell.border = Border(
                                    bottom=Side(style='thin'),
                                    top=Side(style='thin'),
                                    left=Side(style='thin'),
                                    right=Side(style='thin')
                                )
                            
                            # Center alignment for all cells
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            
                            # Apply row fill if defined
                            if row_fill:
                                cell.fill = row_fill
                            
                            # Format numeric cells (assuming currency in column 3 and percentage in column 4)
                            if c_idx == 3 and isinstance(cell.value, (int, float)):  # Encours (€)
                                cell.number_format = '#,##0 '
                            elif c_idx == 4 and isinstance(cell.value, (int, float)):  # Encours (%)
                                cell.number_format = '0.00%'
                            
                            # Bold font for total row
                            if "total" in str(worksheet.cell(row=data_row, column=1).value).lower():
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")  # Light blue for total
                    
                    table_height_in_rows = len(dataframe_to_add) + 1  # +1 for header
                    current_row_offset += table_height_in_rows
                    
                    # Add table borders (outline)
                    if add_table_borders:
                        table_end_row = current_row_offset
                        for col_idx in range(1, len(dataframe_to_add.columns) + 1):
                            # Top border for header
                            worksheet.cell(row=header_row, column=col_idx).border = Border(
                                top=Side(style='medium'),
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            
                            # Left border for first column
                            if col_idx == 1:
                                for row_idx in range(header_row, table_end_row + 1):
                                    worksheet.cell(row=row_idx, column=col_idx).border = Border(
                                        left=Side(style='medium'),
                                        right=Side(style='thin'),
                                        top=Side(style='thin') if row_idx > header_row else Side(style='medium'),
                                        bottom=Side(style='thin') if row_idx < table_end_row else Side(style='medium')
                                    )
                            
                            # Right border for last column
                            if col_idx == len(dataframe_to_add.columns):
                                for row_idx in range(header_row, table_end_row + 1):
                                    worksheet.cell(row=row_idx, column=col_idx).border = Border(
                                        right=Side(style='medium'),
                                        left=Side(style='thin'),
                                        top=Side(style='thin') if row_idx > header_row else Side(style='medium'),
                                        bottom=Side(style='thin') if row_idx < table_end_row else Side(style='medium')
                                    )
                            
                            # Bottom border for last row
                            worksheet.cell(row=table_end_row, column=col_idx).border = Border(
                                bottom=Side(style='medium'),
                                left=Side(style='thin') if col_idx > 1 else Side(style='medium'),
                                right=Side(style='thin') if col_idx < len(dataframe_to_add.columns) else Side(style='medium'),
                                top=Side(style='thin')
                            )
                
                # If there's an image, add it with better positioning
                if image_path is not None and os.path.exists(image_path):
                    img = XLImage(image_path)
                    img.width = image_width if image_width else 600  # Larger default width
                    img.height = image_height if image_height else 400  # Larger default height
                    
                    # Determine anchor for the image with better positioning
                    if dataframe_to_add is not None and not dataframe_to_add.empty:
                        # If table exists, place image to the right or below depending on space
                        if image_anchor_col is None:
                            # Calculate if there's enough space to the right
                            if len(dataframe_to_add.columns) <= 5:  # Small table, place image to the right
                                image_anchor_col_letter = get_column_letter(len(dataframe_to_add.columns) + 2)  # +2 for 1-indexed and a space column
                                img.anchor = f'{image_anchor_col_letter}{start_row_for_content}'
                            else:  # Large table, place image below
                                img.anchor = f'A{current_row_offset + 2}'  # +2 for spacing
                                current_row_offset += (img.height // 15) + 2  # Update offset for image below
                        else:
                            # Use provided anchor column
                            img.anchor = f'{image_anchor_col}{start_row_for_content}'
                            
                        # If image is to the right, check if it extends beyond current row offset
                        if image_anchor_col is not None or len(dataframe_to_add.columns) <= 5:
                            current_row_offset = max(current_row_offset, start_row_for_content + (img.height // 15))
                    else:
                        # If only image, center it better
                        img.anchor = f'C{start_row_for_content}'  # Start at column C for better centering
                        current_row_offset += (img.height // 15) + 2  # +2 for spacing
                    
                    worksheet.add_image(img)
                
                # Add a divider line after the section (HTML-like)
                divider_row = current_row_offset + 1
                worksheet.append([])
                worksheet.merge_cells(start_row=divider_row, start_column=1, end_row=divider_row, end_column=12)
                divider_cell = worksheet[f'A{divider_row}']
                divider_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Gray divider
                worksheet.row_dimensions[divider_row].height = 4  # Thin divider
                
                current_row_offset += 1
                
                # Add spacing after each section
                worksheet.append([])
                current_row_offset += 1
                
                return current_row_offset  # Return the updated offset

            # --- Initialization of image paths and main Excel file path ---
            excel_path = os.path.join(tempfile.gettempdir(), "analyse_registre.xlsx")
            risk_img_path = os.path.join(tempfile.gettempdir(), "donut_risk.png")
            stacked_img_path = os.path.join(tempfile.gettempdir(), "stacked_bar_chart.png")
            investors_count_donut_img_path = os.path.join(tempfile.gettempdir(), "investors_count_donut.png")
            balance_donut_img_path = os.path.join(tempfile.gettempdir(), "balance_donut.png")
            investor_account_bar_img_path = os.path.join(tempfile.gettempdir(), "investor_account_bar.png")
            investor_type_economic_nature_bar_img_path = os.path.join(tempfile.gettempdir(), "investor_type_economic_nature_bar.png") # New image path
            aml_operability_bar_img_path = os.path.join(tempfile.gettempdir(), "aml_operability_bar_chart.png") # New image path for AML Operability chart

            # Create a new workbook
            wb = Workbook()

            # Sheet 1: Raw Data
            ws_raw = wb.active
            ws_raw.title = "Investor Status Report" # Rename sheet

            # Write the full DataFrame to the first sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_raw.append(r)

            # Apply basic formatting to the raw data header (optional, but good practice)
            header_font_raw = Font(bold=True)
            for cell in ws_raw[1]:
                cell.font = header_font_raw

            # Sheet 2: Summary and Visuals
            ws_summary = wb.create_sheet("Synthèse Analyse Registre")

            # Définir une mise en page globale pour la feuille de synthèse
            ws_summary.sheet_properties.tabColor = "1F4E78"  # Couleur de l'onglet
            ws_summary.page_setup.orientation = 'landscape'  # Orientation paysage pour l'impression
            ws_summary.page_setup.fitToWidth = 1  # Ajuster à la largeur
            ws_summary.page_setup.fitToHeight = 0  # Ajuster automatiquement en hauteur

            # Ajouter un en-tête de page avec logo et titre
            header_row = 1
            ws_summary.append(["ANALYSE DU REGISTRE CACEIS"])
            ws_summary.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=12)
            header_cell = ws_summary['A1']
            header_cell.font = Font(bold=True, size=20, color="FFFFFF")
            header_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Bleu foncé
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_summary.row_dimensions[header_row].height = 50

            # Ajouter la date du rapport
            date_row = 2
            ws_summary.append([f"Rapport généré le {dt.datetime.today().strftime('%d/%m/%Y')}"])
            ws_summary.merge_cells(start_row=date_row, start_column=1, end_row=date_row, end_column=12)
            date_cell = ws_summary['A2']
            date_cell.font = Font(italic=True, size=10)
            date_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Ajouter une ligne vide pour l'espacement
            ws_summary.append([])
            row_offset_summary = 3  # Commencer après l'en-tête et la date

            # Define styles for summary tables (re-defined here to ensure scope)
            thin_border = Border(left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))
            header_font_summary = Font(bold=True, color="FFFFFF") # White text
            header_fill_summary = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Blue fill
            alignment_center = Alignment(horizontal='center', vertical='center')

            # 1. Risk Indicators Section (Table + Donut Chart)
            if 'tcd_risk' in locals() and not tcd_risk.empty:
                row_offset_summary = add_section_to_excel(
                    ws_summary, row_offset_summary,
                    "1. Indicateurs de risque : Encours par niveau de vigilance",
                    dataframe_to_add=tcd_risk,
                    image_path=risk_img_path if 'risk_img_path' in locals() and os.path.exists(risk_img_path) else None,
                    image_anchor_col='E', # Place image starting at column E
                    image_width=500, # Specify width
                    image_height=350 # Specify height
                )
            else:
                row_offset_summary = add_section_to_excel(
                    ws_summary, row_offset_summary,
                    "1. Indicateurs de risque : Encours par niveau de vigilance (Aucune donnée)"
                )

            # 2. Country Distribution Section (Two Donut Charts + Table)
            if ('pays_stats' in locals() and not pays_stats.empty and
                'tcd_pays' in locals() and not tcd_pays.empty and
                'pays_balance' in locals() and not pays_balance.empty):
                
                # Save donut charts if they haven't been saved yet
                if 'fig_pays_donut' in locals() and not os.path.exists(investors_count_donut_img_path):
                    try:
                        fig_pays_donut.write_image(investors_count_donut_img_path, width=600, height=400)
                    except Exception as e:
                        print(f"Error saving investors count donut chart: {e}")
                
                if 'fig_balance_donut' in locals() and not os.path.exists(balance_donut_img_path):
                    try:
                        fig_balance_donut.write_image(balance_donut_img_path, width=600, height=400)
                    except Exception as e:
                        print(f"Error saving balance donut chart: {e}")
                
                # Add the section with both charts and table
                row_offset_summary = add_section_to_excel(
                    ws_summary, row_offset_summary,
                    "2. Répartition des investisseurs et encours par pays",
                    dataframe_to_add=tcd_pays,
                    image_path=None # We'll add images separately for this section
                )
                
                # Add both donut charts side by side
                if os.path.exists(investors_count_donut_img_path) and os.path.exists(balance_donut_img_path):
                    img_inv_count = XLImage(investors_count_donut_img_path)
                    img_inv_count.width, img_inv_count.height = 450, 350
                    img_inv_count.anchor = f'A{row_offset_summary + 1}'
                    ws_summary.add_image(img_inv_count)
                    
                    img_bal_donut = XLImage(balance_donut_img_path)
                    img_bal_donut.width, img_bal_donut.height = 450, 350
                    img_bal_donut.anchor = f'G{row_offset_summary + 1}'
                    ws_summary.add_image(img_bal_donut)
                    
                    # Update row offset based on image height
                    row_offset_summary += (img_inv_count.height // 15) + 3
                
                # Add a title for the charts
                ws_summary.append(["Répartition graphique par pays"])
                chart_title_row = row_offset_summary + 1
                ws_summary.merge_cells(start_row=chart_title_row, start_column=1, end_row=chart_title_row, end_column=12)
                chart_title_cell = ws_summary[f'A{chart_title_row}']
                chart_title_cell.font = Font(bold=True, size=12)
                chart_title_cell.alignment = Alignment(horizontal='center', vertical='center')
                row_offset_summary += 2 # For title and spacing
            else:
                row_offset_summary = add_section_to_excel(
                    ws_summary, row_offset_summary,
                    "2. Répartition des investisseurs et encours par pays (Aucune donnée)"
                )

            # 3. Vigilance Level Stacked Bar Chart
            if stacked_img_path is not None and os.path.exists(stacked_img_path):
                row_offset_summary = add_section_to_excel(
                    ws_summary, row_offset_summary,
                    "3. Répartition par pays et niveau de vigilance",
                    image_path=stacked_img_path,
                    image_width=800, image_height=500
                )
            else:
                row_offset_summary = add_section_to_excel(
                    ws_summary, row_offset_summary,
                    "3. Répartition par pays et niveau de vigilance (Aucune donnée)"
                )

            # 4. Investor Type and Economic Nature Group (Table + Bar Chart)
            if (
                'tcd_investor_group' in locals() and not tcd_investor_group.empty and
                'investor_type_col' in locals() and 'economic_nature_col' in locals()
            ):
                # Save the chart image if it was generated
                if 'fig_investor_bar' in locals() and not os.path.exists(investor_type_economic_nature_bar_img_path):
                    try:
                        fig_investor_bar.write_image(investor_type_economic_nature_bar_img_path, width=800, height=500)
                    except Exception as e:
                        print(f"Error saving investor type/economic nature bar chart: {e}")
                        investor_type_economic_nature_bar_img_path = None
                
                row_offset_summary = add_section_to_excel(
                    ws_summary, row_offset_summary,
                    "4. Répartition des encours par Type d'investisseur et nature économique",
                    dataframe_to_add=tcd_investor_group[['Type d\'investisseur', 'Groupe nature économique', 'Encours (€)', 'Encours (%)']],
                    image_path=investor_type_economic_nature_bar_img_path if os.path.exists(investor_type_economic_nature_bar_img_path) else None,
                    image_width=800, image_height=500,
                    image_anchor_col='A', # Place image below table
                    style_alternating_rows=True
                )
            else:
                row_offset_summary = add_section_to_excel(
                    ws_summary, row_offset_summary,
                    "4. Répartition des encours par Type d'investisseur et nature économique (Aucune donnée)"
                )

            # 5. Official Name / Last Name and Investor Account Identifier (Tables + Bar Chart)
            if (
                'tcd_investors_combined' in locals() and not tcd_investors_combined.empty
            ):
                # Save the chart image if it was generated
                if 'fig_top_investors' in locals() and not os.path.exists(investor_account_bar_img_path):
                    try:
                        fig_top_investors.write_image(investor_account_bar_img_path, width=800, height=500)
                    except Exception as e:
                        print(f"Error saving top investors bar chart: {e}")
                        investor_account_bar_img_path = None
                
                row_offset_summary = add_section_to_excel(
                    ws_summary, row_offset_summary,
                    "5. Encours par Nom et identifiants des investisseurs",
                    dataframe_to_add=tcd_investors_combined[[
                        investor_account_id_col,
                        official_name_col,
                        "Encours (€)",
                        "Encours (%)"
                    ]],
                    image_path=investor_account_bar_img_path if os.path.exists(investor_account_bar_img_path) else None,
                    image_width=800, image_height=500,
                    image_anchor_col='A', # Place image below table
                    style_alternating_rows=True
                )
            else:
                row_offset_summary = add_section_to_excel(
                    ws_summary, row_offset_summary,
                    "5. Encours par Nom et identifiants des investisseurs (Aucune donnée)"
                )

            # 6. AML Investor-Fund Operability Filtered Tables and Chart
            if (
                'tcd_investors_combined' in locals() and not tcd_investors_combined.empty and
                AML_investor_col in df_filtered.columns
            ):
                # Check if AML operability chart exists or can be generated
                if 'fig_aml_operability' in locals() and not os.path.exists(aml_operability_bar_img_path):
                    try:
                        fig_aml_operability.write_image(aml_operability_bar_img_path, width=800, height=500)
                    except Exception as e:
                        print(f"Error saving AML operability bar chart: {e}")
                        aml_operability_bar_img_path = None
                
                row_offset_summary = add_section_to_excel(
                    ws_summary, row_offset_summary,
                    "6. Encours par AML Investor-Fund Operability et Investisseur",
                    dataframe_to_add=tcd_investors_combined[[
                        AML_investor_col,
                        investor_account_id_col,
                        official_name_col,
                        "Encours (€)",
                        "Encours (%)"
                    ]] if AML_investor_col in tcd_investors_combined.columns else None,
                    image_path=aml_operability_bar_img_path if os.path.exists(aml_operability_bar_img_path) else None,
                    image_width=800, image_height=500,
                    image_anchor_col='A', # Place image below table
                    style_alternating_rows=True
                )
            else:
                row_offset_summary = add_section_to_excel(
                    ws_summary, row_offset_summary,
                    "6. Encours par AML Investor-Fund Operability et Investisseur (Aucune donnée)"
                )

            # 7. AMF/CFT Levels by Country (if available)
            if 'final_summary_df' in locals() and not final_summary_df.empty:
                # Create a copy to ensure we don't modify the original
                amf_cft_summary = final_summary_df.copy()
                
                # Format the Niveau AMF/CFT column for Excel
                def format_amf_cft_level(level):
                    if isinstance(level, tuple) and level[0] == "Autre":
                        return f"Autre: {level[1]}"
                    return level
                
                if "Niveau AMF/CFT" in amf_cft_summary.columns:
                    amf_cft_summary["Niveau AMF/CFT"] = amf_cft_summary["Niveau AMF/CFT"].apply(format_amf_cft_level)
                
                row_offset_summary = add_section_to_excel(
                    ws_summary, row_offset_summary,
                    "7. Tableau récapitulatif AMF/CFT par pays",
                    dataframe_to_add=amf_cft_summary,
                    style_alternating_rows=True
                )
                
                # Apply conditional formatting for AMF/CFT levels
                if "Niveau AMF/CFT" in amf_cft_summary.columns:
                    amf_col_idx = amf_cft_summary.columns.get_loc("Niveau AMF/CFT") + 1  # +1 for 1-indexed Excel columns
                    
                    # Find the range of rows with data
                    start_row = row_offset_summary - len(amf_cft_summary)
                    end_row = row_offset_summary - 1  # -1 because row_offset_summary is already at the next section
                    
                    for row_idx in range(start_row, end_row + 1):
                        cell = ws_summary.cell(row=row_idx, column=amf_col_idx)
                        level_value = cell.value
                        
                        if level_value == "Low":
                            cell.fill = PatternFill(start_color="4CBB17", end_color="4CBB17", fill_type="solid")  # Green
                        elif level_value == "Medium-Low":
                            cell.fill = PatternFill(start_color="ff9f00", end_color="ff9f00", fill_type="solid")  # Light Orange
                        elif level_value == "Medium":
                            cell.fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")  # Orange
                        elif level_value == "High":
                            cell.fill = PatternFill(start_color="ff2e2e", end_color="ff2e2e", fill_type="solid")  # Red
                        elif "Autre" in str(level_value):
                            cell.fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")  # Light Blue
            else:
                row_offset_summary = add_section_to_excel(
                    ws_summary, row_offset_summary,
                    "7. Tableau récapitulatif AMF/CFT par pays (Aucune donnée)"
                )

            # Améliorer l'auto-dimensionnement des colonnes
            for col_index in range(1, ws_summary.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_index)
                
                # Calculer la largeur optimale
                for row in ws_summary.iter_rows(min_col=col_index, max_col=col_index):
                    for cell in row:
                        try:
                            if cell.value is not None:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                
                # Ajuster la largeur avec un facteur d'échelle selon le type de colonne
                if col_index <= 2:  # Colonnes d'identifiants ou de noms
                    adjusted_width = min(max_length + 2, 40)  # Limiter la largeur maximale
                elif col_index == 3:  # Colonne des montants
                    adjusted_width = max(12, min(max_length + 4, 20))  # Largeur minimale et maximale
                elif col_index == 4:  # Colonne des pourcentages
                    adjusted_width = max(10, min(max_length + 2, 15))
                else:
                    adjusted_width = min(max_length + 2, 30)
                
                ws_summary.column_dimensions[column_letter].width = adjusted_width

            # Ajouter un pied de page
            footer_row = ws_summary.max_row + 2
            ws_summary.append([])
            ws_summary.append(["© Syquant Capital - Document confidentiel"])
            ws_summary.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=12)
            footer_cell = ws_summary[f'A{footer_row}']
            footer_cell.font = Font(size=8, italic=True)
            footer_cell.alignment = Alignment(horizontal='center')

            try:
                # Créer le dossier parent si nécessaire
                os.makedirs(os.path.dirname(excel_path), exist_ok=True)
                
                wb.save(excel_path)
            except PermissionError:
                st.error(f"❌ Erreur de permission lors de l'enregistrement du fichier Excel. Vérifiez vos droits d'accès au dossier réseau.")
                # Fallback sur un dossier temporaire local
                fallback_path = os.path.join(tempfile.gettempdir(), excel_filename)
                try:
                    wb.save(fallback_path)
                    excel_path = r"\\panfs001\Syquant Compliance\PG06 - Politique LCB-FT\Screening\6. Analyse du registre CACEIS" # Mettre à jour le chemin pour les opérations suivantes
                    st.warning(f"⚠️ Excel enregistré dans un dossier temporaire: {excel_path}")
                except Exception as e:
                    st.error(f"❌ Erreur lors de la sauvegarde du fichier Excel dans le dossier temporaire: {e}")
                    excel_path = None
            except Exception as e:
                st.error(f"❌ Erreur lors de la sauvegarde du fichier Excel: {e}")
                excel_path = None

            def excel_to_pdf(excel_path):
                import os
                import datetime as dt
                import streamlit as st
                from reportlab.platypus import (
                    SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, Image, PageBreak
                )
                from reportlab.lib.pagesizes import landscape, letter
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib import colors
                from reportlab.lib.units import inch

                # Chemin d’enregistrement
                network_path = r"\\panfs001\Syquant Compliance\PG06 - Politique LCB-FT\Screening\6. Analyse du registre CACEIS"
                today_str = dt.datetime.today().strftime("%Y%m%d")
                pdf_filename = f"Analyse_Registre_CACEIS_{today_str}.pdf"
                pdf_path = os.path.join(network_path, pdf_filename)
                
                if not os.path.exists(network_path):
                    try:
                        os.makedirs(network_path)
                    except Exception:
                        # Fallback local
                        pdf_path = os.path.join(os.path.dirname(excel_path), pdf_filename)

                # Initialiser le document
                doc = SimpleDocTemplate(pdf_path, pagesize=landscape(letter))
                elements = []
                styles = getSampleStyleSheet()
                title_style = styles['Heading1']
                subtitle_style = styles['Heading2']
                normal_style = styles['Normal']
                footer_style = ParagraphStyle('footer', parent=normal_style, fontSize=7, alignment=1)

                title_style.alignment = 1
                title_style.fontSize = 18

                # ==== TITRE ====
                elements.append(Paragraph("ANALYSE DU REGISTRE CACEIS", title_style))
                elements.append(Spacer(1, 10))
                elements.append(Paragraph(f"Rapport généré le {dt.datetime.today().strftime('%d/%m/%Y')}", normal_style))
                elements.append(Spacer(1, 20))

                # ==== SECTIONS GÉNÉRIQUES ====
                def add_table(dataframe, col_widths=None):
                    if dataframe.empty:
                        return None

                    data = [dataframe.columns.tolist()] + dataframe.values.tolist()
                    # Formatage des données
                    for i in range(1, len(data)):
                        data[i] = [
                            val if not isinstance(val, (int, float)) else f"{val:,.2f}".replace(",", " ") + (" €" if "€" in dataframe.columns[1] else "")
                            for val in data[i]
                        ]

                    table = Table(data, colWidths=col_widths)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F497D")),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ]))
                    return table

                def add_image(img_path, width=6*inch, height=4*inch):
                    if img_path and os.path.exists(img_path):
                        img = Image(img_path, width=width, height=height)
                        table = Table([[img]], colWidths=[width])
                        table.setStyle(TableStyle([
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER')
                        ]))
                        return table
                    return None

                # === SECTION 1 ===
                elements.append(Paragraph("1. Indicateurs de risque : Encours par niveau de vigilance", subtitle_style))
                elements.append(Spacer(1, 10))
                if 'tcd_risk' in globals():
                    table = add_table(tcd_risk, col_widths=[2.5*inch, 2*inch, 1.5*inch])
                    if table:
                        elements.append(table)
                        elements.append(Spacer(1, 10))
                elements.append(add_image(risk_img_path))
                elements.append(PageBreak())

                # === SECTION 2 ===
                elements.append(Paragraph("2. Répartition des encours par pays", subtitle_style))
                elements.append(Spacer(1, 10))
                if 'tcd_pays' in globals():
                    table = add_table(tcd_pays, col_widths=[3*inch, 2*inch, 1.5*inch])
                    if table:
                        elements.append(table)
                        elements.append(Spacer(1, 10))

                # Graphiques côte à côte
                if os.path.exists(investors_count_donut_img_path) and os.path.exists(balance_donut_img_path):
                    img1 = Image(investors_count_donut_img_path, width=3.5*inch, height=3*inch)
                    img2 = Image(balance_donut_img_path, width=3.5*inch, height=3*inch)
                    img_table = Table([[img1, img2]], colWidths=[3.5*inch, 3.5*inch])
                    img_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER')
                    ]))
                    elements.append(Spacer(1, 10))
                    elements.append(img_table)
                elements.append(PageBreak())

                # === SECTION 3 ===
                elements.append(Paragraph("3. Répartition par pays et niveau de vigilance", subtitle_style))
                elements.append(Spacer(1, 10))
                if 'tcd_pays_risk' in globals():
                    table = add_table(tcd_pays_risk, col_widths=[3*inch, 2*inch, 1.5*inch])
                    if table:
                        elements.append(table)
                        elements.append(Spacer(1, 10))
                elements.append(add_image(stacked_img_path, width=7*inch, height=4.5*inch))
                elements.append(PageBreak())

                # === SECTION 4 ===
                elements.append(Paragraph("4. Répartition par type d'investisseur et nature économique", subtitle_style))
                elements.append(Spacer(1, 10))
                if 'tcd_investor_group' in globals():
                    table = add_table(tcd_investor_group, col_widths=[2.5*inch, 2.5*inch, 2*inch, 1.5*inch])
                    if table:
                        elements.append(table)
                        elements.append(Spacer(1, 10))
                elements.append(add_image(investor_type_economic_nature_bar_img_path))
                elements.append(PageBreak())

                # === SECTION 5 ===
                elements.append(Paragraph("5. Encours par Nom et identifiants des investisseurs", subtitle_style))
                elements.append(Spacer(1, 10))
                if 'tcd_investors_combined' in globals():
                    inv_top15 = tcd_investors_combined.head(15)
                    table = add_table(inv_top15, col_widths=[1.5*inch, 3*inch, 2*inch, 1.5*inch])
                    if table:
                        elements.append(table)
                        elements.append(Spacer(1, 10))
                        elements.append(Paragraph("(Top 15 investisseurs affichés)", normal_style))
                elements.append(add_image(investor_account_bar_img_path))
                elements.append(PageBreak())

                # === SECTION 6 ===
                elements.append(Paragraph("6. Tableau récapitulatif AMF/CFT par pays", subtitle_style))
                elements.append(Spacer(1, 10))
                if 'final_summary_df' in globals():
                    table = add_table(final_summary_df, col_widths=[2*inch, 2*inch, 1.5*inch, 2.5*inch])
                    if table:
                        elements.append(table)
                        elements.append(Spacer(1, 10))

                # ==== FOOTER ====
                elements.append(Spacer(1, 20))
                elements.append(Paragraph("© Syquant Capital - Document confidentiel", footer_style))

                try:
                    doc.build(elements)
                    if os.path.exists(pdf_path):
                        return pdf_path
                    else:
                        st.error("❌ Le fichier PDF final n’a pas été généré.")
                except Exception as e:
                    st.error(f"❌ Erreur PDF : {e}")
                    return None

            # Ajouter le bouton dans l'interface Streamlit
            if excel_path and os.path.exists(excel_path):
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("📧 Générer l'email Outlook pour l'analyse registre", key="generate_email_registre"):
                        try:
                            outlook = win32.Dispatch("Outlook.Application")
                            mail = outlook.CreateItem(0)
                            today = dt.datetime.today().strftime("%Y%m%d")
                            mail.Subject = f" {today} - Analyse du Registre CACEIS"
                            mail.To = "rcci@syquant.com"
                            mail.CC = "jeremy.saus@syquant.com;guillaume.bibian@syquant.com"
                            mail.HTMLBody = (
                                "<p>Bonjour,</p>"
                                "<p>Veuillez trouver ci-joint l'analyse complète des registres CACEIS.</p>"
                                "<p>Bien cordialement,<br></p>"
                            )
                            mail.SentOnBehalfOfName = "rcci@syquant.com"
                            mail.Attachments.Add(excel_path)
                            mail.Display()
                            st.success("✅ Email Outlook généré et affiché avec succès.")
                        except Exception as e:
                            st.error(f"Erreur lors de la génération de l'email Outlook : {e}")
                
                with col2:
                    # Bouton pour convertir et télécharger le PDF
                    if st.button("📄 Convertir et télécharger en PDF", key="convert_to_pdf"):
                        with st.spinner("Conversion de l'Excel en PDF en cours..."):
                            pdf_path = excel_to_pdf(excel_path)
                            if pdf_path and os.path.exists(pdf_path):
                                with open(pdf_path, "rb") as pdf_file:
                                    pdf_bytes = pdf_file.read()
                                
                                st.download_button(
                                    label="📥 Télécharger le PDF",
                                    data=pdf_bytes,
                                    file_name=os.path.basename(pdf_path),
                                    mime="application/pdf"
                                )
                                st.success("✅ Conversion réussie. Cliquez sur le bouton ci-dessus pour télécharger.")
                            else:
                                st.error("❌ La conversion en PDF a échoué.")
            else:
                st.error("❌ Le fichier Excel d'analyse du registre n'a pas pu être créé.")

    #------------------------------ Recherche manuelle de sanctions--------------------------------------------------------------------------------------------------------------------
    with tab5:
        st.header("🔍 Recherche manuelle de sanctions")

        # Vérification et téléchargement automatique des listes
        check_and_download_sanctions()

        # Interface de recherche
        st.subheader("Entrez les informations de la personne à rechercher")

        with st.form("manual_search_form"):
            col1, col2 = st.columns(2)
            with col1:
                manual_first_name = st.text_input("Prénom")
            with col2:
                manual_last_name = st.text_input("Nom de famille")

            submitted = st.form_submit_button("🔎 Rechercher dans les listes de sanctions")

        if submitted:
            if not manual_last_name or len(manual_last_name.strip()) < min_name_length:
                st.warning("❗ Veuillez saisir un nom valide d'au moins 3 caractères.")
            else:
                with st.spinner("Recherche en cours..."):
                    sanctions = parse_sdn(file_sdn) + parse_nonsdn(file_nonsdn) + parse_fr(file_fr)
                    person = {"First Name": manual_first_name, "Last Name": manual_last_name}

                    if use_intelligent_scoring:
                        match = intelligent_screening(person, sanctions, threshold_nom, threshold_prenom)
                    else:
                        match = traditional_screening(person, sanctions, threshold_nom, threshold_prenom)

                if match:
                    st.success(f"🎯 Correspondance trouvée : {match['Prénom']} {match['Nom']} ({match['Source']})")

                    # Afficher les détails dans un tableau
                    details_df = pd.DataFrame({
                        "Métrique": ["Score Nom", "Score Prénom", "Score Culturel", "Score Total"],
                        "Valeur": [
                            f"{match['Score Nom']:.1f}%",
                            f"{match['Score Prénom']:.1f}%",
                            f"{match.get('Score Cultural', 0):.1f}%",
                            f"{match['Score Total']:.1f}%"
                        ]
                    })
                    st.table(details_df)

                    st.info(f"**Détails supplémentaires :** {match['Détails']}")
                else:
                    st.info("✅ Aucun résultat trouvé dans les listes de sanctions.")
def load_previous_screening_data(results_dir, file_type_prefix, current_screening_date_yyyymmdd):
    """
    Retourne les dernières correspondances Personnes Physiques (PP) ou Personnes Morales (PM)
    (fichiers Excel) antérieurs à la date spécifiée.
    """
    history_df = None
    latest_file_date = None
    latest_filepath = None

    try:
        files = [f for f in os.listdir(results_dir) if f.endswith(".xlsx")]
    except FileNotFoundError:
        st.warning(f"Dossier des résultats non trouvé : {results_dir}")
        return None
    except Exception as e:
        st.error(f"Erreur lors de la lecture du dossier des résultats {results_dir}: {e}")
        return None

    try:
        current_date_obj = dt.datetime.strptime(current_screening_date_yyyymmdd, '%Y%m%d').date()
    except ValueError:
        st.error(f"Format de date actuel invalide: {current_screening_date_yyyymmdd}. Attendu YYYYMMDD.")
        return None

    for file in files:
        file_date_str = extract_date_from_filename_new_format(file)
        if not file_date_str:
            continue

        try:
            file_date = dt.datetime.strptime(file_date_str, '%Y%m%d').date()
            # print(f"DEBUG: Found file: {file}, Date: {file_date_str}")
        except ValueError:
            # print(f"DEBUG: Skipping file '{file}' - could not parse date '{file_date_str}'.")
            continue

        # Check for file type and date condition
        # Revert: Use < to consider only files strictly before the current screening date
        if file_date < current_date_obj:
            if file_type_prefix == 'PP' and "screening des correspondances PP.xlsx" in file:
                if latest_file_date is None or file_date > latest_file_date:
                    latest_file_date = file_date
                    latest_filepath = os.path.join(results_dir, file)
                    # print(f"DEBUG: Updated latest PP file: {latest_filepath} ({latest_file_date})")
            elif file_type_prefix == 'PM' and "screening des résultats PM.xlsx" in file: # Changed to 'résultats' for PM
                if latest_file_date is None or file_date > latest_file_date:
                    latest_file_date = file_date
                    latest_filepath = os.path.join(results_dir, file)
                    # print(f"DEBUG: Updated latest PM file: {latest_filepath} ({latest_file_date})")

    if latest_filepath:
        try:
            st.info(f"DEBUG: Loading historical file: {os.path.basename(latest_filepath)}")
            history_df = pd.read_excel(latest_filepath, header=1) # Specify header=1 (0-indexed) to correctly read the columns
            # Normalize column names by stripping whitespace
            history_df.columns = history_df.columns.str.strip()
            # Map known problematic column names to their expected clean names
            # This handles potential inconsistencies in column names from older files
            column_name_mapping = {
                'entity_name(crm)': 'Entity Name (CRM)',
                'company/account': 'Company/Account',
                'screening_result': 'Screening Result',
                'commentaire': 'Commentaire',
                'status': 'Status',
                # Add common variations for First Name and Last Name
                'first_name': 'First Name',
                'prenom': 'First Name',
                'last_name': 'Last Name',
                'nom': 'Last Name',
            }
            # Apply the mapping to column names
            history_df = history_df.rename(columns=lambda col: column_name_mapping.get(col.lower(), col))

            # Explicitly convert key columns to string type to ensure consistent normalization
            for col in ["First Name", "Last Name", "Company/Account"]:
                if col in history_df.columns:
                    history_df[col] = history_df[col].astype(str).fillna("")
                else:
                    print(f"DEBUG_HISTORY: Missing expected column '{col}' in historical PP DataFrame after mapping. This may affect key generation.")

            print(f"DEBUG: Columns in loaded historical '{file_type_prefix}' DF: {history_df.columns.tolist()}")
        except Exception as e:
            st.warning(f"Problème lors du chargement du fichier historique '{file_type_prefix}' {os.path.basename(latest_filepath)} : {e}")

    return history_df

def create_history_map(history_df, is_pm):
    """
    Crée une carte (dictionnaire) des commentaires et statuts historiques.

    Args:
        history_df (pd.DataFrame): DataFrame des données historiques.
        is_pm (bool): True pour les personnes morales, False pour les personnes physiques.

    Returns:
        dict: Dictionnaire {unique_key: {'Commentaire': '...', 'Statut': '...'}}
    """
    history_map = {}
    if history_df is None or history_df.empty:
        return history_map

    # Ensure key columns and target columns exist
    if is_pm:
        key_cols = ["Entity Name (CRM)"]
    else:
        key_cols = ["Company/Account", "Screening Result"]
    
    required_cols = key_cols + ["Commentaire", "Status" ,"Company/Account"]

    # Check if all required columns are in the history_df (after stripping whitespace)
    # Convert history_df.columns to a set for efficient lookup
    history_df_cols_set = set(history_df.columns)
    missing_cols = [col for col in required_cols if col not in history_df_cols_set]
    if missing_cols:
        st.warning(f"Le DataFrame historique n'a pas toutes les colonnes requises: {', '.join(missing_cols)}. Le mappage historique sera incomplet.")
        # Proceed with available columns

    for _, row in history_df.iterrows():
        try:
            if is_pm:
                # Access column using .get() for safety and then strip/lower
                entity_name_crm = normalize_text_for_key(row.get("Entity Name (CRM)", ""))
                if entity_name_crm: # Ensure key is not empty
                    key = entity_name_crm
                else:
                    continue # Skip if key column is missing or empty
            else:
                # Apply normalization to PP key components
                first_name = normalize_text_for_key(row.get("First Name", ""))
                last_name = normalize_text_for_key(row.get("Last Name", ""))

                key = f"{first_name}|{last_name}"
                if not first_name or not last_name : # Check if key is empty after normalization
                    continue

            comment = row.get("Commentaire", "")
            status = row.get("Status", "")
            Company=row.get("Company/Account", "")

            # Ensure comment and status are strings, handling NaN/None
            comment = str(comment).strip() if pd.notna(comment) else ""
            status = str(status).strip() if pd.notna(status) else ""
            Company = str(Company).strip() if pd.notna(Company) else ""

            history_map[key] = {"Commentaire": comment, "Statut": status,"Company/Account":Company}
            print(f"DEBUG_HISTORY: Added to history_map: key='{key}', Comment='{comment}', Status='{status}'")
        except Exception as e:
            st.error(f"Erreur lors de la création de la carte historique pour la ligne : {e}")
            print(f"DEBUG_HISTORY: Error processing history row: {row.to_dict()}, Error: {e}")
            continue
            
    return history_map

def apply_history_to_current_screening(current_df, history_map, is_pm):
    """
    Applique les commentaires et statuts historiques au DataFrame de screening actuel.

    Args:
        current_df (pd.DataFrame): DataFrame des résultats de screening actuels.
        history_map (dict): Dictionnaire {unique_key: {'Commentaire': '...', 'Statut': '...'}}.
        is_pm (bool): True pour les personnes morales, False pour les personnes physiques.

    Returns:
        pd.DataFrame: DataFrame de screening mis à jour.
    """
    if current_df is None or current_df.empty or not history_map:
        return current_df

    # Ensure 'Commentaire' and 'Status' columns exist, initialize if not
    if "Commentaire" not in current_df.columns:
        current_df["Commentaire"] = ""
    if "Status" not in current_df.columns:
        current_df["Status"] = ""
    if "Company/Account" not in current_df.columns:
        current_df["Company/Account"] = ""
        
    # Normalize column names of the current DataFrame before applying history
    current_df.columns = current_df.columns.str.strip()

    for index, row in current_df.iterrows():
        try:
            if is_pm:
                # Access column using .get() for safety and then strip/lower
                entity_name_crm = normalize_text_for_key(row.get("Entity Name (CRM)", ""))
                if entity_name_crm: # Ensure key is not empty
                    key = entity_name_crm
                else:
                    print(f"DEBUG: Skipping current row (PM) due to empty Entity Name (CRM) after normalization: {row.to_dict()}")
                    continue
            else:
                # Apply normalization to PP key components
                first_name = normalize_text_for_key(row.get("First Name", ""))
                last_name = normalize_text_for_key(row.get("Last Name", ""))

                key = f"{first_name}|{last_name}"
                if not first_name or not last_name: # Check if key is empty after normalization
                    continue

            if key in history_map:
                # Use .get() for safety when accessing Commentaire and Statut from history_map entry
                current_df.at[index, "Commentaire"] = history_map[key].get("Commentaire", "")
                current_df.at[index, "Status"] = history_map[key].get("Statut", "")
                current_df.at[index, "Company/Account"] = history_map[key].get("Company/Account", "")
            else:
                if not is_pm: # Only show UI info for PP
                    st.info(f"")
        except Exception as e:
            st.error(f"Erreur lors de l'application de l'historique pour la ligne {index}: {e}")
            print(f"DEBUG: Error processing current row {index}: {row.to_dict()}, Error: {e}")
            continue

    # st.success(f"Historique appliqué avec succès au screening {'PM' if is_pm else 'PP'}.")
    return current_df

def merge_interactive_changes_back(main_df_key, interactive_df, is_pm):
    """
    Merges 'Commentaire' and 'Status' from the interactively updated DataFrame
    back into the main DataFrame stored in st.session_state.

    Args:
        main_df_key (str): The key in st.session_state where the main DataFrame is stored (e.g., 'screening_results', 'pm_matches').
        interactive_df (pd.DataFrame): The DataFrame that was interactively updated (e.g., 'matches' or 'pm_results_df' after display_matches_interactive).
        is_pm (bool): True for PM, False for PP.
    """
    if main_df_key not in st.session_state or st.session_state[main_df_key] is None or st.session_state[main_df_key].empty:
        print(f"DEBUG: Main DataFrame '{main_df_key}' not found or empty in session state. Skipping merge.")
        return

    main_df = st.session_state[main_df_key]

    # Ensure 'Commentaire' and 'Status' columns exist in main_df
    if "Commentaire" not in main_df.columns:
        main_df["Commentaire"] = ""
    if "Status" not in main_df.columns:
        main_df["Status"] = ""

    # Normalize column names in interactive_df if they haven't been already
    interactive_df.columns = interactive_df.columns.str.strip()

    # Create the unique key for merging
    if is_pm:
        key_cols = ["Entity Name (CRM)"]
    else:
        key_cols = ["Company/Account", "Screening Result"]
    
    # Ensure key columns exist in both DataFrames
    missing_keys_main = [col for col in key_cols if col not in main_df.columns]
    missing_keys_interactive = [col for col in key_cols if col not in interactive_df.columns]
    if missing_keys_main or missing_keys_interactive:
        print(f"DEBUG: Missing key columns in main_df ({missing_keys_main}) or interactive_df ({missing_keys_interactive}). Cannot merge interactive changes.")
        return

    main_df['__merge_key__'] = main_df.apply(
        lambda row: normalize_text_for_key(row.get(key_cols[0], "")) if is_pm
        else f"{normalize_text_for_key(row.get(key_cols[0], ''))}|{normalize_text_for_key(row.get(key_cols[1], ''))}",
        axis=1
    )
    interactive_df['__merge_key__'] = interactive_df.apply(
        lambda row: normalize_text_for_key(row.get(key_cols[0], "")) if is_pm
        else f"{normalize_text_for_key(row.get(key_cols[0], ''))}|{normalize_text_for_key(row.get(key_cols[1], ''))}",
        axis=1
    )

    # Create a mapping from interactive_df's keys to its comments/statuses
    interactive_changes_map = interactive_df.set_index('__merge_key__')[['Commentaire', 'Status']].to_dict('index')

    updated_count = 0
    for index, row in main_df.iterrows():
        key = row['__merge_key__']
        if key in interactive_changes_map:
            new_comment = interactive_changes_map[key].get('Commentaire', '')
            new_status = interactive_changes_map[key].get('Status', '')

            current_comment = main_df.at[index, "Commentaire"]
            current_status = main_df.at[index, "Status"]

            if current_comment != new_comment or current_status != new_status:
                main_df.at[index, "Commentaire"] = new_comment
                main_df.at[index, "Status"] = new_status
                updated_count += 1
                print(f"DEBUG: Updated {main_df_key} for key '{key}': Comment='{new_comment}', Status='{new_status}'")
            else:
                print(f"DEBUG: No change for {main_df_key} key '{key}'.")

    # Drop the temporary merge key
    main_df.drop(columns=['__merge_key__'], inplace=True, errors='ignore')
    interactive_df.drop(columns=['__merge_key__'], inplace=True, errors='ignore') # Also drop from the copy passed in

    print(f"DEBUG: Merged {updated_count} interactive changes back into '{main_df_key}'.")
    st.session_state[main_df_key] = main_df # Ensure session state is explicitly updated with the modified DF (even if it's already a reference)

def delete_current_day_screening_files(results_dir, current_date_yyyymmdd):
    files_deleted = []
    patterns = [
        f"{current_date_yyyymmdd} - screening des correspondances PP.xlsx",
        f"{current_date_yyyymmdd} - screening des résultats PP.xlsx",
        f"{current_date_yyyymmdd} - screening des valeurs aberrantes PP.xlsx",
        f"{current_date_yyyymmdd} - screening des correspondances PM.xlsx",
        f"{current_date_yyyymmdd} - screening des résultats PM.xlsx",
        f"{current_date_yyyymmdd} - screening toutes entités PM.xlsx", # Also delete this one if it's generated
    ]
    
    for pattern in patterns:
        filepath = os.path.join(results_dir, pattern)
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
                files_deleted.append(os.path.basename(filepath))
                print(f"DEBUG: Deleted existing file: {filepath}")
            except Exception as e:
                print(f"DEBUG: Error deleting file {filepath}: {e}")
    return files_deleted

def check_for_existing_screening_files(results_dir, current_date_yyyymmdd):
    files_found = []
    patterns = [
        f"{current_date_yyyymmdd} - screening des correspondances PP.xlsx",
        f"{current_date_yyyymmdd} - screening des résultats PP.xlsx",
        f"{current_date_yyyymmdd} - screening des valeurs aberrantes PP.xlsx",
        f"{current_date_yyyymmdd} - screening des correspondances PM.xlsx",
        f"{current_date_yyyymmdd} - screening des résultats PM.xlsx",
        f"{current_date_yyyymmdd} - screening toutes entités PM.xlsx",
    ]

    for pattern in patterns:
        filepath = os.path.join(results_dir, pattern)
        if os.path.exists(filepath):
            files_found.append(os.path.basename(filepath))
    return files_found

if __name__ == "__main__":
    pythoncom.CoInitialize()
    main()             

IOSCO_FILENAME = "IOSCO_List_of_Firms.csv"
IOSCO_URL = "https://www.iosco.org/investor_alert/?obj=investor_alert_data.csv"
IOSCO_FILEPATH = os.path.join(os.path.dirname(__file__), IOSCO_FILENAME)

def download_file(url, filepath):
    """Downloads a file from a URL to a specified filepath."""
    try:
        st.info(f"Attempting to download file from {url}")
        response = requests.get(url, stream=True)
        response.raise_for_status() # Raise an HTTPError for bad responses (4xx or 5xx)
        with open(filepath, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
        st.success(f"Downloaded {os.path.basename(filepath)}")
    except requests.exceptions.RequestException as e:
        st.error(f"Error downloading file from {url}: {e}")
        return False
    except Exception as e:
        st.error(f"An unexpected error occurred during download: {e}")
        return False
    return True

def parse_iosco_csv(filepath):
    """Parses the IOSCO CSV file and extracts/cleans company names."""
    if not os.path.exists(filepath):
        st.error(f"IOSCO file not found at {filepath}")
        return []

    try:
        st.info(f"Attempting to parse file at {filepath}")
        # Read the CSV, assuming it's comma-delimited and has a header
        # Columns are 0-indexed, so G, H, I are columns 6, 7, 8
        df = pd.read_csv(filepath, usecols=[6, 7, 8], header=0, encoding='utf-8')
        df.columns = ['commercial_name', 'other_commercial_names', 'corporate_names']

        # Combine all name columns into a single list, handling NaNs
        all_names = []
        for index, row in df.iterrows():
            names = []
            # Ensure data is string and handle NaNs before splitting/appending
            commercial = str(row['commercial_name']) if pd.notna(row['commercial_name']) else ''
            other = str(row['other_commercial_names']) if pd.notna(row['other_commercial_names']) else ''
            corporate = str(row['corporate_names']) if pd.notna(row['corporate_names']) else ''

            if commercial:
                names.append(commercial)
            if other:
                # Split aliases by comma or semicolon and strip spaces
                other_names = re.split(r'[,;]', other)
                names.extend([name.strip() for name in other_names if name.strip()])
            if corporate:
                # Split aliases by comma or semicolon and strip spaces
                corporate_names = re.split(r'[,;]', corporate)
                names.extend([name.strip() for name in corporate_names if name.strip()])

            all_names.extend(names)

        # Remove leading/trailing spaces and duplicates
        # Convert to lowercase for case-insensitive matching later
        cleaned_names = sorted(list(set([name.strip().lower() for name in all_names if name.strip()])))

        st.success(f"Parsed {len(cleaned_names)} unique company names from IOSCO file.")
        return cleaned_names

    except FileNotFoundError:
        st.error(f"Error parsing file: The file was not found at {filepath}")
        return []
    except pd.errors.EmptyDataError:
        st.error(f"Error parsing file: The file at {filepath} is empty.")
        return []
    except Exception as e:
        st.error(f"An unexpected error occurred during parsing: {e}")
        return []

def extract_crm_company_names(filepath, column_names):
    """Extracts and cleans company names from specified columns in a CRM file."""
    if not os.path.exists(filepath):
        st.error(f"CRM file not found at {filepath}")
        return []

    try:
        st.info(f"Attempting to read CRM file at {filepath}")
        # Assuming CSV format for CRM files
        df = pd.read_csv(filepath)

        all_names = []
        for col_name in column_names:
            if col_name in df.columns:
                # Extract names, handle NaNs, convert to string, strip spaces
                names = df[col_name].dropna().astype(str).str.strip().tolist()
                all_names.extend(names)
            else:
                st.warning(f"Column '{col_name}' not found in {os.path.basename(filepath)}")

        # Remove duplicates and convert to lowercase for case-insensitive matching
        cleaned_names = sorted(list(set([name.lower() for name in all_names if name])))

        st.success(f"Extracted {len(cleaned_names)} company names from CRM file {os.path.basename(filepath)}.")
        return cleaned_names

    except FileNotFoundError:
        st.error(f"Error reading file: The file was not found at {filepath}")
        return []
    except pd.errors.EmptyDataError:
        st.error(f"Error reading file: The file at {filepath} is empty.")
        return []
    except Exception as e:
        st.error(f"An unexpected error occurred during reading CRM file: {e}")
        return []

def perform_fuzzy_matching(crm_names, ioscio_names, threshold=80):
    """Performs fuzzy matching between CRM names and IOSCO names."""
    matches = []
    if not ioscio_names:
        st.warning("IOSCO names list is empty, skipping fuzzy matching.")
        return matches

    st.info(f"Performing fuzzy matching with threshold: {threshold}%")
    # Using process.extractOne for efficiency when looking for a single best match
    for crm_name in crm_names:
        # Ensure crm_name is a string before matching
        if pd.notna(crm_name) and isinstance(crm_name, str) and crm_name.strip():
            best_match = process.extractOne(crm_name.strip().lower(), ioscio_names, score_cutoff=threshold)
            if best_match:
                # best_match is a tuple: (matched_string, score)
                matches.append({
                    'CRM Name': crm_name.strip(),
                    'IOSCO Match': best_match[0],
                    'Score': best_match[1]
                })

    st.success(f"Found {len(matches)} potential matches.")
    return matches


import re

def find_companies_with_all_keywords(list_a, list_b):
    """
    Trouve les paires de noms (name_a, name_b) où name_b contient tous les mots
    de name_a (ignorants la casse et les espaces superflus).

    Args:
        list_a (list): Liste de noms d'entreprises sources.
        list_b (list): Liste de noms d'entreprises cibles.

    Returns:
        list: Une liste de tuples (nom_a, nom_b) pour les correspondances trouvées.
    """
    matches = []

    for name_a in list_a:
        if not isinstance(name_a, str) or not name_a.strip():
            continue # Ignorer les entrées vides ou non string dans la liste A

        # Normaliser et obtenir les mots de name_a
        # Utiliser re.findall pour gérer plusieurs espaces et signes de ponctuation basiques
        words_a = set(re.findall(r'\b\w+\b', name_a.lower()))

        if not words_a:
            continue # Si name_a n'a pas de mots, ignorer

        for name_b in list_b:
            if not isinstance(name_b, str) or not name_b.strip():
                continue # Ignorer les entrées vides ou non string dans la liste B

            # Normaliser et obtenir les mots de name_b
            words_b = set(re.findall(r'\b\w+\b', name_b.lower()))

            # Vérifier si tous les mots de name_a sont présents dans name_b
            # C'est équivalent à vérifier si words_a est un sous-ensemble de words_b
            if words_a.issubset(words_b):
                matches.append((name_a.strip(), name_b.strip()))

    return matches


